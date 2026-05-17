from datetime import datetime, timedelta
from io import BytesIO
import os
import re

import pandas as pd
from flask import Blueprint, Response, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def register_admin_legacy_routes(app, legacy):
    """Registro incremental de rutas legacy del area administrativa."""

    def admin_promo():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        if request.method == "POST":
            id_producto = pd.to_numeric(request.form.get("id_producto", ""), errors="coerce")
            nombre = request.form.get("nombre", "").strip()
            descripcion = request.form.get("descripcion", "").strip()
            tipo_descuento = request.form.get("tipo_descuento", "porcentaje").strip().lower()
            if tipo_descuento not in ["porcentaje", "valor_fijo"]:
                flash("Selecciona un tipo de descuento válido.", "warning")
                return redirect(url_for("admin_promo"))
            try:
                valor_descuento = float(request.form.get("valor_descuento", ""))
            except (TypeError, ValueError):
                flash("Ingresa un valor de descuento válido.", "warning")
                return redirect(url_for("admin_promo"))
            if valor_descuento <= 0:
                flash("El descuento debe ser mayor a cero. No se permiten valores negativos.", "warning")
                return redirect(url_for("admin_promo"))
            if tipo_descuento == "porcentaje" and valor_descuento > 100:
                flash("El porcentaje de descuento no puede superar el 100%.", "warning")
                return redirect(url_for("admin_promo"))

            codigo = request.form.get("codigo", "").strip().upper()
            fecha_inicio = request.form.get("fecha_inicio", "")
            fecha_fin = request.form.get("fecha_fin", "")
            activo = request.form.get("activo") == "on"
            inicio_date = legacy.parsear_fecha_promocion(fecha_inicio)
            fin_date = legacy.parsear_fecha_promocion(fecha_fin)
            if not inicio_date or not fin_date:
                flash("La fecha de inicio y la fecha de finalización son obligatorias.", "warning")
                return redirect(url_for("admin_promo"))
            if inicio_date > fin_date:
                flash("La fecha de inicio no puede ser mayor que la fecha de fin.", "warning")
                return redirect(url_for("admin_promo"))
            if pd.isna(id_producto):
                flash("Debes seleccionar un producto para la promoción.", "warning")
                return redirect(url_for("admin_promo"))
            productos_ref = legacy.cargar_productos_activos_df()
            id_producto_num = int(id_producto)
            existe_producto = not productos_ref[
                pd.to_numeric(productos_ref.get("id_producto", 0), errors="coerce") == id_producto_num
            ].empty
            if not existe_producto:
                flash("El producto seleccionado no existe o está inactivo.", "warning")
                return redirect(url_for("admin_promo"))
            producto_ref = productos_ref[
                pd.to_numeric(productos_ref.get("id_producto", 0), errors="coerce") == id_producto_num
            ].iloc[0]
            precio_producto_raw = pd.to_numeric(producto_ref.get("precio", 0), errors="coerce")
            precio_producto = float(precio_producto_raw) if pd.notna(precio_producto_raw) else 0.0
            if tipo_descuento == "valor_fijo" and valor_descuento >= precio_producto:
                flash("El descuento fijo debe ser inferior al precio establecido de la prenda.", "warning")
                return redirect(url_for("admin_promo"))

            promos = legacy.cargar_promociones_df()
            if codigo:
                if not re.fullmatch(r"[A-Z0-9_-]{3,30}", codigo):
                    flash("El código promocional debe tener entre 3 y 30 caracteres: letras, números, guion o guion bajo.", "warning")
                    return redirect(url_for("admin_promo"))
                existe_codigo = promos[promos["codigo"].astype(str).str.upper() == codigo]
                if not existe_codigo.empty:
                    flash("El código promocional ya existe. Usa otro.", "warning")
                    return redirect(url_for("admin_promo"))
            next_id = int(pd.to_numeric(promos["id_promo"], errors="coerce").max() + 1) if not promos.empty else 1
            if not nombre:
                nombre = f"Promoción {producto_ref.get('nombre', 'producto')}"
            nuevo = {
                "id_promo": next_id,
                "nombre": nombre,
                "descripcion": descripcion,
                "tipo_descuento": tipo_descuento,
                "valor_descuento": valor_descuento,
                "id_producto": int(id_producto),
                "codigo": codigo,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "activo": activo,
            }
            promos = pd.concat([promos, pd.DataFrame([nuevo])], ignore_index=True)
            legacy.guardar_promociones_df(promos)
            detalle_desc = (
                legacy.formatear_cop(valor_descuento) if tipo_descuento == "valor_fijo" else f"{valor_descuento:.2f}%"
            )
            legacy.registrar_actividad(
                f"Promoción creada: {nombre}\n"
                f"- producto ID: {int(id_producto)}\n"
                f"- descuento: {detalle_desc}\n"
                f"- código: {codigo or 'N/A'}"
            )
            return redirect(url_for("admin_promo"))

        promos = legacy.cargar_promociones_df()
        hoy = datetime.now().date()
        productos = legacy.cargar_productos_activos_df()
        productos["id_producto"] = pd.to_numeric(productos.get("id_producto", 0), errors="coerce").fillna(0).astype(int)
        lista_productos = productos.to_dict(orient="records")
        productos_por_fuerza = {fuerza: [] for fuerza in legacy.FUERZAS_OPCIONES}

        for producto in lista_productos:
            imagen_principal = str(producto.get("imagen_url", "")).strip()
            galeria = legacy.obtener_galeria_producto(producto.get("id_producto"), imagen_principal)
            producto["imagenes"] = galeria
            if galeria:
                producto["imagen_url"] = galeria[0]

            fuerza_producto = str(producto.get("fuerza", "")).strip().lower()
            for fuerza in legacy.FUERZAS_OPCIONES:
                if fuerza_producto == fuerza.lower():
                    productos_por_fuerza[fuerza].append(producto)
                    break

        pagos = legacy.cargar_pagos_df()
        if not pagos.empty:
            if "id_promo" not in pagos.columns:
                pagos["id_promo"] = pd.Series(dtype="float")
            if "monto_descuento" not in pagos.columns:
                pagos["monto_descuento"] = 0.0
            pagos["id_promo"] = pd.to_numeric(pagos["id_promo"], errors="coerce")
            pagos["monto_descuento"] = pd.to_numeric(pagos["monto_descuento"], errors="coerce").fillna(0.0)
            impacto = pagos.dropna(subset=["id_promo"]).groupby("id_promo", as_index=False).agg(
                usos=("id_pago", "count"),
                descuento_total=("monto_descuento", "sum"),
            )
        else:
            impacto = pd.DataFrame(columns=["id_promo", "usos", "descuento_total"])

        lista_promos = promos.to_dict(orient="records")
        impacto_map = {}
        for _, row in impacto.iterrows():
            impacto_map[int(row["id_promo"])] = {
                "usos": int(row["usos"]),
                "descuento_total": float(row["descuento_total"]),
            }
        for promo in lista_promos:
            promo["estado_vigencia"] = legacy.estado_vigencia_promocion(promo, hoy)
            promo["es_aplicable"] = legacy.promocion_esta_aplicable(promo, hoy)
            promo["id_producto"] = int(pd.to_numeric(promo.get("id_producto"), errors="coerce") or 0)
            key = int(pd.to_numeric(promo.get("id_promo"), errors="coerce") or 0)
            metrica = impacto_map.get(key, {"usos": 0, "descuento_total": 0.0})
            promo["usos"] = metrica["usos"]
            promo["descuento_total"] = metrica["descuento_total"]
        nombre_producto = {int(row["id_producto"]): str(row.get("nombre", "Producto")) for _, row in productos.iterrows()}
        for promo in lista_promos:
            promo["producto_nombre"] = nombre_producto.get(promo["id_producto"], "Producto no encontrado")
        return render_template(
            "Administrador/Promociones/adim_promo.html",
            promos=lista_promos,
            productos=lista_productos,
            productos_por_fuerza=productos_por_fuerza,
            fuerzas=legacy.FUERZAS_OPCIONES,
        )

    def admin_promo_toggle(id_promo):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        promos = legacy.cargar_promociones_df()
        idx = promos[promos["id_promo"] == id_promo].index
        if not idx.empty:
            promo_actual = promos.loc[idx[0]].to_dict()
            if legacy.estado_vigencia_promocion(promo_actual, datetime.now().date()) == "vencida":
                flash("La promoción está vencida y no se puede reactivar ni cambiar de estado.", "warning")
                return redirect(url_for("admin_promo"))
            promos.loc[idx, "activo"] = ~promos.loc[idx, "activo"]
            legacy.guardar_promociones_df(promos)
            legacy.registrar_actividad(
                f"Promoción {'activada' if promos.loc[idx, 'activo'].iloc[0] else 'desactivada'}: {promos.loc[idx, 'nombre'].iloc[0]}"
            )
        return redirect(url_for("admin_promo"))

    def obtener_datos_charts(periodo, fecha_desde_raw, fecha_hasta_raw):
        detalle = legacy.cargar_detalle_pedido_df()
        productos = legacy.cargar_productos_df()
        pedidos = legacy.cargar_pedidos_df()
        pagos = legacy.cargar_pagos_df()

        if "id_producto" not in detalle.columns:
            detalle["id_producto"] = pd.Series(dtype="int")
        if "cantidad" not in detalle.columns:
            detalle["cantidad"] = 0
        if "subtotal" not in detalle.columns:
            detalle["subtotal"] = 0
        if "id_pedido" not in detalle.columns:
            detalle["id_pedido"] = pd.Series(dtype="int")

        detalle["cantidad"] = pd.to_numeric(detalle["cantidad"], errors="coerce").fillna(0)
        detalle["subtotal"] = pd.to_numeric(detalle["subtotal"], errors="coerce").fillna(0)

        if "id_producto" not in productos.columns:
            productos["id_producto"] = pd.Series(dtype="int")
        if "nombre" not in productos.columns:
            productos["nombre"] = ""
        if "precio" not in productos.columns:
            productos["precio"] = 0
        productos["precio"] = pd.to_numeric(productos["precio"], errors="coerce").fillna(0)

        if "id_pedido" not in pedidos.columns:
            pedidos["id_pedido"] = pd.Series(dtype="int")
        if "fecha_pedido" not in pedidos.columns:
            pedidos["fecha_pedido"] = ""
        pedidos["fecha_pedido"] = pd.to_datetime(pedidos["fecha_pedido"], errors="coerce")

        if "monto" not in pagos.columns:
            pagos["monto"] = 0
        if "metodo_pago" not in pagos.columns:
            pagos["metodo_pago"] = ""
        pagos["monto"] = pd.to_numeric(pagos["monto"], errors="coerce").fillna(0)
        pagos["metodo_pago"] = pagos["metodo_pago"].fillna("").astype(str)

        def filtrar_pedidos_por_rango(df_pedidos, desde=None, hasta=None):
            filtrado = df_pedidos.copy().dropna(subset=["fecha_pedido"])
            if desde is not None:
                filtrado = filtrado[filtrado["fecha_pedido"].dt.date >= desde]
            if hasta is not None:
                filtrado = filtrado[filtrado["fecha_pedido"].dt.date <= hasta]
            return filtrado

        def calcular_kpis_desde_pedidos(pedidos_base):
            ids = pedidos_base["id_pedido"].tolist()
            detalle_base = detalle[detalle["id_pedido"].isin(ids)].copy()
            pagos_base = pagos[pagos["id_pedido"].isin(ids)].copy()
            total_ventas_base = float(pagos_base["monto"].sum()) if not pagos_base.empty else float(detalle_base["subtotal"].sum())
            total_pedidos_base = int(len(pedidos_base))
            total_items_base = int(detalle_base["cantidad"].sum()) if not detalle_base.empty else 0
            ticket_promedio_base = (total_ventas_base / total_pedidos_base) if total_pedidos_base > 0 else 0
            return {
                "total_ventas": total_ventas_base,
                "total_pedidos": total_pedidos_base,
                "ticket_promedio": ticket_promedio_base,
                "total_items": total_items_base,
            }

        def variacion_kpi(valor_actual, valor_anterior):
            if valor_anterior == 0:
                if valor_actual == 0:
                    return {"trend": "flat", "texto": "0.0% vs periodo anterior"}
                return {"trend": "up", "texto": "Nuevo vs periodo anterior"}

            delta_pct = ((valor_actual - valor_anterior) / abs(valor_anterior)) * 100
            trend = "up" if delta_pct > 0 else "down" if delta_pct < 0 else "flat"
            return {"trend": trend, "texto": f"{delta_pct:+.1f}% vs periodo anterior"}

        hoy = datetime.now().date()
        fecha_desde = None
        fecha_hasta = None

        if periodo == "today":
            fecha_desde = hoy
            fecha_hasta = hoy
        elif periodo == "week":
            fecha_desde = hoy - timedelta(days=6)
            fecha_hasta = hoy
        elif periodo == "month":
            fecha_desde = hoy.replace(day=1)
            fecha_hasta = hoy
        elif periodo == "range":
            if fecha_desde_raw:
                try:
                    fecha_desde = datetime.strptime(fecha_desde_raw, "%Y-%m-%d").date()
                except ValueError:
                    fecha_desde = None
            if fecha_hasta_raw:
                try:
                    fecha_hasta = datetime.strptime(fecha_hasta_raw, "%Y-%m-%d").date()
                except ValueError:
                    fecha_hasta = None
            if fecha_desde is not None and fecha_hasta is not None and fecha_desde > fecha_hasta:
                fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

        pedidos_filtrados = filtrar_pedidos_por_rango(pedidos, fecha_desde, fecha_hasta)
        ids_pedidos = pedidos_filtrados["id_pedido"].tolist()
        detalle_filtrado = detalle[detalle["id_pedido"].isin(ids_pedidos)].copy()
        pagos_filtrados = pagos[pagos["id_pedido"].isin(ids_pedidos)].copy()

        ventas_por_producto_df = detalle_filtrado.groupby("id_producto", as_index=False)["cantidad"].sum()
        ventas_por_producto_df = pd.merge(
            ventas_por_producto_df,
            productos[["id_producto", "nombre", "precio"]],
            on="id_producto",
            how="left",
        )
        ventas_por_producto_df["nombre"] = ventas_por_producto_df["nombre"].fillna("Producto sin nombre")
        ventas_por_producto_df["precio"] = pd.to_numeric(ventas_por_producto_df["precio"], errors="coerce").fillna(0)

        pedidos_fechas = pedidos_filtrados.copy()
        pedidos_fechas["mes"] = pedidos_fechas["fecha_pedido"].dt.to_period("M").astype(str)

        ventas_por_mes_df = detalle_filtrado.groupby("id_pedido", as_index=False)["subtotal"].sum()
        ventas_por_mes_df = pd.merge(ventas_por_mes_df, pedidos_fechas[["id_pedido", "mes"]], on="id_pedido", how="left")
        ventas_por_mes_df = ventas_por_mes_df.dropna(subset=["mes"])
        ventas_por_mes_df = ventas_por_mes_df.groupby("mes", as_index=False)["subtotal"].sum().sort_values("mes")

        metodos_pago_df = pagos_filtrados.groupby("metodo_pago", as_index=False)["monto"].sum()
        metodos_pago_df = metodos_pago_df[metodos_pago_df["metodo_pago"].str.strip() != ""]

        kpis = calcular_kpis_desde_pedidos(pedidos_filtrados)

        top_productos_df = ventas_por_producto_df.sort_values("cantidad", ascending=False).head(5).copy()
        if "precio" not in top_productos_df.columns:
            top_productos_df = pd.merge(
                top_productos_df,
                productos[["id_producto", "precio"]],
                on="id_producto",
                how="left",
            )
        top_productos_df["precio"] = pd.to_numeric(top_productos_df.get("precio", 0), errors="coerce").fillna(0)

        kpis_anterior = {"total_ventas": 0, "total_pedidos": 0, "ticket_promedio": 0, "total_items": 0}
        comparacion_texto = "Sin comparacion de periodo"
        if fecha_desde is not None and fecha_hasta is not None:
            dias_periodo = max(1, (fecha_hasta - fecha_desde).days + 1)
            anterior_hasta = fecha_desde - timedelta(days=1)
            anterior_desde = anterior_hasta - timedelta(days=dias_periodo - 1)
            pedidos_anterior = filtrar_pedidos_por_rango(pedidos, anterior_desde, anterior_hasta)
            kpis_anterior = calcular_kpis_desde_pedidos(pedidos_anterior)
            comparacion_texto = (
                f"Comparando contra {anterior_desde.strftime('%Y-%m-%d')} a " f"{anterior_hasta.strftime('%Y-%m-%d')}"
            )
        elif periodo == "all":
            comparacion_texto = "Selecciona un periodo para activar comparacion"

        kpi_variaciones = {
            "total_ventas": variacion_kpi(kpis["total_ventas"], kpis_anterior["total_ventas"]),
            "total_pedidos": variacion_kpi(kpis["total_pedidos"], kpis_anterior["total_pedidos"]),
            "ticket_promedio": variacion_kpi(kpis["ticket_promedio"], kpis_anterior["ticket_promedio"]),
            "total_items": variacion_kpi(kpis["total_items"], kpis_anterior["total_items"]),
        }

        rango_texto = "Todo el historial"
        if fecha_desde is not None and fecha_hasta is not None:
            rango_texto = f"{fecha_desde.strftime('%Y-%m-%d')} a {fecha_hasta.strftime('%Y-%m-%d')}"
        elif fecha_desde is not None:
            rango_texto = f"Desde {fecha_desde.strftime('%Y-%m-%d')}"
        elif fecha_hasta is not None:
            rango_texto = f"Hasta {fecha_hasta.strftime('%Y-%m-%d')}"

        filtros = {
            "periodo": periodo if periodo in {"all", "today", "week", "month", "range"} else "all",
            "fecha_desde": fecha_desde_raw,
            "fecha_hasta": fecha_hasta_raw,
        }

        return {
            "ventas_producto": ventas_por_producto_df.to_dict(orient="records"),
            "ventas_mes": ventas_por_mes_df.to_dict(orient="records"),
            "metodos_pago": metodos_pago_df.to_dict(orient="records"),
            "top_productos": top_productos_df.to_dict(orient="records"),
            "kpis": kpis,
            "kpi_variaciones": kpi_variaciones,
            "filtros": filtros,
            "rango_texto": rango_texto,
            "comparacion_texto": comparacion_texto,
            "ventas_producto_df": ventas_por_producto_df,
            "ventas_mes_df": ventas_por_mes_df,
            "metodos_pago_df": metodos_pago_df,
            "top_productos_df": top_productos_df,
        }

    def admin_charts():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        periodo = request.args.get("periodo", "all").strip().lower()
        fecha_desde_raw = request.args.get("fecha_desde", "").strip()
        fecha_hasta_raw = request.args.get("fecha_hasta", "").strip()
        datos = obtener_datos_charts(periodo, fecha_desde_raw, fecha_hasta_raw)

        return render_template(
            "Administrador/Informes/admin_charts_dashboard.html",
            ventas_producto=datos["ventas_producto"],
            ventas_mes=datos["ventas_mes"],
            metodos_pago=datos["metodos_pago"],
            top_productos=datos["top_productos"],
            kpis=datos["kpis"],
            kpi_variaciones=datos["kpi_variaciones"],
            filtros=datos["filtros"],
            rango_texto=datos["rango_texto"],
            comparacion_texto=datos["comparacion_texto"],
        )

    def admin_charts_export_excel():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        periodo = request.args.get("periodo", "all").strip().lower()
        fecha_desde_raw = request.args.get("fecha_desde", "").strip()
        fecha_hasta_raw = request.args.get("fecha_hasta", "").strip()
        datos = obtener_datos_charts(periodo, fecha_desde_raw, fecha_hasta_raw)

        resumen_df = pd.DataFrame(
            [
                {"metrica": "Ventas totales", "valor": legacy.formatear_cop(datos["kpis"]["total_ventas"])},
                {"metrica": "Pedidos", "valor": datos["kpis"]["total_pedidos"]},
                {"metrica": "Ticket promedio", "valor": legacy.formatear_cop(datos["kpis"]["ticket_promedio"])},
                {"metrica": "Items vendidos", "valor": datos["kpis"]["total_items"]},
                {"metrica": "Rango aplicado", "valor": datos["rango_texto"]},
                {"metrica": "Comparacion", "valor": datos["comparacion_texto"]},
            ]
        )

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            resumen_df.to_excel(writer, sheet_name="Resumen", index=False)
            datos["ventas_producto_df"].to_excel(writer, sheet_name="Ventas por producto", index=False)
            datos["ventas_mes_df"].to_excel(writer, sheet_name="Ventas por mes", index=False)
            datos["metodos_pago_df"].to_excel(writer, sheet_name="Metodos de pago", index=False)
            datos["top_productos_df"].to_excel(writer, sheet_name="Top productos", index=False)

            wb = writer.book
            ws_producto = wb["Ventas por producto"]
            ws_mes = wb["Ventas por mes"]
            ws_metodos = wb["Metodos de pago"]
            ws_top = wb["Top productos"]
            formato_cop = '[>=1000]#,##0.00 "COP";0.00 "COP"'

            for row in range(2, ws_mes.max_row + 1):
                ws_mes.cell(row=row, column=2).number_format = formato_cop
            for row in range(2, ws_metodos.max_row + 1):
                ws_metodos.cell(row=row, column=2).number_format = formato_cop
            for row in range(2, ws_top.max_row + 1):
                ws_top.cell(row=row, column=4).number_format = formato_cop

            if ws_producto.max_row > 1:
                chart_prod = BarChart()
                chart_prod.title = "Cantidad vendida por producto"
                chart_prod.y_axis.title = "Cantidad"
                chart_prod.x_axis.title = "Producto"
                data_prod = Reference(ws_producto, min_col=3, min_row=1, max_row=ws_producto.max_row)
                cats_prod = Reference(ws_producto, min_col=2, min_row=2, max_row=ws_producto.max_row)
                chart_prod.add_data(data_prod, titles_from_data=True)
                chart_prod.set_categories(cats_prod)
                chart_prod.height = 8
                chart_prod.width = 14
                if chart_prod.series:
                    chart_prod.series[0].graphicalProperties.solidFill = "1882A9"
                ws_producto.add_chart(chart_prod, "E2")

            if ws_mes.max_row > 1:
                chart_mes = LineChart()
                chart_mes.title = "Ventas por mes"
                chart_mes.y_axis.title = "Subtotal"
                chart_mes.x_axis.title = "Mes"
                data_mes = Reference(ws_mes, min_col=2, min_row=1, max_row=ws_mes.max_row)
                cats_mes = Reference(ws_mes, min_col=1, min_row=2, max_row=ws_mes.max_row)
                chart_mes.add_data(data_mes, titles_from_data=True)
                chart_mes.set_categories(cats_mes)
                chart_mes.height = 8
                chart_mes.width = 14
                if chart_mes.series:
                    chart_mes.series[0].graphicalProperties.line.solidFill = "0A2962"
                ws_mes.add_chart(chart_mes, "D2")

            if ws_metodos.max_row > 1:
                chart_met = PieChart()
                chart_met.title = "Distribucion por metodo de pago"
                data_met = Reference(ws_metodos, min_col=2, min_row=1, max_row=ws_metodos.max_row)
                cats_met = Reference(ws_metodos, min_col=1, min_row=2, max_row=ws_metodos.max_row)
                chart_met.add_data(data_met, titles_from_data=True)
                chart_met.set_categories(cats_met)
                chart_met.height = 8
                chart_met.width = 12
                if chart_met.series:
                    colores = ["0A2962", "1882A9", "1ABC9C", "F1C40F", "E74C3C", "6C5CE7"]
                    points = []
                    num_points = max(0, ws_metodos.max_row - 1)
                    for i in range(num_points):
                        point = DataPoint(idx=i)
                        point.graphicalProperties.solidFill = colores[i % len(colores)]
                        points.append(point)
                    chart_met.series[0].dPt = points
                ws_metodos.add_chart(chart_met, "D2")

            if ws_top.max_row > 1:
                chart_top = BarChart()
                chart_top.title = "Top productos vendidos"
                chart_top.y_axis.title = "Cantidad"
                chart_top.x_axis.title = "Producto"
                data_top = Reference(ws_top, min_col=3, min_row=1, max_row=ws_top.max_row)
                cats_top = Reference(ws_top, min_col=2, min_row=2, max_row=ws_top.max_row)
                chart_top.add_data(data_top, titles_from_data=True)
                chart_top.set_categories(cats_top)
                chart_top.height = 8
                chart_top.width = 14
                if chart_top.series:
                    chart_top.series[0].graphicalProperties.solidFill = "0A2962"
                ws_top.add_chart(chart_top, "F2")
        buffer.seek(0)

        nombre_archivo = f"graficas_informe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def agregar_producto():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        productos = legacy.cargar_productos_df()
        nuevo_id = legacy.next_id("producto", "id_producto")
        fuerza = request.form.get("fuerza", "").strip()
        intendencia = request.form.get("intendencia", "").strip()
        if fuerza not in legacy.FUERZAS_OPCIONES or intendencia not in legacy.INTENDENCIAS_OPCIONES:
            flash("Selecciona una fuerza e intendencia validas.", "danger")
            return redirect(url_for("admin_productos"))

        imagenes = [a for a in request.files.getlist("imagenes") if a and str(getattr(a, "filename", "")).strip()]
        if not imagenes:
            imagen_unica = request.files.get("imagen")
            if imagen_unica and str(getattr(imagen_unica, "filename", "")).strip():
                imagenes = [imagen_unica]

        if len(imagenes) > legacy.MAX_IMAGES_PER_PRODUCT:
            flash(f"Solo puedes subir hasta {legacy.MAX_IMAGES_PER_PRODUCT} imagenes por producto.", "danger")
            return redirect(url_for("admin_productos"))

        for imagen in imagenes:
            error_validacion = legacy.validar_archivo_imagen(imagen)
            if error_validacion:
                flash(error_validacion, "danger")
                return redirect(url_for("admin_productos"))

        galeria_guardada = legacy.guardar_galeria_producto(
            nuevo_id,
            imagenes,
            reemplazar=True,
            fuerza=fuerza,
            intendencia=intendencia,
        )
        imagen_url = galeria_guardada[0] if galeria_guardada else ""

        nuevo_producto = {
            "id_producto": nuevo_id,
            "nombre": request.form["nombre"],
            "descripcion": request.form["descripcion"],
            "precio": float(request.form["precio"]),
            "stock": int(request.form["stock"]),
            "id_categoria": 1,
            "fuerza": fuerza,
            "intendencia": intendencia,
            "imagen_url": imagen_url,
            "eliminado": False,
        }

        productos = pd.concat([productos, pd.DataFrame([nuevo_producto])], ignore_index=True)
        legacy.guardar_productos_df(productos)
        legacy.registrar_actividad(
            f"Creo producto '{request.form['nombre']}' (ID {nuevo_id})\n"
            f"- precio: {legacy.formatear_cop(float(request.form['precio']))}\n"
            f"- stock: {int(request.form['stock'])}\n"
            f"- fuerza: {fuerza}\n"
            f"- intendencia: {intendencia}\n"
            f"- imagenes: {len(galeria_guardada)}"
        )
        return redirect(url_for("admin_productos"))

    def export_sales():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        pedidos = legacy.cargar_pedidos_df()
        pagos = legacy.cargar_pagos_df()
        detalle = legacy.cargar_detalle_pedido_df()

        informe = pd.merge(pedidos, pagos, on="id_pedido", how="left")
        totales_pedido = detalle.groupby("id_pedido")["subtotal"].sum().reset_index()
        totales_pedido.columns = ["id_pedido", "total_productos"]
        informe = pd.merge(informe, totales_pedido, on="id_pedido", how="left")
        csv_data = informe.to_csv(index=False)

        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"},
        )

    def subir_imagen(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        imagenes = [a for a in request.files.getlist("imagenes") if a and str(getattr(a, "filename", "")).strip()]
        if not imagenes:
            imagen_unica = request.files.get("imagen")
            if imagen_unica and str(getattr(imagen_unica, "filename", "")).strip():
                imagenes = [imagen_unica]
        if not imagenes:
            flash("No se selecciono ninguna imagen.")
            return redirect(url_for("admin_productos"))

        if len(imagenes) > legacy.MAX_IMAGES_PER_PRODUCT:
            flash(f"Solo puedes subir hasta {legacy.MAX_IMAGES_PER_PRODUCT} imagenes por producto.", "danger")
            return redirect(url_for("admin_productos"))

        for imagen in imagenes:
            error_validacion = legacy.validar_archivo_imagen(imagen)
            if error_validacion:
                flash(error_validacion, "danger")
                return redirect(url_for("admin_productos"))

        productos = legacy.cargar_productos_df()
        idx = productos[productos["id_producto"] == id_producto].index
        if not idx.empty:
            fuerza = str(productos.at[idx[0], "fuerza"]) if "fuerza" in productos.columns else ""
            intendencia = str(productos.at[idx[0], "intendencia"]) if "intendencia" in productos.columns else ""
            galeria_guardada = legacy.guardar_galeria_producto(
                id_producto,
                imagenes,
                reemplazar=True,
                fuerza=fuerza,
                intendencia=intendencia,
            )
            productos.at[idx[0], "imagen_url"] = galeria_guardada[0] if galeria_guardada else ""
            legacy.guardar_productos_df(productos)
            flash(f"Galeria reemplazada ({len(galeria_guardada)} imagenes).", "success")
        else:
            flash("Producto no encontrado.", "danger")
        return redirect(url_for("admin_productos"))

    def agregar_imagenes_producto(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        imagenes = [a for a in request.files.getlist("imagenes_agregar") if a and str(getattr(a, "filename", "")).strip()]
        if not imagenes:
            imagenes = [a for a in request.files.getlist("imagenes") if a and str(getattr(a, "filename", "")).strip()]
        if not imagenes:
            flash("No se seleccionaron imagenes para agregar.", "warning")
            return redirect(url_for("admin_productos"))

        for imagen in imagenes:
            error_validacion = legacy.validar_archivo_imagen(imagen)
            if error_validacion:
                flash(error_validacion, "danger")
                return redirect(url_for("admin_productos"))

        productos = legacy.cargar_productos_df()
        idx = productos[productos["id_producto"] == id_producto].index
        if idx.empty:
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("admin_productos"))

        imagen_principal = str(productos.at[idx[0], "imagen_url"]) if "imagen_url" in productos.columns else ""
        galeria_actual = legacy.obtener_galeria_producto(id_producto, imagen_principal)
        if len(galeria_actual) + len(imagenes) > legacy.MAX_IMAGES_PER_PRODUCT:
            disponibles = max(0, legacy.MAX_IMAGES_PER_PRODUCT - len(galeria_actual))
            flash(
                f"Este producto ya tiene {len(galeria_actual)} imagen(es). "
                f"Solo puedes agregar {disponibles} mas (maximo {legacy.MAX_IMAGES_PER_PRODUCT}).",
                "danger",
            )
            return redirect(url_for("admin_productos"))

        fuerza = str(productos.at[idx[0], "fuerza"]) if "fuerza" in productos.columns else ""
        intendencia = str(productos.at[idx[0], "intendencia"]) if "intendencia" in productos.columns else ""
        legacy.guardar_galeria_producto(
            id_producto,
            imagenes,
            reemplazar=False,
            fuerza=fuerza,
            intendencia=intendencia,
        )
        galeria_actualizada = legacy.obtener_galeria_producto(id_producto, "")
        productos.at[idx[0], "imagen_url"] = galeria_actualizada[0] if galeria_actualizada else ""
        legacy.guardar_productos_df(productos)
        flash(f"Se agregaron {len(imagenes)} imagenes. Total: {len(galeria_actualizada)}.", "success")
        return redirect(url_for("admin_productos"))

    def eliminar_imagen_producto(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        imagen_a_eliminar = legacy.normalizar_imagen_url(request.form.get("imagen_a_eliminar", "").strip())
        if not imagen_a_eliminar:
            flash("Selecciona una imagen para eliminar.", "warning")
            return redirect(url_for("admin_productos"))

        productos = legacy.cargar_productos_df()
        idx = productos[productos["id_producto"] == id_producto].index
        if idx.empty:
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("admin_productos"))

        imagen_principal = str(productos.at[idx[0], "imagen_url"]) if "imagen_url" in productos.columns else ""
        galeria_actual = legacy.obtener_galeria_producto(id_producto, imagen_principal)
        if imagen_a_eliminar not in galeria_actual:
            flash("La imagen seleccionada no pertenece a este producto.", "danger")
            return redirect(url_for("admin_productos"))

        ruta_absoluta = legacy.ruta_imagen_producto_absoluta(imagen_a_eliminar)
        if ruta_absoluta and os.path.exists(ruta_absoluta):
            os.remove(ruta_absoluta)

        galeria_actualizada = legacy.obtener_galeria_producto(id_producto, "")
        productos.at[idx[0], "imagen_url"] = galeria_actualizada[0] if galeria_actualizada else ""
        legacy.guardar_productos_df(productos)
        flash(f"Imagen eliminada. Total actual: {len(galeria_actualizada)}.", "success")
        return redirect(url_for("admin_productos"))

    def eliminar_producto(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        productos = legacy.cargar_productos_df()
        idx = productos[productos["id_producto"] == id_producto].index
        if not idx.empty:
            productos.at[idx[0], "eliminado"] = True
            legacy.guardar_productos_df(productos)
            nombre = productos.at[idx[0], "nombre"]
            legacy.registrar_actividad(f"Elimino producto: {nombre} (ID {id_producto})")
        return redirect(url_for("admin_productos"))

    def eliminar_definitivo(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        productos = legacy.cargar_productos_df()
        idx = productos[productos["id_producto"] == id_producto].index
        if not idx.empty:
            nombre = productos.at[idx[0], "nombre"]
            productos = productos.drop(index=idx)
            legacy.guardar_productos_df(productos)
            legacy.registrar_actividad(f"Elimino definitivamente el producto: {nombre} (ID {id_producto})")
        return redirect(url_for("admin_papelera"))

    def restaurar_producto(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        productos = legacy.cargar_productos_df()
        idx = productos[productos["id_producto"] == id_producto].index
        if not idx.empty:
            productos.at[idx[0], "eliminado"] = False
            legacy.guardar_productos_df(productos)
            nombre = productos.at[idx[0], "nombre"]
            legacy.registrar_actividad(f"Restauro producto: {nombre} (ID {id_producto})")
        return redirect(url_for("admin_papelera"))

    def admin_papelera():
        if session.get("rol") != "admin":
            return "Acceso denegado"
        productos = legacy.cargar_productos_df()
        eliminados = productos[productos["eliminado"] == True].to_dict(orient="records")
        return render_template("Administrador/Papelera/admin_papelera.html", productos=eliminados)

    def admin_usuarios():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        usuarios = legacy.cargar_usuarios_df()
        pedidos = legacy.cargar_pedidos_df()
        usuarios["id_usuario"] = pd.to_numeric(usuarios["id_usuario"], errors="coerce")
        usuarios["telefono"] = usuarios["telefono"].fillna("").astype(str)
        usuarios["direccion"] = usuarios["direccion"].fillna("").astype(str)
        usuarios["email_verified"] = usuarios["email_verified"].fillna(False).astype(bool)

        buscar = request.args.get("buscar", "").strip()
        rol = request.args.get("rol", "").strip()
        estado = request.args.get("estado", "").strip().lower()
        orden = request.args.get("orden", "reciente").strip()
        edit_id = request.args.get("edit_id", type=int)
        pagina_actual = legacy._parse_positive_int(request.args.get("page", 1), default=1)

        filtrados = usuarios.copy()
        if buscar:
            mask = filtrados["nombre"].astype(str).str.contains(buscar, case=False, na=False) | filtrados["email"].astype(str).str.contains(
                buscar, case=False, na=False
            )
            filtrados = filtrados[mask]

        if rol:
            filtrados = filtrados[filtrados["rol"].astype(str).str.lower() == rol.lower()]
        if estado in ["activo", "inactivo"]:
            filtrados = filtrados[filtrados["estado"].astype(str).str.lower() == estado]

        if orden == "antiguo":
            filtrados = filtrados.sort_values(by="id_usuario", ascending=True, na_position="last")
        elif orden == "nombre":
            filtrados = filtrados.sort_values(by="nombre", ascending=True, na_position="last")
        else:
            filtrados = filtrados.sort_values(by="id_usuario", ascending=False, na_position="last")

        usuario_editar = None
        if edit_id is not None:
            candidato = usuarios[usuarios["id_usuario"] == edit_id]
            if not candidato.empty:
                usuario_editar = candidato.iloc[0].to_dict()
                if pd.notna(usuario_editar.get("id_usuario")):
                    usuario_editar["id_usuario"] = int(usuario_editar["id_usuario"])
                usuario_editar["estado"] = str(usuario_editar.get("estado", "activo")).strip().lower()
                usuario_editar["telefono"] = str(usuario_editar.get("telefono", "") or "").strip()
                usuario_editar["direccion"] = str(usuario_editar.get("direccion", "") or "").strip()
                usuario_editar["email_verified"] = bool(usuario_editar.get("email_verified", False))

        pedidos_por_usuario = {}
        if not pedidos.empty and "id_usuario" in pedidos.columns:
            pedidos_tmp = pedidos.copy()
            pedidos_tmp["id_usuario"] = pedidos_tmp["id_usuario"].fillna("").astype(str).str.strip().str.lower()
            for _, usuario_row in usuarios.iterrows():
                usuario_id = str(int(usuario_row["id_usuario"])) if pd.notna(usuario_row["id_usuario"]) else ""
                usuario_email = legacy.normalizar_email(usuario_row.get("email", ""))
                conteo = pedidos_tmp[pedidos_tmp["id_usuario"].isin({usuario_id.lower(), usuario_email})].shape[0]
                pedidos_por_usuario[usuario_email] = int(conteo)

        lista_usuarios = filtrados.fillna("").to_dict(orient="records")
        for usuario in lista_usuarios:
            if usuario.get("id_usuario") != "":
                usuario["id_usuario"] = int(usuario["id_usuario"])
            usuario["estado"] = str(usuario.get("estado", "activo")).strip().lower()
            usuario["telefono"] = str(usuario.get("telefono", "") or "").strip()
            usuario["direccion"] = str(usuario.get("direccion", "") or "").strip()
            usuario["email_verified"] = bool(usuario.get("email_verified", False))
            usuario["pedidos_total"] = pedidos_por_usuario.get(legacy.normalizar_email(usuario.get("email", "")), 0)

        usuarios_vista, paginacion_usuarios = legacy._paginar_lista(lista_usuarios, pagina_actual, per_page=10)

        filtros = {"buscar": buscar, "rol": rol, "estado": estado, "orden": orden, "page": paginacion_usuarios["page"]}
        resumen = {
            "total": int(len(usuarios)),
            "admins": int((usuarios["rol"].astype(str).str.lower() == "admin").sum()),
            "clientes": int((usuarios["rol"].astype(str).str.lower() == "normal").sum()),
            "activos": int((usuarios["estado"].astype(str).str.lower() == "activo").sum()),
            "verificados": int(usuarios["email_verified"].sum()),
        }
        return render_template(
            "Administrador/Gestion usuarios/admin_usuario.html",
            usuarios=usuarios_vista,
            usuario_editar=usuario_editar,
            filtros=filtros,
            resumen=resumen,
            paginacion_usuarios=paginacion_usuarios,
        )

    def admin_guardar_usuario():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        buscar = request.form.get("buscar", "").strip()
        rol_filtro = request.form.get("rol_filtro", "").strip()
        estado_filtro = request.form.get("estado_filtro", "").strip().lower()
        orden = request.form.get("orden", "reciente").strip()
        pagina_actual = legacy._parse_positive_int(request.form.get("page", 1), default=1)
        id_usuario_raw = request.form.get("id_usuario", "").strip()
        nombre = request.form.get("nombre", "").strip()
        email = legacy.normalizar_email(request.form.get("email", ""))
        password = request.form.get("password", "").strip()
        telefono = re.sub(r"\D", "", request.form.get("telefono", ""))
        direccion = re.sub(r"\s+", " ", request.form.get("direccion", "")).strip()
        rol = request.form.get("rol", "normal").strip().lower()
        estado = request.form.get("estado", "activo").strip().lower()

        retorno_kwargs = {
            "buscar": buscar,
            "rol": rol_filtro,
            "estado": estado_filtro,
            "orden": orden,
            "page": pagina_actual,
        }

        if rol not in ["admin", "normal"]:
            rol = "normal"
        if estado not in ["activo", "inactivo"]:
            estado = "activo"

        if not nombre or not email:
            flash("Nombre y email son obligatorios.", "danger")
            return redirect(url_for("admin_usuarios", **retorno_kwargs))

        if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
            flash("Debes ingresar un correo electronico valido.", "danger")
            return redirect(url_for("admin_usuarios", **retorno_kwargs))

        if telefono and len(telefono) != 10:
            flash("El telefono debe contener solo numeros y tener exactamente 10 digitos.", "danger")
            return redirect(url_for("admin_usuarios", **retorno_kwargs))

        usuarios = legacy.cargar_usuarios_df()
        usuarios["id_usuario"] = pd.to_numeric(usuarios["id_usuario"], errors="coerce")
        usuarios["email"] = usuarios["email"].astype(str)

        edit_id = None
        if id_usuario_raw:
            try:
                edit_id = int(float(id_usuario_raw))
            except ValueError:
                flash("ID de usuario invalido.", "danger")
                return redirect(url_for("admin_usuarios", **retorno_kwargs))

        if edit_id is None and not password:
            flash("La contrasena es obligatoria para crear usuarios.", "danger")
            return redirect(url_for("admin_usuarios", **retorno_kwargs))

        if password and not legacy.password_cumple_estandares(password):
            flash(
                "La contrasena debe tener minimo 8 caracteres, mayuscula, minuscula, numero y caracter especial.",
                "danger",
            )
            return redirect(url_for("admin_usuarios", **retorno_kwargs))

        existe_email = usuarios[usuarios["email"].str.lower() == email.lower()]
        if edit_id is not None:
            existe_email = existe_email[existe_email["id_usuario"] != edit_id]
        if not existe_email.empty:
            flash("Ese email ya esta registrado por otro usuario.", "danger")
            return redirect(url_for("admin_usuarios", **retorno_kwargs))

        if edit_id is not None:
            idx = usuarios[usuarios["id_usuario"] == edit_id].index
            if idx.empty:
                flash("Usuario no encontrado para edicion.", "danger")
                return redirect(url_for("admin_usuarios", **retorno_kwargs))

            usuarios.at[idx[0], "nombre"] = nombre
            usuarios.at[idx[0], "email"] = email
            usuarios.at[idx[0], "rol"] = rol
            usuarios.at[idx[0], "estado"] = estado
            usuarios.at[idx[0], "telefono"] = telefono
            usuarios.at[idx[0], "direccion"] = direccion
            if password:
                usuarios.at[idx[0], "password_hash"] = legacy.crear_hash_password(password)

            email_verified_actual = bool(usuarios.at[idx[0], "email_verified"]) if "email_verified" in usuarios.columns else False

            legacy.guardar_usuarios_df(usuarios[legacy.USUARIO_COLUMNS])
            legacy.registrar_actividad(
                f"Actualizo usuario {email} (ID {edit_id})\n- rol: {rol}\n- estado: {estado}\n- verificado: {email_verified_actual}"
            )
            flash("Usuario actualizado correctamente.", "success")
        else:
            ultimo_id = pd.to_numeric(usuarios["id_usuario"], errors="coerce").max()
            nuevo_id = int(ultimo_id + 1) if pd.notna(ultimo_id) else 1

            nuevo_usuario = {
                "id_usuario": nuevo_id,
                "nombre": nombre,
                "email": email,
                "password_hash": legacy.crear_hash_password(password),
                "rol": rol,
                "estado": estado,
                "fecha_registro": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "telefono": telefono,
                "direccion": direccion,
                "email_verified": False,
                "verification_code": "",
                "verification_code_expiry": "",
                "password_change_code": "",
                "password_change_code_expiry": "",
                "reset_token": "",
                "reset_token_expiry": "",
            }
            usuarios = pd.concat([usuarios, pd.DataFrame([nuevo_usuario])], ignore_index=True)
            legacy.guardar_usuarios_df(usuarios[legacy.USUARIO_COLUMNS])
            legacy.registrar_actividad(f"Creo usuario {email} (ID {nuevo_id})\n- rol: {rol}\n- estado: {estado}\n- verificado: False")
            flash("Usuario creado correctamente.", "success")

        return redirect(url_for("admin_usuarios", **retorno_kwargs))

    def admin_toggle_usuario_estado(id_usuario):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        buscar = request.form.get("buscar", "").strip()
        rol = request.form.get("rol", "").strip()
        estado_filtro = request.form.get("estado_filtro", "").strip().lower()
        orden = request.form.get("orden", "reciente").strip()
        pagina_actual = legacy._parse_positive_int(request.form.get("page", 1), default=1)

        usuarios = legacy.cargar_usuarios_df()
        usuarios["id_usuario"] = pd.to_numeric(usuarios["id_usuario"], errors="coerce")

        idx = usuarios[usuarios["id_usuario"] == id_usuario].index
        if idx.empty:
            flash("Usuario no encontrado.", "danger")
            return redirect(url_for("admin_usuarios", buscar=buscar, rol=rol, estado=estado_filtro, orden=orden, page=pagina_actual))

        usuario_email = legacy.normalizar_email(usuarios.at[idx[0], "email"])
        if usuario_email == legacy.normalizar_email(session.get("usuario", "")) and str(usuarios.at[idx[0], "estado"]).strip().lower() == "activo":
            flash("No puedes desactivar tu propia cuenta de administrador desde esta vista.", "warning")
            return redirect(url_for("admin_usuarios", buscar=buscar, rol=rol, estado=estado_filtro, orden=orden, page=pagina_actual))

        estado_actual = str(usuarios.at[idx[0], "estado"]).strip().lower()
        nuevo_estado = "inactivo" if estado_actual == "activo" else "activo"
        usuarios.at[idx[0], "estado"] = nuevo_estado
        legacy.guardar_usuarios_df(usuarios[legacy.USUARIO_COLUMNS])
        legacy.registrar_actividad(f"Cambio de estado para usuario {usuario_email} (ID {id_usuario}) -> {nuevo_estado}")
        flash(f"Usuario actualizado a estado {nuevo_estado}.", "success")
        return redirect(url_for("admin_usuarios", buscar=buscar, rol=rol, estado=estado_filtro, orden=orden, page=pagina_actual))

    def admin_registros():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        registros = legacy.cargar_registros_df()
        filtro_usuario = request.args.get("usuario", "").strip()
        filtro_fecha = request.args.get("fecha", "").strip()

        if filtro_usuario:
            registros = registros[
                registros["id_usuario"].astype(str).str.contains(filtro_usuario, case=False, na=False)
                | registros["accion"].astype(str).str.contains(filtro_usuario, case=False, na=False)
            ]

        if filtro_fecha:
            registros = registros[registros["fecha_accion"].astype(str).str.startswith(filtro_fecha)]

        registros["id_registro"] = pd.to_numeric(registros["id_registro"], errors="coerce")
        registros = registros.sort_values(by="id_registro", ascending=False, na_position="last")

        lista_registros = registros.fillna("").to_dict(orient="records")
        for registro in lista_registros:
            if registro.get("id_registro") != "":
                registro["id_registro"] = int(registro["id_registro"])

        filtros = {"usuario": filtro_usuario, "fecha": filtro_fecha}
        return render_template("Administrador/Gestion usuarios/admin_registros.html", registros=lista_registros, filtros=filtros)

    def admin_registros_export_excel():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        registros = legacy.cargar_registros_df()
        filtro_usuario = request.args.get("usuario", "").strip()
        filtro_fecha = request.args.get("fecha", "").strip()

        if filtro_usuario:
            registros = registros[
                registros["id_usuario"].astype(str).str.contains(filtro_usuario, case=False, na=False)
                | registros["accion"].astype(str).str.contains(filtro_usuario, case=False, na=False)
            ]
        if filtro_fecha:
            registros = registros[registros["fecha_accion"].astype(str).str.startswith(filtro_fecha)]

        buffer = BytesIO()
        registros.to_excel(buffer, index=False)
        buffer.seek(0)

        nombre_archivo = f"registros_bd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def admin_ajustes():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        if request.method == "POST":
            producto = request.form.get("producto", "").strip()
            try:
                precio = float(request.form.get("precio", 0))
            except ValueError:
                precio = 0
            if precio < 0:
                precio = 0
            if legacy.actualizar_precio_orden_personalizada(producto, precio):
                flash("Precio de prenda personalizada actualizado correctamente.", "success")
                legacy.registrar_actividad(f"Actualizo precio personalizado '{producto}' a {legacy.formatear_cop(precio)}")
            else:
                flash("No fue posible actualizar el precio seleccionado.", "warning")
            return redirect(url_for("admin_ajustes"))

        ordenes_df = legacy.cargar_ordenes_personalizadas_df()
        if not ordenes_df.empty and "estado" in ordenes_df.columns:
            ordenes_df = ordenes_df[~ordenes_df["estado"].astype(str).str.strip().str.lower().eq("pendiente_pago")].copy()
        if not ordenes_df.empty:
            ordenes_df = ordenes_df.sort_values(by="id_orden_personalizada", ascending=False, na_position="last")
        precios_df = legacy.cargar_precios_orden_personalizada_df().sort_values(by="nombre", ascending=True)
        nombres_producto = {row["producto"]: row["nombre"] for _, row in precios_df.iterrows()}
        if not ordenes_df.empty:
            ordenes_df = ordenes_df.copy()
            ordenes_df["producto_nombre"] = ordenes_df["producto"].apply(
                lambda valor: nombres_producto.get(legacy.producto_personalizado_canonico(valor), str(valor or "").replace("_", " ").title())
            )
        pendientes = int(ordenes_df["estado"].astype(str).str.strip().str.lower().eq("pendiente").sum()) if not ordenes_df.empty else 0
        return render_template(
            "Administrador/Prendas personalizadas/admin_prendas_personalizadas.html",
            ordenes_personalizadas=ordenes_df.to_dict("records"),
            precios_personalizados=precios_df.to_dict("records"),
            total_personalizadas=len(ordenes_df),
            pendientes_personalizadas=pendientes,
        )

    def admin_orden_personalizada_estado(id_orden):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        estado = request.form.get("estado", "pendiente").strip().lower()
        estados_validos = set(legacy.ORDEN_PERSONALIZADA_ESTADOS_PAGO.keys())
        if estado not in estados_validos:
            flash("Estado no valido para la solicitud personalizada.", "warning")
            return redirect(url_for("admin_ajustes"))
        legacy.asegurar_tablas_orden_personalizada()
        with legacy.engine.begin() as conn:
            conn.execute(
                legacy.sa.text(
                    """
                UPDATE orden_personalizada
                SET estado = :estado
                WHERE id_orden_personalizada = :id_orden
                """
                ),
                {"estado": estado, "id_orden": int(id_orden)},
            )
        legacy.registrar_actividad(f"Actualizo solicitud personalizada #{id_orden} a {estado}")
        flash("Estado de la solicitud actualizado.", "success")
        return redirect(url_for("admin_ajustes"))

    def admin_informes():
        if session.get("rol") != "admin":
            return "Acceso denegado"
        return render_template("Administrador/Informes/admin_infor_dashboard.html")

    def admin_pos():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        recibo_pedido_id = None
        descargar_recibo = request.args.get("descargar_recibo", "").strip()
        if descargar_recibo:
            try:
                recibo_pedido_id = int(descargar_recibo)
                if recibo_pedido_id <= 0:
                    recibo_pedido_id = None
            except ValueError:
                recibo_pedido_id = None

        productos = legacy.cargar_productos_activos_df()
        promos = legacy.cargar_promociones_df()
        hoy = datetime.now().date()
        mejor_promo_por_producto = legacy.obtener_mejor_promocion_por_producto(productos, promos, hoy)

        lista_productos = productos.to_dict(orient="records")
        for producto in lista_productos:
            id_producto = int(pd.to_numeric(producto.get("id_producto", 0), errors="coerce") or 0)
            galeria_producto = legacy.obtener_galeria_producto(id_producto, producto.get("imagen_url", ""))
            if galeria_producto:
                producto["imagen_url"] = legacy.normalizar_imagen_url(galeria_producto[0])
            else:
                producto["imagen_url"] = legacy.normalizar_imagen_url(producto.get("imagen_url", ""))
            producto["requiere_talla"] = bool(legacy.producto_requiere_talla(producto.get("intendencia", "")))
            precio_base = float(pd.to_numeric(producto.get("precio", 0), errors="coerce") or 0)
            promo = mejor_promo_por_producto.get(id_producto)

            if promo:
                descuento_unitario = legacy.calcular_descuento_promocion(precio_base, promo)
                precio_venta = max(0.0, precio_base - descuento_unitario)
                producto["promo_activa"] = True
                producto["promo_id"] = promo.get("id_promo", "")
                producto["promo_codigo"] = promo.get("codigo", "")
                producto["promo_nombre"] = promo.get("nombre", "")
                producto["promo_tipo_descuento"] = promo.get("tipo_descuento", "")
                producto["promo_valor_descuento"] = float(
                    pd.to_numeric(promo.get("valor_descuento", 0), errors="coerce") or 0
                )
                producto["promo_descuento_unitario"] = float(descuento_unitario)
                producto["precio_original"] = float(precio_base)
                producto["precio_con_descuento"] = float(precio_venta)
                producto["precio_venta"] = float(precio_venta)
            else:
                producto["promo_activa"] = False
                producto["promo_id"] = ""
                producto["promo_codigo"] = ""
                producto["promo_nombre"] = ""
                producto["promo_tipo_descuento"] = ""
                producto["promo_valor_descuento"] = 0.0
                producto["promo_descuento_unitario"] = 0.0
                producto["precio_original"] = float(precio_base)
                producto["precio_con_descuento"] = float(precio_base)
                producto["precio_venta"] = float(precio_base)

        return render_template(
            "Administrador/Sistema POS/admin_pos_dashboard.html",
            productos=lista_productos,
            fuerzas=legacy.FUERZAS_OPCIONES,
            tallas=legacy.TALLAS_OPCIONES,
            recibo_pedido_id=recibo_pedido_id,
        )

    def admin_pos_recibo_pdf(id_pedido):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        try:
            pdf_bytes = legacy.generar_pdf_recibo_pos(id_pedido)
        except ValueError as exc:
            return Response(str(exc), status=404, mimetype="text/plain; charset=utf-8")
        except Exception as exc:
            return Response(f"No se pudo generar el recibo PDF. {str(exc)}", status=500, mimetype="text/plain; charset=utf-8")

        return send_file(
            BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"recibo_pos_{id_pedido}.pdf",
        )

    def admin_pos_recibo_html(id_pedido):
        if session.get("rol") != "admin":
            return "Acceso denegado"
        try:
            return legacy.render_html_recibo_pos(id_pedido)
        except ValueError as exc:
            return Response(str(exc), status=404, mimetype="text/plain; charset=utf-8")
        except Exception as exc:
            return Response(f"No se pudo generar la vista del recibo. {str(exc)}", status=500, mimetype="text/plain; charset=utf-8")

    def admin_pedidos():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        pedidos = legacy.cargar_pedidos_df()
        pagos = legacy.cargar_pagos_df()
        detalle = legacy.cargar_detalle_pedido_df()
        productos = legacy.cargar_productos_df()
        usuarios = legacy.cargar_usuarios_df()

        pedidos, pagos, detalle, productos = legacy._normalizar_dataframes_admin_pedidos(pedidos, pagos, detalle, productos)
        pedidos_view = legacy._construir_vista_pedidos(pedidos, pagos, detalle, usuarios, productos)

        filtros, pago_filtros = legacy._leer_filtros_admin_pedidos(request.args)
        pedidos_view, filtros, paginacion, hay_filtros_activos = legacy._filtrar_y_paginar_pedidos(pedidos_view, filtros, per_page=10)
        pagina_curso_actual = legacy._parse_positive_int(request.args.get("curso_page", 1), default=1)
        lista_pedidos_base = legacy._enriquecer_pedidos_con_tracking(
            legacy._serializar_pedidos_admin(
                legacy._construir_vista_pedidos(pedidos, pagos, detalle, usuarios, productos).sort_values(
                    by="id_pedido", ascending=False, na_position="last"
                )
            )
        )
        pedidos_activos = [pedido for pedido in lista_pedidos_base if pedido.get("estado_activo")]
        pedidos_activos_vista, paginacion_curso = legacy._paginar_lista(pedidos_activos, pagina_curso_actual, per_page=6)
        resumen_estados = {
            "activos": len(pedidos_activos),
            "confirmados": sum(1 for pedido in pedidos_activos if pedido.get("estado_ui") == "confirmado"),
            "empaquetados": sum(1 for pedido in pedidos_activos if pedido.get("estado_ui") == "empaquetado"),
            "enviados": sum(1 for pedido in pedidos_activos if pedido.get("estado_ui") == "enviado"),
        }
        lista_pedidos = legacy._enriquecer_pedidos_con_tracking(legacy._serializar_pedidos_admin(pedidos_view))

        return render_template(
            "Administrador/Gestion pedidos/admin_orders.html",
            pedidos=lista_pedidos,
            filtros=filtros,
            paginacion=paginacion,
            hay_filtros_activos=hay_filtros_activos,
            pedidos_activos=pedidos_activos_vista,
            paginacion_curso=paginacion_curso,
            resumen_estados=resumen_estados,
            estados_pedido=legacy.PEDIDO_STATUS_FLOW,
        )

    def admin_pedidos_estado(id_pedido):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        filtros = {
            "q": str(request.form.get("f_q", "")).strip(),
            "estado": str(request.form.get("f_estado", "todos")).strip().lower(),
            "estado_pago": str(request.form.get("f_estado_pago", "todos")).strip().lower(),
            "fecha_desde": str(request.form.get("f_fecha_desde", "")).strip(),
            "fecha_hasta": str(request.form.get("f_fecha_hasta", "")).strip(),
            "page": str(request.form.get("f_page", "1")).strip(),
        }
        pago_filtros = {
            "q": str(request.form.get("f_pago_q", "")).strip(),
            "metodo": str(request.form.get("f_pago_metodo", "todos")).strip().lower(),
            "estado": str(request.form.get("f_pago_estado", "todos")).strip().lower(),
            "fecha_desde": str(request.form.get("f_pago_fecha_desde", "")).strip(),
            "fecha_hasta": str(request.form.get("f_pago_fecha_hasta", "")).strip(),
            "page": str(request.form.get("f_pago_page", "1")).strip(),
        }
        ajustes_page = legacy._parse_positive_int(request.form.get("ajustes_page", "1"), default=1)
        curso_page = legacy._parse_positive_int(request.form.get("curso_page", "1"), default=1)

        estado_nuevo = request.form.get("estado", "").strip().lower()
        origen = str(request.form.get("origen", "")).strip().lower()
        estados_validos = {clave for clave, _ in legacy.PEDIDO_STATUS_FLOW} | {"cancelado"}
        if estado_nuevo not in estados_validos:
            flash("Estado de pedido invalido.", "danger")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        pedidos = legacy.cargar_pedidos_df()
        if "id_pedido" not in pedidos.columns:
            flash("Estructura de pedidos invalida.", "danger")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)
        if "estado" not in pedidos.columns:
            pedidos["estado"] = "confirmado"

        pedidos["id_pedido"] = pd.to_numeric(pedidos["id_pedido"], errors="coerce")
        idx = pedidos[pedidos["id_pedido"] == id_pedido].index
        if idx.empty:
            flash("Pedido no encontrado.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        estado_anterior = str(pedidos.at[idx[0], "estado"]).strip().lower()
        id_usuario_pedido = pedidos.at[idx[0], "id_usuario"] if "id_usuario" in pedidos.columns else ""
        pedidos.at[idx[0], "estado"] = estado_nuevo
        legacy.guardar_pedidos_df(pedidos)

        if estado_anterior != estado_nuevo:
            legacy.registrar_actividad(f"Actualizo estado de pedido #{id_pedido}: {estado_anterior} -> {estado_nuevo}")
            flash(f'Pedido #{id_pedido} actualizado a "{estado_nuevo}".', "success")
            resultado_notificacion = legacy._notificar_actualizacion_pedido_cliente(
                id_pedido=id_pedido,
                id_usuario=id_usuario_pedido,
                estado_pedido=estado_nuevo,
                tipo_actualizacion="pedido",
            )
            legacy._flash_resultado_notificacion_pedido(resultado_notificacion, id_pedido)
        else:
            flash(f'El pedido #{id_pedido} ya estaba en "{estado_nuevo}".', "info")

        return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

    def admin_pedidos_pago(id_pedido):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        filtros = {
            "q": str(request.form.get("f_q", "")).strip(),
            "estado": str(request.form.get("f_estado", "todos")).strip().lower(),
            "estado_pago": str(request.form.get("f_estado_pago", "todos")).strip().lower(),
            "fecha_desde": str(request.form.get("f_fecha_desde", "")).strip(),
            "fecha_hasta": str(request.form.get("f_fecha_hasta", "")).strip(),
            "page": str(request.form.get("f_page", "1")).strip(),
        }
        pago_filtros = {
            "q": str(request.form.get("f_pago_q", "")).strip(),
            "metodo": str(request.form.get("f_pago_metodo", "todos")).strip().lower(),
            "estado": str(request.form.get("f_pago_estado", "todos")).strip().lower(),
            "fecha_desde": str(request.form.get("f_pago_fecha_desde", "")).strip(),
            "fecha_hasta": str(request.form.get("f_pago_fecha_hasta", "")).strip(),
            "page": str(request.form.get("f_pago_page", "1")).strip(),
        }
        ajustes_page = legacy._parse_positive_int(request.form.get("ajustes_page", "1"), default=1)
        curso_page = legacy._parse_positive_int(request.form.get("curso_page", "1"), default=1)
        origen = str(request.form.get("origen", "")).strip().lower()

        accion = str(request.form.get("accion_pago", "")).strip().lower()
        mapa_acciones = {"revision": "en_revision", "aprobar": "aprobado", "rechazar": "rechazado"}
        estado_pago_nuevo = mapa_acciones.get(accion, "")
        if not estado_pago_nuevo:
            flash("Accion de pago invalida.", "danger")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        pagos = legacy.cargar_pagos_df()
        if pagos.empty or "id_pedido" not in pagos.columns:
            flash("No se encontro un pago registrado para este pedido.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        pagos = pagos.copy()
        pagos["id_pedido_num"] = pd.to_numeric(pagos["id_pedido"], errors="coerce")
        pagos["id_pago_num"] = pd.to_numeric(pagos["id_pago"], errors="coerce")
        pagos_ordenados = pagos.sort_values(by="id_pago_num", ascending=False, na_position="last")
        idx_pago = pagos_ordenados[pagos_ordenados["id_pedido_num"] == id_pedido].index
        if idx_pago.empty:
            flash("No se encontro un pago registrado para este pedido.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        idx_pago = idx_pago[0]
        estado_pago_anterior = str(pagos.at[idx_pago, "estado_pago"]).strip().lower()
        metodo_pago = str(pagos.at[idx_pago, "metodo_pago"]).strip().lower()
        comprobante_url = str(pagos.at[idx_pago, "comprobante_url"]).strip() if "comprobante_url" in pagos.columns else ""
        if metodo_pago != "transferencia":
            flash("Solo las transferencias requieren validacion manual.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)
        if estado_pago_nuevo == "aprobado" and not comprobante_url:
            flash("No puedes aprobar la transferencia sin un comprobante adjunto en la web.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        pedidos = legacy.cargar_pedidos_df()
        if pedidos.empty or "id_pedido" not in pedidos.columns:
            flash("No se encontro el pedido asociado.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        pedidos = pedidos.copy()
        pedidos["id_pedido_num"] = pd.to_numeric(pedidos["id_pedido"], errors="coerce")
        idx_pedido = pedidos[pedidos["id_pedido_num"] == id_pedido].index
        if idx_pedido.empty:
            flash("No se encontro el pedido asociado.", "warning")
            return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

        idx_pedido = idx_pedido[0]
        estado_pedido_anterior = str(pedidos.at[idx_pedido, "estado"]).strip().lower() or "confirmado"
        id_usuario_pedido = pedidos.at[idx_pedido, "id_usuario"] if "id_usuario" in pedidos.columns else ""

        pagos.at[idx_pago, "estado_pago"] = estado_pago_nuevo
        if estado_pago_nuevo == "aprobado":
            pedidos.at[idx_pedido, "estado"] = "confirmado"
        elif estado_pago_nuevo == "rechazado":
            pedidos.at[idx_pedido, "estado"] = "cancelado"
        else:
            pedidos.at[idx_pedido, "estado"] = "pago_en_revision"
        estado_pedido_nuevo = str(pedidos.at[idx_pedido, "estado"]).strip().lower()

        if "id_pedido_num" in pedidos.columns:
            pedidos = pedidos.drop(columns=["id_pedido_num"])
        if "id_pedido_num" in pagos.columns:
            pagos = pagos.drop(columns=["id_pedido_num"])
        if "id_pago_num" in pagos.columns:
            pagos = pagos.drop(columns=["id_pago_num"])

        legacy.guardar_pagos_df(pagos)
        legacy.guardar_pedidos_df(pedidos)

        if estado_pago_anterior != estado_pago_nuevo or estado_pedido_anterior != estado_pedido_nuevo:
            legacy.registrar_actividad(
                f"Actualizo revision de pago para pedido #{id_pedido}: "
                f"{estado_pago_anterior or 'sin_estado'} -> {estado_pago_nuevo}"
            )
            if estado_pago_nuevo == "aprobado":
                flash(f"Transferencia del pedido #{id_pedido} aprobada. El pedido quedo confirmado.", "success")
            elif estado_pago_nuevo == "rechazado":
                flash(f"Transferencia del pedido #{id_pedido} rechazada. El pedido quedo cancelado.", "warning")
            else:
                flash(f"El pago del pedido #{id_pedido} quedo en revision.", "info")
            resultado_notificacion = legacy._notificar_actualizacion_pedido_cliente(
                id_pedido=id_pedido,
                id_usuario=id_usuario_pedido,
                estado_pedido=estado_pedido_nuevo,
                estado_pago=estado_pago_nuevo,
                tipo_actualizacion="pago",
            )
            legacy._flash_resultado_notificacion_pedido(resultado_notificacion, id_pedido)
        else:
            flash(f'El pago del pedido #{id_pedido} ya estaba en "{estado_pago_nuevo}".', "info")

        return legacy._redirigir_admin_pedidos_por_origen(origen, filtros, pago_filtros, ajustes_page, curso_page)

    def admin_dashboard():
        if session.get("rol") != "admin":
            return "Acceso denegado"
        return render_template("Administrador/admin_dashboard_principal.html", **legacy._construir_contexto_admin_dashboard())

    def admin_actualizar_destacados():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        resultado = legacy._actualizar_destacados_dashboard()
        if resultado is None:
            flash("No hay productos para destacar.", "warning")
            return redirect(url_for("admin_dashboard"))

        ids_ejercito = resultado["ids_ejercito"]
        ids_policia = resultado["ids_policia"]
        ids_armada = resultado["ids_armada"]
        ids_gaula = resultado["ids_gaula"]
        ids_variado = resultado["ids_variado"]
        ids_accesorios = resultado["ids_accesorios"]
        total_destacados = resultado["total_destacados"]

        legacy.registrar_actividad(
            "Administrador actualizo prendas destacadas por categoria:\n"
            f"- Ejercito: {len(ids_ejercito)} ({', '.join(str(x) for x in ids_ejercito) if ids_ejercito else 'ninguna'})\n"
            f"- Policia: {len(ids_policia)} ({', '.join(str(x) for x in ids_policia) if ids_policia else 'ninguna'})\n"
            f"- Armada: {len(ids_armada)} ({', '.join(str(x) for x in ids_armada) if ids_armada else 'ninguna'})\n"
            f"- Gaula: {len(ids_gaula)} ({', '.join(str(x) for x in ids_gaula) if ids_gaula else 'ninguna'})\n"
            f"- Variado: {len(ids_variado)} ({', '.join(str(x) for x in ids_variado) if ids_variado else 'ninguna'})\n"
            f"- Accesorios: {len(ids_accesorios)} ({', '.join(str(x) for x in ids_accesorios) if ids_accesorios else 'ninguna'})\n"
            f"- Total destacadas: {total_destacados}"
        )

        flash("Prendas destacadas actualizadas correctamente.", "success")
        return redirect(url_for("admin_dashboard"))

    def admin_productos():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        busqueda = request.args.get("q", "").strip()
        fuerza_filtro = request.args.get("fuerza", "").strip()
        intendencia_filtro = request.args.get("intendencia", "").strip()
        contexto = legacy._construir_contexto_admin_productos(busqueda, fuerza_filtro, intendencia_filtro)

        return render_template(
            "Administrador/Gestion productos/admin_prod_dashboard.html",
            productos=contexto["productos"],
            productos_por_fuerza=contexto["productos_por_fuerza"],
            fuerzas=contexto["fuerzas"],
            intendencias=contexto["intendencias"],
            search_query=contexto["search_query"],
            selected_fuerza=contexto["selected_fuerza"],
            selected_intendencia=contexto["selected_intendencia"],
            has_filters=contexto["has_filters"],
            total_productos=contexto["total_productos"],
            total_eliminados=contexto["total_eliminados"],
            max_images_per_product=legacy.MAX_IMAGES_PER_PRODUCT,
        )

    def admin_pos_checkout():
        if session.get("rol") != "admin":
            return "Acceso denegado"

        cliente_nombre = request.form.get("cliente_nombre", "").strip()
        cliente_correo = request.form.get("cliente_correo", "").strip().lower()
        cliente_documento = request.form.get("cliente_documento", "").strip()
        cliente_telefono = request.form.get("cliente_telefono", "").strip()
        error_cliente = legacy._validar_cliente_pos(cliente_nombre, cliente_correo, cliente_documento, cliente_telefono)
        if error_cliente:
            flash(error_cliente[0], error_cliente[1])
            return redirect(url_for("admin_pos"))

        items_raw = request.form.get("items_json", "[]")
        metodo_pago = legacy._normalizar_metodo_pago_pos(request.form.get("metodo_pago", "efectivo"))

        items, error_items = legacy._parsear_items_checkout_pos(items_raw)
        if error_items:
            flash(error_items[0], error_items[1])
            return redirect(url_for("admin_pos"))

        productos = legacy.cargar_productos_df()
        if productos.empty:
            flash("No existe el catalogo de productos.", "danger")
            return redirect(url_for("admin_pos"))

        promos = legacy.cargar_promociones_df()
        hoy = datetime.now().date()
        mejor_promo_por_producto = legacy.obtener_mejor_promocion_por_producto(productos, promos, hoy)

        resultado_carrito, error_carrito = legacy._validar_y_preparar_carrito_pos(items, productos, mejor_promo_por_producto)
        if error_carrito:
            flash(error_carrito[0], error_carrito[1])
            return redirect(url_for("admin_pos"))

        carrito_validado = resultado_carrito["carrito_validado"]
        total_bruto = resultado_carrito["total_bruto"]
        total_descuento = resultado_carrito["total_descuento"]
        total = resultado_carrito["total"]
        legacy.guardar_productos_df(resultado_carrito["productos"])

        next_pedido_id = legacy._registrar_venta_pos_admin(carrito_validado, metodo_pago, total, total_descuento)

        total_bruto = round(total_bruto, 2)
        total_descuento = round(total_descuento, 2)
        total = round(total, 2)
        total_cop = legacy.formatear_cop(total)
        total_bruto_cop = legacy.formatear_cop(total_bruto)
        descuento_cop = legacy.formatear_cop(total_descuento)
        legacy.registrar_actividad(
            f"POS registro venta #{next_pedido_id} por {total_cop} ({len(carrito_validado)} producto(s))\n"
            f"- total bruto: {total_bruto_cop}\n"
            f"- descuento aplicado: {descuento_cop}\n"
            f"- cliente: {cliente_nombre}\n"
            f"- correo: {cliente_correo}\n"
            f"- documento: {cliente_documento}\n"
            f"- telefono: {cliente_telefono}"
        )
        if total_descuento > 0:
            flash(
                f"Venta POS registrada. Pedido #{next_pedido_id} por {total_cop}. "
                f"Descuento aplicado: {descuento_cop}.",
                "success",
            )
        else:
            flash(f"Venta POS registrada. Pedido #{next_pedido_id} por {total_cop}.", "success")

        session["ultimo_recibo_pos"] = {
            "id_pedido": int(next_pedido_id),
            "cliente_nombre": cliente_nombre,
            "cliente_correo": cliente_correo,
            "cliente_documento": cliente_documento,
            "cliente_telefono": cliente_telefono,
        }
        session.modified = True
        return redirect(url_for("admin_pos", descargar_recibo=next_pedido_id))

    def editar_producto(id_producto):
        if session.get("rol") != "admin":
            return "Acceso denegado"

        productos = legacy.cargar_productos_df()
        producto = productos[productos["id_producto"] == id_producto]
        if producto.empty:
            return "Producto no encontrado."

        if request.method == "POST":
            idx = producto.index[0]
            anterior = {
                "nombre": str(productos.at[idx, "nombre"]) if "nombre" in productos.columns else "",
                "descripcion": str(productos.at[idx, "descripcion"]) if "descripcion" in productos.columns else "",
                "precio": float(pd.to_numeric(productos.at[idx, "precio"], errors="coerce")) if "precio" in productos.columns else 0.0,
                "stock": int(pd.to_numeric(productos.at[idx, "stock"], errors="coerce")) if "stock" in productos.columns else 0,
                "id_categoria": str(productos.at[idx, "id_categoria"]) if "id_categoria" in productos.columns else "",
                "fuerza": str(productos.at[idx, "fuerza"]) if "fuerza" in productos.columns else "",
                "intendencia": str(productos.at[idx, "intendencia"]) if "intendencia" in productos.columns else "",
            }

            nuevo_nombre = request.form["nombre"].strip()
            nueva_descripcion = request.form["descripcion"].strip()
            nuevo_precio = float(request.form["precio"])
            nuevo_stock = int(request.form["stock"])
            nueva_fuerza = request.form.get("fuerza", anterior["fuerza"]).strip()
            nueva_intendencia = request.form.get("intendencia", anterior["intendencia"]).strip()
            if nueva_fuerza not in legacy.FUERZAS_OPCIONES or nueva_intendencia not in legacy.INTENDENCIAS_OPCIONES:
                flash("Selecciona una fuerza e intendencia validas.", "danger")
                return redirect(url_for("admin_productos"))

            nueva_categoria = anterior["id_categoria"]
            if "id_categoria" in productos.columns:
                categoria_form = request.form.get("id_categoria", "").strip()
                if categoria_form:
                    try:
                        nueva_categoria = int(float(categoria_form))
                    except ValueError:
                        nueva_categoria = anterior["id_categoria"]
                nueva_categoria_num = pd.to_numeric(nueva_categoria, errors="coerce")
                if pd.isna(nueva_categoria_num):
                    nueva_categoria_num = pd.to_numeric(productos.at[idx, "id_categoria"], errors="coerce")
                nueva_categoria = int(nueva_categoria_num) if pd.notna(nueva_categoria_num) else 0

            productos.at[idx, "nombre"] = nuevo_nombre
            productos.at[idx, "descripcion"] = nueva_descripcion
            productos.at[idx, "precio"] = nuevo_precio
            productos.at[idx, "stock"] = nuevo_stock
            productos.at[idx, "fuerza"] = nueva_fuerza
            productos.at[idx, "intendencia"] = nueva_intendencia
            if "id_categoria" in productos.columns:
                productos.at[idx, "id_categoria"] = nueva_categoria

            legacy.guardar_productos_df(productos)

            cambios = []
            if anterior["nombre"] != nuevo_nombre:
                cambios.append(f"- nombre: '{anterior['nombre']}' -> '{nuevo_nombre}'")
            if anterior["descripcion"] != nueva_descripcion:
                cambios.append(f"- descripcion: '{anterior['descripcion']}' -> '{nueva_descripcion}'")
            if round(float(anterior["precio"]), 2) != round(float(nuevo_precio), 2):
                cambios.append(f"- precio: {legacy.formatear_cop(anterior['precio'])} -> {legacy.formatear_cop(nuevo_precio)}")
            if int(anterior["stock"]) != int(nuevo_stock):
                cambios.append(f"- stock: {anterior['stock']} -> {nuevo_stock}")
            if anterior["fuerza"] != nueva_fuerza:
                cambios.append(f"- fuerza: {anterior['fuerza']} -> {nueva_fuerza}")
            if anterior["intendencia"] != nueva_intendencia:
                cambios.append(f"- intendencia: {anterior['intendencia']} -> {nueva_intendencia}")
            if str(anterior["id_categoria"]) != str(nueva_categoria):
                cambios.append(f"- categoria: {anterior['id_categoria']} -> {nueva_categoria}")

            encabezado = f"Actualizo producto '{nuevo_nombre}' (ID {id_producto})"
            if cambios:
                legacy.registrar_actividad(encabezado + "\n" + "\n".join(cambios))
            else:
                legacy.registrar_actividad(encabezado + "\n- sin cambios detectados")

            return redirect(url_for("admin_productos"))

        return render_template(
            "Administrador/Gestion productos/editar_producto.html",
            producto=producto.iloc[0],
            fuerzas=legacy.FUERZAS_OPCIONES,
            intendencias=legacy.INTENDENCIAS_OPCIONES,
        )


    app.add_url_rule("/admin/promo", endpoint="admin_promo", view_func=admin_promo, methods=["GET", "POST"])
    app.add_url_rule("/admin/promo/toggle/<int:id_promo>", endpoint="admin_promo_toggle", view_func=admin_promo_toggle, methods=["POST"])
    app.add_url_rule("/admin/charts", endpoint="admin_charts", view_func=admin_charts)
    app.add_url_rule("/admin/charts/export_excel", endpoint="admin_charts_export_excel", view_func=admin_charts_export_excel)
    app.add_url_rule("/admin/agregar_producto", endpoint="agregar_producto", view_func=agregar_producto, methods=["POST"])
    app.add_url_rule("/admin/sales/export", endpoint="export_sales", view_func=export_sales)
    app.add_url_rule("/admin/imagen/<int:id_producto>", endpoint="subir_imagen", view_func=subir_imagen, methods=["POST"])
    app.add_url_rule(
        "/admin/imagen/agregar/<int:id_producto>",
        endpoint="agregar_imagenes_producto",
        view_func=agregar_imagenes_producto,
        methods=["POST"],
    )
    app.add_url_rule(
        "/admin/imagen/eliminar/<int:id_producto>",
        endpoint="eliminar_imagen_producto",
        view_func=eliminar_imagen_producto,
        methods=["POST"],
    )
    app.add_url_rule("/admin/eliminar/<int:id_producto>", endpoint="eliminar_producto", view_func=eliminar_producto, methods=["POST"])
    app.add_url_rule(
        "/admin/eliminar_definitivo/<int:id_producto>",
        endpoint="eliminar_definitivo",
        view_func=eliminar_definitivo,
        methods=["POST"],
    )
    app.add_url_rule("/admin/restaurar/<int:id_producto>", endpoint="restaurar_producto", view_func=restaurar_producto, methods=["POST"])
    app.add_url_rule("/admin/papelera", endpoint="admin_papelera", view_func=admin_papelera)
    app.add_url_rule("/admin/usuarios", endpoint="admin_usuarios", view_func=admin_usuarios)
    app.add_url_rule("/admin/usuarios/guardar", endpoint="admin_guardar_usuario", view_func=admin_guardar_usuario, methods=["POST"])
    app.add_url_rule(
        "/admin/usuarios/estado/<int:id_usuario>",
        endpoint="admin_toggle_usuario_estado",
        view_func=admin_toggle_usuario_estado,
        methods=["POST"],
    )
    app.add_url_rule("/admin/registros", endpoint="admin_registros", view_func=admin_registros)
    app.add_url_rule("/admin/registros/export_excel", endpoint="admin_registros_export_excel", view_func=admin_registros_export_excel)
    app.add_url_rule("/admin/ajustes", endpoint="admin_ajustes", view_func=admin_ajustes, methods=["GET", "POST"])
    app.add_url_rule(
        "/admin/ajustes/orden-personalizada/<int:id_orden>/estado",
        endpoint="admin_orden_personalizada_estado",
        view_func=admin_orden_personalizada_estado,
        methods=["POST"],
    )
    app.add_url_rule("/admin/informes", endpoint="admin_informes", view_func=admin_informes)
    app.add_url_rule("/admin/pos", endpoint="admin_pos", view_func=admin_pos)
    app.add_url_rule("/admin/pos/recibo/<int:id_pedido>.pdf", endpoint="admin_pos_recibo_pdf", view_func=admin_pos_recibo_pdf)
    app.add_url_rule("/admin/pos/recibo/<int:id_pedido>/html", endpoint="admin_pos_recibo_html", view_func=admin_pos_recibo_html)
    app.add_url_rule("/admin/pedidos", endpoint="admin_pedidos", view_func=admin_pedidos)
    app.add_url_rule("/admin/pedidos/estado/<int:id_pedido>", endpoint="admin_pedidos_estado", view_func=admin_pedidos_estado, methods=["POST"])
    app.add_url_rule("/admin/pedidos/pago/<int:id_pedido>", endpoint="admin_pedidos_pago", view_func=admin_pedidos_pago, methods=["POST"])
    app.add_url_rule("/admin", endpoint="admin_dashboard", view_func=admin_dashboard)
    app.add_url_rule("/admin/destacados", endpoint="admin_actualizar_destacados", view_func=admin_actualizar_destacados, methods=["POST"])
    app.add_url_rule("/admin/productos", endpoint="admin_productos", view_func=admin_productos)
    app.add_url_rule("/admin/pos/checkout", endpoint="admin_pos_checkout", view_func=admin_pos_checkout, methods=["POST"])
    app.add_url_rule("/admin/editar/<int:id_producto>", endpoint="editar_producto", view_func=editar_producto, methods=["GET", "POST"])
