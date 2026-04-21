import os, json, re, random, shutil, subprocess, base64, secrets, hashlib, pandas as pd, sqlalchemy as sa
from io import BytesIO
from pathlib import Path
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from flask import Flask, render_template, request, redirect, url_for, session, Response, flash, send_file, jsonify
from datetime import datetime, timedelta, date
from typing import Any, Mapping, Optional
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from email_service import (
    mail,
    generar_codigo_verificacion,
    enviar_codigo_verificacion,
    enviar_recuperacion_password,
)
from stripe_service import (
    stripe_estado_configuracion,
    crear_checkout_sesion_tarjeta,
    obtener_checkout_sesion,
)
load_dotenv()
import config_email
from db_utils import engine, read_table_df, replace_table_df, next_id  # Inicializa DB y proxys al importar

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True
app.secret_key = os.getenv("SECRET_KEY", "").strip()
if not app.secret_key:
    app.secret_key = os.urandom(32)
    print(
        "ADVERTENCIA: SECRET_KEY no está configurada en el entorno. "
        "Usando una clave temporal para esta ejecución."
    )

# Configuración de Flask-Mail para Gmail
app.config['MAIL_SERVER'] = config_email.MAIL_SERVER
app.config['MAIL_PORT'] = config_email.MAIL_PORT
app.config['MAIL_USE_TLS'] = config_email.MAIL_USE_TLS
app.config['MAIL_USERNAME'] = config_email.MAIL_USERNAME
app.config['MAIL_PASSWORD'] = config_email.MAIL_PASSWORD
app.config['MAIL_DEFAULT_SENDER'] = config_email.MAIL_DEFAULT_SENDER
app.config['PROJECT_NAME'] = 'NACHOHER'

# Inicializar Flask-Mail
mail.init_app(app)

CURRENCY_CODE = "COP"
CURRENCY_NAME = "Peso colombiano"

USUARIO_COLUMNS = [
    'id_usuario',
    'nombre',
    'email',
    'password_hash',
    'rol',
    'estado',
    'fecha_registro',
    'telefono',
    'direccion',
    'email_verified',
    'verification_code',
    'verification_code_expiry',
    'reset_token',
    'reset_token_expiry',
    'password_change_code',
    'password_change_code_expiry',
]
REGISTRO_COLUMNS = ['id_registro', 'id_usuario', 'accion', 'fecha_accion']
PRODUCTO_COLUMNS = [
    'id_producto', 'nombre', 'descripcion', 'precio', 'stock',
    'id_categoria', 'fuerza', 'intendencia', 'imagen_url',
    'eliminado', 'destacado_dashboard'
]
PEDIDO_COLUMNS = ['id_pedido', 'id_usuario', 'fecha_pedido', 'estado', 'cliente_telefono', 'cliente_direccion']
DETALLE_PEDIDO_COLUMNS = ['id_detalle', 'id_pedido', 'id_producto', 'cantidad', 'subtotal', 'talla']
PAGO_COLUMNS = ['id_pago', 'id_pedido', 'monto', 'metodo_pago', 'fecha_pago', 'estado_pago', 'id_promo', 'codigo_promo', 'tipo_descuento', 'valor_descuento', 'monto_descuento']
PEDIDO_STATUS_FLOW = [
    ('pendiente', 'Pendiente'),
    ('confirmado', 'Confirmado'),
    ('en_preparacion', 'En preparación'),
    ('empaquetado', 'Empaquetado'),
    ('enviado', 'Enviado'),
    ('entregado', 'Entregado'),
]
PEDIDO_STATUS_EXTRA = {
    'cancelado': 'Cancelado',
    'pendiente_revision': 'Pendiente de revisión',
    'completado': 'Completado',
}
PEDIDO_STATUS_ALIAS = {
    'pendiente_revision': 'confirmado',
    'completado': 'entregado',
}
PEDIDO_STATUS_LABELS = {
    **{clave: etiqueta for clave, etiqueta in PEDIDO_STATUS_FLOW},
    **PEDIDO_STATUS_EXTRA,
}

# Column definitions for promociones
PROMO_COLUMNS = [
    'id_promo', 'nombre', 'descripcion',
    'tipo_descuento', 'valor_descuento',
    'id_producto',
    'codigo',
    'fecha_inicio', 'fecha_fin',
    'activo'
]

FUERZAS_OPCIONES = ["Policia", "Ejercito", "Armada", "Gaula"]
INTENDENCIAS_OPCIONES = [
    "Busos", "Camibusos", "Gorras", "Pañoletas", "Sudaderas",
    "Pantalonetas", "Colchas", "Tendidos", "Chuspas para ropa sucia",
    "Fundas para almohadas", "Camuflados", "Accesorios", "Presillas"
]
TALLAS_OPCIONES = ["XS", "S", "M", "L", "XL", "XXL", "XXXL"]
INTENDENCIAS_SIN_TALLA = {
    "pañoleta", "pañoletas",
    "panoleta", "panoletas",
    "gorra", "gorras",
    "colcha", "colchas",
    "tendido", "tendidos",
    "chuspa para ropa sucia", "chuspas para ropa sucia",
    "funda para almohadas", "fundas para almohadas",
    "accesorio", "accesorios",
    "presilla", "presillas",
}
ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
MAX_IMAGE_SIZE_BYTES = 3 * 1024 * 1024
MAX_IMAGES_PER_PRODUCT = 5
REGISTER_CODE_EXP_MINUTES = 7
PASSWORD_RESET_EXP_MINUTES = 30
PASSWORD_CHANGE_CODE_EXP_MINUTES = 7
PENDING_REGISTRATIONS = {}

# DB init y proxys se hacen al importar db_utils


def normalizar_intendencia(valor):
    texto = str(valor or "").strip().lower()
    return re.sub(r"\s+", " ", texto)


def producto_requiere_talla(intendencia):
    return normalizar_intendencia(intendencia) not in INTENDENCIAS_SIN_TALLA


def cargar_usuarios_df():
    usuarios = read_table_df('usuarios')
    if usuarios.empty:
        usuarios = pd.DataFrame(columns=USUARIO_COLUMNS)

    for col in USUARIO_COLUMNS:
        if col not in usuarios.columns:
            if col == 'estado':
                usuarios[col] = 'activo'
            elif col == 'email_verified':
                usuarios[col] = False
            else:
                usuarios[col] = ''
    
    # Asegurar que las columnas de tokens sean tipo string
    if 'verification_code' in usuarios.columns:
        usuarios['verification_code'] = usuarios['verification_code'].fillna('').astype(str)
    if 'verification_code_expiry' in usuarios.columns:
        usuarios['verification_code_expiry'] = usuarios['verification_code_expiry'].fillna('').astype(str)
    if 'reset_token' in usuarios.columns:
        usuarios['reset_token'] = usuarios['reset_token'].fillna('').astype(str)
    if 'reset_token_expiry' in usuarios.columns:
        usuarios['reset_token_expiry'] = usuarios['reset_token_expiry'].fillna('').astype(str)
    if 'password_change_code' in usuarios.columns:
        usuarios['password_change_code'] = usuarios['password_change_code'].fillna('').astype(str)
    if 'password_change_code_expiry' in usuarios.columns:
        usuarios['password_change_code_expiry'] = usuarios['password_change_code_expiry'].fillna('').astype(str)
    if 'email_verified' in usuarios.columns:
        usuarios['email_verified'] = usuarios['email_verified'].fillna(False).astype(bool)
    
    usuarios['estado'] = usuarios['estado'].fillna('activo').astype(str).str.strip().str.lower()
    usuarios.loc[~usuarios['estado'].isin(['activo', 'inactivo']), 'estado'] = 'activo'
    return usuarios[USUARIO_COLUMNS]


def guardar_usuarios_df(usuarios):
    """
    Persiste el DataFrame de usuarios en PostgreSQL manteniendo tipos correctos
    para fechas y booleanos, y normalizando los tokens a texto.
    """
    df = usuarios.copy()

    # Normalizar columnas de texto
    for col in ['verification_code', 'reset_token', 'password_change_code']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str)

    # Convertir columnas de fecha/hora a datetime o None
    for col in ['fecha_registro', 'verification_code_expiry', 'reset_token_expiry', 'password_change_code_expiry']:
        if col in df.columns:
            col_series = df[col].replace('', pd.NA)
            parsed = pd.to_datetime(col_series, errors='coerce')
            df[col] = parsed.where(parsed.notna(), None)

    # Asegurar booleanos
    if 'email_verified' in df.columns:
        df['email_verified'] = df['email_verified'].fillna(False).astype(bool)

    replace_table_df('usuarios', df[USUARIO_COLUMNS])


def cargar_registros_df():
    registros = read_table_df('registros')
    if registros.empty:
        registros = pd.DataFrame(columns=REGISTRO_COLUMNS)

    for col in REGISTRO_COLUMNS:
        if col not in registros.columns:
            registros[col] = ''
    registros['id_registro'] = pd.to_numeric(registros['id_registro'], errors='coerce')
    registros['id_usuario'] = registros['id_usuario'].fillna('')
    registros['accion'] = registros['accion'].fillna('')
    registros['fecha_accion'] = registros['fecha_accion'].fillna('')
    return registros[REGISTRO_COLUMNS]


def guardar_registros_df(registros):
    replace_table_df('registros', registros[REGISTRO_COLUMNS])


def cargar_promociones_df():
    """Carga el DataFrame de promociones desde la tabla SQL."""
    promos = read_table_df('promociones')
    if promos.empty:
        promos = pd.DataFrame(columns=PROMO_COLUMNS)

    for col in PROMO_COLUMNS:
        if col not in promos.columns:
            promos[col] = ''

    # compatibilidad con estructura anterior
    if 'valor_descuento' not in promos.columns and 'descuento' in promos.columns:
        promos['valor_descuento'] = promos['descuento']
    if 'tipo_descuento' not in promos.columns:
        promos['tipo_descuento'] = 'porcentaje'
    if 'codigo' not in promos.columns:
        promos['codigo'] = ''
    if 'id_producto' not in promos.columns:
        promos['id_producto'] = pd.Series(dtype='float')

    # asegurar tipos
    promos['valor_descuento'] = pd.to_numeric(promos['valor_descuento'], errors='coerce').fillna(0)
    promos['tipo_descuento'] = promos['tipo_descuento'].astype(str).str.strip().str.lower()
    promos.loc[~promos['tipo_descuento'].isin(['porcentaje', 'valor_fijo']), 'tipo_descuento'] = 'porcentaje'
    promos['id_producto'] = pd.to_numeric(promos['id_producto'], errors='coerce')
    promos['codigo'] = promos['codigo'].fillna('').astype(str).str.strip().str.upper()
    promos['activo'] = promos['activo'].astype(bool)
    return promos[PROMO_COLUMNS]


def guardar_promociones_df(promos):
    """Guarda las promociones en la tabla SQL."""
    replace_table_df('promociones', promos[PROMO_COLUMNS])


def normalizar_imagen_url(valor):
    ruta = str(valor or '').strip().replace('\\', '/')
    if not ruta or ruta.lower() == 'nan':
        return ''
    if ruta.startswith('img/Empresa/') or ruta.startswith('img/Pagina/') or ruta.startswith('img/catalogo/'):
        return ruta
    if ruta.startswith('img/'):
        return f"img/Empresa/{ruta.split('/')[-1]}"
    return ruta


def normalizar_imagenes_productos(productos):
    if 'imagen_url' not in productos.columns:
        productos['imagen_url'] = ''
    productos['imagen_url'] = productos['imagen_url'].apply(normalizar_imagen_url)
    return productos


def extension_imagen(nombre_archivo):
    if not nombre_archivo or '.' not in nombre_archivo:
        return ''
    return nombre_archivo.rsplit('.', 1)[-1].lower()


def normalizar_email(valor):
    return str(valor or '').strip().lower()


def email_es_valido(email):
    return bool(re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', str(email or '').strip()))


def limpiar_registros_pendientes():
    ahora = datetime.now()
    expirados = []
    for email, data in PENDING_REGISTRATIONS.items():
        expiry_at = data.get('expiry_at')
        if not expiry_at or ahora > expiry_at:
            expirados.append(email)
    for email in expirados:
        PENDING_REGISTRATIONS.pop(email, None)


def obtener_registro_pendiente(email):
    limpiar_registros_pendientes()
    return PENDING_REGISTRATIONS.get(normalizar_email(email))


def guardar_registro_pendiente(email, codigo, nombre='', password=''):
    expiry_at = datetime.now() + timedelta(minutes=REGISTER_CODE_EXP_MINUTES)
    PENDING_REGISTRATIONS[normalizar_email(email)] = {
        'code': str(codigo).strip(),
        'expiry_at': expiry_at,
        'nombre': str(nombre).strip(),
        'password': str(password)
    }
    return expiry_at


def password_esta_hasheado(valor):
    texto = str(valor or '').strip()
    return texto.startswith('pbkdf2:') or texto.startswith('scrypt:')


def crear_hash_password(password):
    return generate_password_hash(str(password or ''))


def password_coincide(password_guardado, password_plano):
    guardado = str(password_guardado or '')
    plano = str(password_plano or '')
    if not guardado or not plano:
        return False

    if password_esta_hasheado(guardado):
        try:
            return check_password_hash(guardado, plano)
        except Exception:
            return False

    return guardado == plano


def password_cumple_estandares(password):
    patron = r'(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9]).{8,}'
    return bool(re.fullmatch(patron, str(password or '')))


def timestamp_expirado(valor_expiry):
    expiry_dt = pd.to_datetime(valor_expiry, errors='coerce')
    if pd.isna(expiry_dt):
        return True

    tz = getattr(expiry_dt, 'tz', None)
    if tz is not None:
        ahora = pd.Timestamp.now(tz=tz)
    else:
        ahora = pd.Timestamp.now()
    return ahora > expiry_dt


def asegurar_columnas_cambio_password(usuarios):
    if 'password_change_code' not in usuarios.columns:
        usuarios['password_change_code'] = ''
    if 'password_change_code_expiry' not in usuarios.columns:
        usuarios['password_change_code_expiry'] = ''
    usuarios['password_change_code'] = usuarios['password_change_code'].fillna('').astype(str)
    usuarios['password_change_code_expiry'] = usuarios['password_change_code_expiry'].fillna('').astype(str)
    return usuarios


def codigo_cambio_password_expirado(usuario):
    expiry_raw = str(usuario.get('password_change_code_expiry', '') or '').strip()
    if not expiry_raw:
        return True

    return timestamp_expirado(expiry_raw)


def limpiar_codigo_cambio_password(usuarios, idx_usuario):
    usuarios.at[idx_usuario, 'password_change_code'] = ''
    usuarios.at[idx_usuario, 'password_change_code_expiry'] = ''


def generar_token_recuperacion():
    return secrets.token_urlsafe(32)


def obtener_usuario_por_token_recuperacion(usuarios, token):
    token = str(token or '').strip()
    if not token:
        return None

    usuarios['reset_token'] = usuarios['reset_token'].fillna('').astype(str)
    candidatos = usuarios[usuarios['reset_token'] == token]
    if candidatos.empty:
        return None

    idx = candidatos.index[0]
    return idx, candidatos.loc[idx]


def token_recuperacion_expirado(usuario):
    expiry_raw = str(usuario.get('reset_token_expiry', '') or '').strip()
    if not expiry_raw:
        return True

    return timestamp_expirado(expiry_raw)


def limpiar_token_recuperacion(usuarios, idx_usuario):
    usuarios.at[idx_usuario, 'reset_token'] = ''
    usuarios.at[idx_usuario, 'reset_token_expiry'] = ''


def enviar_codigo_registro(email, codigo):
    envio_ok = enviar_codigo_verificacion(
        email,
        codigo,
        tipo='registro',
        minutos_expiracion=REGISTER_CODE_EXP_MINUTES
    )
    if envio_ok:
        return True, 'Código enviado correctamente. Revisa tu correo.'
    return False, (
        'No fue posible enviar el código de verificación al correo indicado. '
        'Verifica la configuración SMTP del sistema e intenta nuevamente.'
    )


def validar_archivo_imagen(archivo):
    extension = extension_imagen(getattr(archivo, 'filename', ''))
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        return "Formato de imagen no permitido. Usa .jpg, .jpeg, .png, .gif o .webp."

    archivo.seek(0, os.SEEK_END)
    tamano = archivo.tell()
    archivo.seek(0)
    if tamano > MAX_IMAGE_SIZE_BYTES:
        return "La imagen excede el tamaño máximo permitido (3MB)."
    return None


def ruta_imagen_producto_absoluta(ruta_relativa):
    ruta = str(ruta_relativa or '').strip().replace('\\', '/').lstrip('/')
    if not ruta.startswith('img/Empresa/'):
        return ''

    base = os.path.abspath(os.path.join('static', 'img', 'Empresa'))
    candidata = os.path.abspath(os.path.join('static', ruta))
    if os.path.commonpath([base, candidata]) != base:
        return ''
    return candidata


def listar_archivos_galeria_producto(id_producto):
    carpeta_destino = os.path.join('static', 'img', 'Empresa')
    prefijo = f'producto_{id_producto}_'
    resultados = []

    if not os.path.isdir(carpeta_destino):
        return resultados

    for nombre in os.listdir(carpeta_destino):
        if not nombre.lower().startswith(prefijo.lower()):
            continue
        if extension_imagen(nombre) not in ALLOWED_IMAGE_EXTENSIONS:
            continue
        sufijo = nombre[len(prefijo):]
        posicion = sufijo.split('.', 1)[0]
        if not posicion.isdigit():
            continue
        resultados.append((int(posicion), nombre))

    resultados.sort(key=lambda item: item[0])
    return resultados


def migrar_legacy_a_galeria(id_producto):
    if listar_archivos_galeria_producto(id_producto):
        return

    carpeta_destino = os.path.join('static', 'img', 'Empresa')
    if not os.path.isdir(carpeta_destino):
        return

    for extension in sorted(ALLOWED_IMAGE_EXTENSIONS):
        legacy_name = f'producto_{id_producto}.{extension}'
        legacy_path = os.path.join(carpeta_destino, legacy_name)
        if not os.path.exists(legacy_path):
            continue
        nuevo_nombre = f'producto_{id_producto}_1.{extension}'
        nuevo_path = os.path.join(carpeta_destino, nuevo_nombre)
        os.replace(legacy_path, nuevo_path)
        break


def limpiar_imagenes_producto(id_producto):
    carpeta_destino = os.path.join('static', 'img', 'Empresa')
    if not os.path.isdir(carpeta_destino):
        return

    prefijo_galeria = f'producto_{id_producto}_'.lower()
    prefijo_legacy = f'producto_{id_producto}.'.lower()
    for nombre in os.listdir(carpeta_destino):
        nombre_lower = nombre.lower()
        if not (nombre_lower.startswith(prefijo_galeria) or nombre_lower.startswith(prefijo_legacy)):
            continue
        if extension_imagen(nombre_lower) not in ALLOWED_IMAGE_EXTENSIONS:
            continue
        ruta = os.path.join(carpeta_destino, nombre)
        if os.path.isfile(ruta):
            os.remove(ruta)


def guardar_galeria_producto(id_producto, imagenes, reemplazar=True):
    carpeta_destino = os.path.join('static', 'img', 'Empresa')
    os.makedirs(carpeta_destino, exist_ok=True)

    imagenes_validas = [img for img in imagenes if img and str(getattr(img, 'filename', '')).strip()]
    if reemplazar:
        limpiar_imagenes_producto(id_producto)
        indice_inicial = 1
    else:
        migrar_legacy_a_galeria(id_producto)
        existentes = listar_archivos_galeria_producto(id_producto)
        indice_inicial = (existentes[-1][0] + 1) if existentes else 1

    rutas_guardadas = []
    for indice, imagen in enumerate(imagenes_validas, start=indice_inicial):
        extension = extension_imagen(imagen.filename)
        nombre_archivo = f'producto_{id_producto}_{indice}.{extension}'
        ruta_absoluta = os.path.join(carpeta_destino, nombre_archivo)
        imagen.save(ruta_absoluta)
        rutas_guardadas.append(f"img/Empresa/{nombre_archivo}".replace('\\', '/'))
    return rutas_guardadas


def obtener_galeria_producto(id_producto, imagen_principal=''):
    try:
        id_producto_int = int(float(id_producto))
    except (TypeError, ValueError):
        return [normalizar_imagen_url(imagen_principal)] if str(imagen_principal).strip() else []

    galeria = listar_archivos_galeria_producto(id_producto_int)
    rutas = [f"img/Empresa/{nombre}".replace('\\', '/') for _, nombre in galeria]

    if not rutas:
        carpeta_destino = os.path.join('static', 'img', 'Empresa')
        if str(imagen_principal).strip():
            rutas.append(normalizar_imagen_url(imagen_principal))
        else:
            for extension in sorted(ALLOWED_IMAGE_EXTENSIONS):
                legacy_name = f'producto_{id_producto_int}.{extension}'
                legacy_path = os.path.join(carpeta_destino, legacy_name)
                if os.path.exists(legacy_path):
                    rutas.append(f"img/Empresa/{legacy_name}")
                    break
    return rutas


def cargar_productos_df():
    productos = read_table_df('producto')
    for column in PRODUCTO_COLUMNS:
        if column not in productos.columns:
            if column in {'eliminado', 'destacado_dashboard'}:
                productos[column] = False
            elif column in {'precio'}:
                productos[column] = 0.0
            elif column in {'stock', 'id_categoria'}:
                productos[column] = 0
            else:
                productos[column] = ''

    productos = normalizar_imagenes_productos(productos)
    productos['id_producto'] = pd.to_numeric(productos['id_producto'], errors='coerce')
    productos['precio'] = pd.to_numeric(productos['precio'], errors='coerce').fillna(0.0).astype(float)
    productos['stock'] = pd.to_numeric(productos['stock'], errors='coerce').fillna(0).astype(int)
    productos['id_categoria'] = pd.to_numeric(productos['id_categoria'], errors='coerce').fillna(0).astype(int)
    # Normalizar bandera eliminado aun si viene como texto ('t','f','true','false') o nulos.
    def _to_bool(valor: Any) -> bool:
        if pd.isna(valor):
            return False
        if isinstance(valor, (int, float, bool)):
            return bool(valor)
        if isinstance(valor, str):
            return valor.strip().lower() in {'true', 't', '1', 'yes', 'si', 'y'}
        return bool(valor)

    productos['eliminado'] = productos['eliminado'].apply(_to_bool)
    if 'destacado_dashboard' in productos.columns:
        productos['destacado_dashboard'] = productos['destacado_dashboard'].apply(_to_bool)
    else:
        productos['destacado_dashboard'] = False
    productos['fuerza'] = productos['fuerza'].fillna('').astype(str)
    productos['intendencia'] = productos['intendencia'].fillna('').astype(str)
    productos['nombre'] = productos['nombre'].fillna('').astype(str)
    productos['descripcion'] = productos['descripcion'].fillna('').astype(str)
    productos['imagen_url'] = productos['imagen_url'].fillna('').astype(str)
    return productos[PRODUCTO_COLUMNS]


def guardar_productos_df(productos):
    productos = productos.copy()
    for column in PRODUCTO_COLUMNS:
        if column not in productos.columns:
            if column in {'eliminado', 'destacado_dashboard'}:
                productos[column] = False
            elif column in {'precio'}:
                productos[column] = 0.0
            elif column in {'stock', 'id_categoria'}:
                productos[column] = 0
            else:
                productos[column] = ''

    productos['eliminado'] = productos['eliminado'].fillna(False).astype(bool)
    productos['destacado_dashboard'] = productos['destacado_dashboard'].fillna(False).astype(bool)
    replace_table_df('producto', productos[PRODUCTO_COLUMNS])


def cargar_productos_activos_df():
    productos = cargar_productos_df()
    return productos[productos['eliminado'] == False].copy()


def cargar_pedidos_df():
    pedidos = read_table_df('pedidos')
    for column in PEDIDO_COLUMNS:
        if column not in pedidos.columns:
            pedidos[column] = '' if column in {'id_usuario', 'fecha_pedido', 'estado', 'cliente_telefono', 'cliente_direccion'} else 0
    pedidos['id_pedido'] = pd.to_numeric(pedidos['id_pedido'], errors='coerce')
    pedidos['id_usuario'] = pedidos['id_usuario'].fillna('')
    pedidos['fecha_pedido'] = pedidos['fecha_pedido'].fillna('')
    pedidos['estado'] = pedidos['estado'].fillna('pendiente')
    pedidos['cliente_telefono'] = pedidos['cliente_telefono'].fillna('').astype(str)
    pedidos['cliente_direccion'] = pedidos['cliente_direccion'].fillna('').astype(str)
    return pedidos[PEDIDO_COLUMNS]


def guardar_pedidos_df(pedidos):
    pedidos = pedidos.copy()
    for column in PEDIDO_COLUMNS:
        if column not in pedidos.columns:
            pedidos[column] = '' if column in {'id_usuario', 'fecha_pedido', 'estado', 'cliente_telefono', 'cliente_direccion'} else 0
    replace_table_df('pedidos', pedidos[PEDIDO_COLUMNS])


def cargar_detalle_pedido_df():
    detalle = read_table_df('detalle_pedido')
    for column in DETALLE_PEDIDO_COLUMNS:
        if column not in detalle.columns:
            detalle[column] = '' if column == 'talla' else 0
    detalle['id_detalle'] = pd.to_numeric(detalle['id_detalle'], errors='coerce')
    detalle['id_pedido'] = pd.to_numeric(detalle['id_pedido'], errors='coerce')
    detalle['id_producto'] = pd.to_numeric(detalle['id_producto'], errors='coerce')
    detalle['cantidad'] = pd.to_numeric(detalle['cantidad'], errors='coerce').fillna(0)
    detalle['subtotal'] = pd.to_numeric(detalle['subtotal'], errors='coerce').fillna(0.0)
    if 'talla' in detalle.columns:
        detalle['talla'] = detalle['talla'].fillna('').astype(str)
    return detalle[DETALLE_PEDIDO_COLUMNS]


def guardar_detalle_pedido_df(detalle):
    detalle = detalle.copy()
    for column in DETALLE_PEDIDO_COLUMNS:
        if column not in detalle.columns:
            detalle[column] = '' if column == 'talla' else 0
    replace_table_df('detalle_pedido', detalle[DETALLE_PEDIDO_COLUMNS])


def cargar_pagos_df():
    pagos = read_table_df('pagos')
    defaults = {
        'id_pago': 0,
        'id_pedido': 0,
        'monto': 0.0,
        'metodo_pago': '',
        'fecha_pago': '',
        'estado_pago': '',
        'id_promo': '',
        'codigo_promo': '',
        'tipo_descuento': '',
        'valor_descuento': 0.0,
        'monto_descuento': 0.0,
    }
    for column, default_value in defaults.items():
        if column not in pagos.columns:
            pagos[column] = default_value
    pagos['id_pago'] = pd.to_numeric(pagos['id_pago'], errors='coerce')
    pagos['id_pedido'] = pd.to_numeric(pagos['id_pedido'], errors='coerce')
    pagos['monto'] = pd.to_numeric(pagos['monto'], errors='coerce').fillna(0.0)
    pagos['valor_descuento'] = pd.to_numeric(pagos['valor_descuento'], errors='coerce').fillna(0.0)
    pagos['monto_descuento'] = pd.to_numeric(pagos['monto_descuento'], errors='coerce').fillna(0.0)
    pagos['metodo_pago'] = pagos['metodo_pago'].fillna('')
    pagos['fecha_pago'] = pagos['fecha_pago'].fillna('')
    pagos['estado_pago'] = pagos['estado_pago'].fillna('')
    pagos['codigo_promo'] = pagos['codigo_promo'].fillna('')
    pagos['tipo_descuento'] = pagos['tipo_descuento'].fillna('')
    return pagos[PAGO_COLUMNS]


def guardar_pagos_df(pagos):
    pagos = pagos.copy()
    defaults = {
        'id_pago': 0,
        'id_pedido': 0,
        'monto': 0.0,
        'metodo_pago': '',
        'fecha_pago': '',
        'estado_pago': '',
        'id_promo': '',
        'codigo_promo': '',
        'tipo_descuento': '',
        'valor_descuento': 0.0,
        'monto_descuento': 0.0,
    }
    for column, default_value in defaults.items():
        if column not in pagos.columns:
            pagos[column] = default_value
    replace_table_df('pagos', pagos[PAGO_COLUMNS])


def cargar_productos_por_fuerza(fuerza):
    productos = cargar_productos_activos_df()
    fuerza_norm = fuerza.strip().lower()
    productos = productos[productos['fuerza'].str.strip().str.lower() == fuerza_norm]
    return productos.to_dict(orient='records')

def cargar_productos_por_intendencia(intendencia):
    productos = cargar_productos_activos_df()
    intendencia_norm = intendencia.strip().lower()
    productos = productos[productos['intendencia'].str.strip().str.lower() == intendencia_norm]
    return productos.to_dict(orient='records')

def parsear_fecha_promocion(valor: Any) -> Optional[date]:
    if valor is None:
        return None
    if isinstance(valor, date) and not isinstance(valor, datetime):
        return valor

    convertido = pd.to_datetime(valor, errors='coerce')
    if pd.isna(convertido):
        return None
    try:
        return convertido.date()
    except Exception:
        return None


def estado_vigencia_promocion(promo: Mapping[str, Any], fecha_ref: Optional[date] = None) -> str:
    if fecha_ref is None:
        fecha_ref = datetime.now().date()
    inicio = parsear_fecha_promocion(promo.get('fecha_inicio', ''))
    fin = parsear_fecha_promocion(promo.get('fecha_fin', ''))
    if inicio is not None and fecha_ref < inicio:
        return 'programada'
    if fin is not None and fecha_ref > fin:
        return 'vencida'
    return 'vigente'


def promocion_esta_aplicable(promo: Mapping[str, Any], fecha_ref: Optional[date] = None) -> bool:
    return bool(promo.get('activo', False)) and estado_vigencia_promocion(promo, fecha_ref) == 'vigente'


def calcular_descuento_promocion(subtotal, promo):
    subtotal = max(0.0, float(subtotal))
    tipo = str(promo.get('tipo_descuento', 'porcentaje')).strip().lower()
    valor = float(pd.to_numeric(promo.get('valor_descuento', 0), errors='coerce') or 0)

    if tipo == 'valor_fijo':
        descuento = max(0.0, valor)
    else:
        descuento = subtotal * max(0.0, valor) / 100.0

    return min(descuento, subtotal)


def obtener_mejor_promocion_por_producto(
    productos_df: pd.DataFrame,
    promos_df: pd.DataFrame,
    fecha_ref: Optional[date] = None
):
    if fecha_ref is None:
        fecha_ref = datetime.now().date()

    productos_ref = productos_df.copy()
    if 'id_producto' not in productos_ref.columns:
        productos_ref['id_producto'] = pd.Series(dtype='int')
    if 'precio' not in productos_ref.columns:
        productos_ref['precio'] = 0.0

    productos_ref['id_producto'] = pd.to_numeric(productos_ref['id_producto'], errors='coerce')
    productos_ref['precio'] = pd.to_numeric(productos_ref['precio'], errors='coerce').fillna(0.0)

    precio_por_producto = {}
    for _, producto in productos_ref.iterrows():
        pid = pd.to_numeric(producto.get('id_producto'), errors='coerce')
        if pd.notna(pid):
            precio_por_producto[int(pid)] = float(pd.to_numeric(producto.get('precio', 0), errors='coerce') or 0.0)

    mejor_promo_por_producto = {}
    mejor_descuento_por_producto = {}
    for _, row in promos_df.iterrows():
        promo = row.to_dict()
        if not promocion_esta_aplicable(promo, fecha_ref):
            continue

        id_producto_promo = pd.to_numeric(promo.get('id_producto'), errors='coerce')
        if pd.isna(id_producto_promo):
            continue

        id_prod_int = int(id_producto_promo)
        if id_prod_int not in precio_por_producto:
            continue

        precio_ref = precio_por_producto.get(id_prod_int, 0.0)
        descuento_actual = calcular_descuento_promocion(precio_ref, promo)
        if descuento_actual > mejor_descuento_por_producto.get(id_prod_int, -1):
            mejor_descuento_por_producto[id_prod_int] = descuento_actual
            mejor_promo_por_producto[id_prod_int] = promo

    return mejor_promo_por_producto


def buscar_promocion_por_codigo(promos_df, codigo, fecha_ref=None):
    codigo_norm = str(codigo).strip().upper()
    if not codigo_norm:
        return None

    candidatos = promos_df[promos_df['codigo'].astype(str).str.upper() == codigo_norm]
    if candidatos.empty:
        return None

    for _, row in candidatos.iterrows():
        promo = row.to_dict()
        if promocion_esta_aplicable(promo, fecha_ref):
            return promo
    return None


def registrar_actividad(accion):
    registros = cargar_registros_df()
    nuevo_id = int(pd.to_numeric(registros['id_registro'], errors='coerce').max() + 1) if not registros.empty else 1
    actor = session.get('usuario', 'admin')
    nuevo_registro = {
        'id_registro': nuevo_id,
        'id_usuario': actor,
        'accion': accion,
        'fecha_accion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    registros = pd.concat([registros, pd.DataFrame([nuevo_registro])], ignore_index=True)
    guardar_registros_df(registros)


def obtener_nombre_sesion():
    nombre = str(session.get('nombre', '')).strip()
    if nombre:
        return nombre

    email = str(session.get('usuario', '')).strip()
    if not email:
        return 'Administrador'

    usuarios = cargar_usuarios_df()
    if {'email', 'nombre'}.issubset(usuarios.columns):
        coincidencia = usuarios[usuarios['email'].astype(str).str.lower() == email.lower()]
        if not coincidencia.empty:
            nombre_excel = str(coincidencia.iloc[0].get('nombre', '')).strip()
            if nombre_excel:
                session['nombre'] = nombre_excel
                return nombre_excel

    return email.split('@')[0] if '@' in email else email


def formatear_cop(valor):
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        numero = 0.0
    valor_formateado = f"{numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{CURRENCY_CODE} {valor_formateado}"


def _construir_datos_recibo_pos(id_pedido):
    detalle = cargar_detalle_pedido_df()
    if detalle.empty:
        raise ValueError(f'No existe detalle para el pedido #{id_pedido}.')

    detalle = detalle.copy()
    detalle['id_pedido'] = pd.to_numeric(detalle['id_pedido'], errors='coerce')
    detalle = detalle[detalle['id_pedido'] == int(id_pedido)]
    if detalle.empty:
        raise ValueError(f'No existe detalle para el pedido #{id_pedido}.')

    if 'id_detalle' in detalle.columns:
        detalle = detalle.sort_values(by='id_detalle', ascending=True, na_position='last')

    productos_df = cargar_productos_df()
    productos_df['id_producto'] = pd.to_numeric(productos_df['id_producto'], errors='coerce')
    productos_map = {}
    for _, row in productos_df.iterrows():
        if pd.notna(row.get('id_producto')):
            productos_map[int(row['id_producto'])] = str(row.get('nombre', '')).strip()

    items = []
    for _, row in detalle.iterrows():
        id_producto = int(pd.to_numeric(row.get('id_producto'), errors='coerce') or 0)
        cantidad = int(pd.to_numeric(row.get('cantidad', 0), errors='coerce') or 0)
        subtotal = float(pd.to_numeric(row.get('subtotal', 0), errors='coerce') or 0)
        talla = str(row.get('talla', '') or '').strip().upper()
        if talla not in TALLAS_OPCIONES:
            talla = '-'
        if cantidad <= 0:
            cantidad = 1
        valor_unitario = subtotal / cantidad if cantidad else subtotal
        descripcion = productos_map.get(id_producto, f"Producto #{id_producto}")
        items.append({
            'cantidad': cantidad,
            'talla': talla,
            'descripcion': descripcion,
            'valor_unitario': float(valor_unitario),
            'total': float(subtotal)
        })

    pagos = cargar_pagos_df()
    pagos['id_pedido'] = pd.to_numeric(pagos['id_pedido'], errors='coerce')
    pago_pedido = pagos[pagos['id_pedido'] == int(id_pedido)].copy()
    pago_row = None
    if not pago_pedido.empty:
        pago_pedido['id_pago'] = pd.to_numeric(pago_pedido['id_pago'], errors='coerce')
        pago_pedido = pago_pedido.sort_values(by='id_pago', ascending=False, na_position='last')
        pago_row = pago_pedido.iloc[0]

    metodo_pago = ''
    monto_total = float(sum(item['total'] for item in items))
    descuento_total = 0.0
    if pago_row is not None:
        metodo_pago = str(pago_row.get('metodo_pago', '')).strip().lower()
        monto_total = float(pd.to_numeric(pago_row.get('monto', monto_total), errors='coerce') or monto_total)
        descuento_total = float(pd.to_numeric(pago_row.get('monto_descuento', 0), errors='coerce') or 0)

    metodo_label = {
        'efectivo': 'Efectivo',
        'tarjeta': 'Tarjeta',
        'transferencia': 'Transferencia',
        'qr': 'QR'
    }.get(metodo_pago, metodo_pago.title() if metodo_pago else 'No especificado')

    pedidos = cargar_pedidos_df()
    pedidos['id_pedido'] = pd.to_numeric(pedidos['id_pedido'], errors='coerce')
    pedido_row = pedidos[pedidos['id_pedido'] == int(id_pedido)]
    fecha_compra = datetime.now()
    if not pedido_row.empty:
        fecha_raw = pedido_row.iloc[0].get('fecha_pedido', '')
        fecha_parseada = pd.to_datetime(fecha_raw, errors='coerce')
        if pd.notna(fecha_parseada):
            fecha_compra = fecha_parseada.to_pydatetime()

    cliente_nombre = ''
    cliente_correo = ''
    cliente_documento = ''
    cliente_telefono = ''
    ultimo_recibo = session.get('ultimo_recibo_pos', {})
    if str(ultimo_recibo.get('id_pedido', '')) == str(id_pedido):
        cliente_nombre = str(ultimo_recibo.get('cliente_nombre', '')).strip()
        cliente_correo = str(ultimo_recibo.get('cliente_correo', '')).strip()
        cliente_documento = str(ultimo_recibo.get('cliente_documento', '')).strip()
        cliente_telefono = str(ultimo_recibo.get('cliente_telefono', '')).strip()

    subtotal_bruto = max(0.0, monto_total + descuento_total)
    return {
        'id_pedido': int(id_pedido),
        'fecha_compra': fecha_compra,
        'metodo_pago': metodo_label,
        'items': items,
        'subtotal_bruto': float(subtotal_bruto),
        'descuento_total': float(descuento_total),
        'total': float(monto_total),
        'cliente_nombre': cliente_nombre,
        'cliente_correo': cliente_correo,
        'cliente_documento': cliente_documento,
        'cliente_telefono': cliente_telefono
    }


def _particionar_items_recibo(items, max_items_por_pagina=8):
    items = list(items or [])
    if not items:
        return [[]]
    return [items[i:i + max_items_por_pagina] for i in range(0, len(items), max_items_por_pagina)]


def _construir_contexto_recibo_pos(id_pedido):
    datos = _construir_datos_recibo_pos(id_pedido)
    fecha_txt = datos['fecha_compra'].strftime('%d / %m / %Y')
    hora_txt = datos['fecha_compra'].strftime('%H:%M:%S')

    items_formateados = []
    for item in datos.get('items', []):
        items_formateados.append({
            'cantidad': int(pd.to_numeric(item.get('cantidad', 0), errors='coerce') or 0),
            'talla': str(item.get('talla', '-') or '-').strip().upper() or '-',
            'descripcion': str(item.get('descripcion', '') or '').strip(),
            'valor_unitario': formatear_cop(item.get('valor_unitario', 0)),
            'total': formatear_cop(item.get('total', 0))
        })

    paginas_items = _particionar_items_recibo(items_formateados, max_items_por_pagina=8)
    total_paginas = len(paginas_items)
    paginas = []
    for i, items_pagina in enumerate(paginas_items, start=1):
        paginas.append({
            'numero': i,
            'detalle_items': items_pagina,
            'es_ultima': i == total_paginas
        })

    cliente_campos = []
    cliente_nombre = str(datos.get('cliente_nombre', '') or '').strip()
    cliente_doc = str(datos.get('cliente_documento', '') or '').strip()
    cliente_tel = str(datos.get('cliente_telefono', '') or '').strip()
    cliente_mail = str(datos.get('cliente_correo', '') or '').strip()
    if cliente_nombre:
        cliente_campos.append({'label': 'Nombre', 'valor': cliente_nombre})
    if cliente_doc:
        cliente_campos.append({'label': 'Documento', 'valor': cliente_doc})
    if cliente_tel:
        cliente_campos.append({'label': 'Telefono', 'valor': cliente_tel})
    if cliente_mail:
        cliente_campos.append({'label': 'Correo', 'valor': cliente_mail})

    return {
        'id_pedido': int(datos['id_pedido']),
        'codigo_recibo': f"POS-{int(datos['id_pedido']):06d}",
        'fecha_txt': fecha_txt,
        'hora_txt': hora_txt,
        'paginas': paginas,
        'total_paginas': total_paginas,
        'cliente_campos': cliente_campos,
        'cantidad_items': len(items_formateados),
        'metodo_pago': str(datos.get('metodo_pago', 'No especificado') or 'No especificado'),
        'subtotal_txt': formatear_cop(datos.get('subtotal_bruto', 0)),
        'descuento_txt': formatear_cop(datos.get('descuento_total', 0)),
        'total_txt': formatear_cop(datos.get('total', 0)),
        'observaciones': (
            "Prendas confeccionadas bajo estandares de resistencia, "
            "durabilidad y ergonomia para uso tactico y operativo."
        )
    }


def _cargar_css_recibo_pos():
    ruta_css = os.path.join(app.root_path, 'static', 'css', 'administrador', 'admin_pos_recibo.css')
    if not os.path.exists(ruta_css):
        return ''
    try:
        with open(ruta_css, 'r', encoding='utf-8') as f_css:
            return f_css.read()
    except UnicodeDecodeError:
        with open(ruta_css, 'r', encoding='latin-1') as f_css:
            return f_css.read()


def _cargar_marca_agua_recibo_src():
    carpeta = os.path.join(app.root_path, 'static', 'img', 'Pagina')
    candidatos = ('recibo.png', 'nachoher_fondo_logo.jpg', 'logo.jpeg')
    mime_por_ext = {
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg'
    }

    for nombre in candidatos:
        ruta = Path(carpeta) / nombre
        if not ruta.exists():
            continue
        try:
            return ruta.resolve().as_uri()
        except ValueError:
            pass

        ext = ruta.suffix.lower()
        mime = mime_por_ext.get(ext, 'application/octet-stream')
        with open(ruta, 'rb') as f_img:
            contenido_b64 = base64.b64encode(f_img.read()).decode('ascii')
        return f"data:{mime};base64,{contenido_b64}"
    return ''


def render_html_recibo_pos(id_pedido):
    contexto = _construir_contexto_recibo_pos(id_pedido)
    css_content = _cargar_css_recibo_pos()
    return render_template(
        'Administrador/Sistema POS/recibo_pos_pdf.html',
        css_content=css_content,
        watermark_image_src=_cargar_marca_agua_recibo_src(),
        **contexto
    )


def _buscar_msedge():
    candidatos = []

    edge_env = str(os.environ.get('EDGE_PATH', '')).strip()
    if edge_env:
        candidatos.append(edge_env)

    for exe in ('msedge.exe', 'msedge'):
        encontrado = shutil.which(exe)
        if encontrado:
            candidatos.append(encontrado)

    candidatos.extend([
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
    ])

    for ruta in candidatos:
        if ruta and os.path.exists(ruta):
            return ruta
    return None


def generar_pdf_recibo_pos(id_pedido):
    html = render_html_recibo_pos(id_pedido)
    ruta_edge = _buscar_msedge()
    if not ruta_edge:
        raise RuntimeError(
            "No se encontro Microsoft Edge para convertir el recibo a PDF. "
            "Configura EDGE_PATH o instala Edge."
        )

    tmp_path = Path(app.root_path) / f"recibo_pos_tmp_{id_pedido}_{os.getpid()}_{int(datetime.now().timestamp() * 1000)}"
    tmp_path.mkdir(parents=True, exist_ok=True)

    try:
        html_path = tmp_path / f"recibo_pos_{id_pedido}.html"
        pdf_path = tmp_path / f"recibo_pos_{id_pedido}.pdf"
        user_data_dir = tmp_path

        html_path.write_text(html, encoding='utf-8')

        comando = [
            ruta_edge,
            "--headless",
            "--disable-gpu",
            "--no-pdf-header-footer",
            "--print-to-pdf-no-header",
            "--no-sandbox",
            "--disable-breakpad",
            "--no-first-run",
            "--no-default-browser-check",
            "--allow-file-access-from-files",
            f"--user-data-dir={str(user_data_dir)}",
            f"--print-to-pdf={str(pdf_path)}",
            html_path.resolve().as_uri()
        ]

        resultado = subprocess.run(
            comando,
            capture_output=True,
            text=True,
            timeout=90
        )

        if resultado.returncode != 0 or not pdf_path.exists():
            detalle_error = (resultado.stderr or resultado.stdout or '').strip()
            if detalle_error:
                detalle_error = detalle_error.splitlines()[0]
            raise RuntimeError(f"No se pudo convertir el recibo a PDF. {detalle_error}")

        return pdf_path.read_bytes()
    finally:
        shutil.rmtree(tmp_path, ignore_errors=True)


def obtener_productos_agotados():
    productos = cargar_productos_df()
    if productos.empty:
        return []

    agotados = productos[(productos['eliminado'] == False) & (productos['stock'] <= 0)].copy()
    if agotados.empty:
        return []

    agotados = agotados.sort_values(by='nombre', na_position='last')
    return agotados[['id_producto', 'nombre', 'stock']].to_dict(orient='records')


def normalizar_carrito_por_stock(carrito):
    if not carrito:
        return [], []

    productos = cargar_productos_df()
    if 'id_producto' not in productos.columns:
        return carrito, []

    inventario = {}
    for _, row in productos.iterrows():
        if pd.isna(row['id_producto']):
            continue
        inventario[int(row['id_producto'])] = {
            'nombre': str(row.get('nombre', '')).strip(),
            'precio': float(row.get('precio', 0.0)),
            'stock': int(row.get('stock', 0)),
            'eliminado': bool(row.get('eliminado', False))
        }

    carrito_limpio = []
    cambios = []

    for item in carrito:
        id_producto = pd.to_numeric(item.get('id_producto'), errors='coerce')
        cantidad_item = pd.to_numeric(item.get('cantidad', 0), errors='coerce')
        if pd.isna(id_producto) or pd.isna(cantidad_item):
            continue

        id_producto = int(id_producto)
        cantidad = max(0, int(cantidad_item))
        if cantidad <= 0:
            continue

        referencia = inventario.get(id_producto)
        nombre_item = str(item.get('nombre', f'ID {id_producto}'))
        if referencia is None or referencia['eliminado'] or referencia['stock'] <= 0:
            cambios.append(f'Se removio "{nombre_item}" porque ya no tiene stock.')
            continue

        stock_actual = int(referencia['stock'])
        cantidad_final = min(cantidad, stock_actual)
        if cantidad_final < cantidad:
            cambios.append(f'Se ajusto "{nombre_item}" de {cantidad} a {cantidad_final} por stock disponible.')

        precio_final = float(referencia['precio']) if referencia['precio'] > 0 else float(item.get('precio', 0))
        talla_item = str(item.get('talla', '')).strip()
        carrito_limpio.append({
            'id_producto': id_producto,
            'nombre': referencia['nombre'] or nombre_item,
            'cantidad': cantidad_final,
            'precio': precio_final,
            'subtotal': float(precio_final) * cantidad_final,
            'talla': talla_item
        })

    return carrito_limpio, cambios


def _obtener_carrito_guardado_usuario(email):
    email_norm = normalizar_email(email)
    if not email_norm:
        return []

    try:
        with engine.connect() as conn:
            row = conn.execute(
                sa.text("SELECT carrito_json FROM carrito_usuario WHERE email = :email"),
                {"email": email_norm}
            ).fetchone()
        if not row:
            return []

        carrito = json.loads(str(row[0] or '[]'))
        if isinstance(carrito, list):
            return carrito
        return []
    except Exception as exc:
        print(f"No se pudo cargar carrito guardado para {email_norm}: {exc}")
        return []


def _guardar_carrito_guardado_usuario(email, carrito):
    email_norm = normalizar_email(email)
    if not email_norm:
        return

    payload = carrito if isinstance(carrito, list) else []
    try:
        with engine.begin() as conn:
            conn.execute(
                sa.text(
                    """
                    INSERT INTO carrito_usuario (email, carrito_json, updated_at)
                    VALUES (:email, :carrito_json, NOW())
                    ON CONFLICT (email)
                    DO UPDATE SET
                        carrito_json = EXCLUDED.carrito_json,
                        updated_at = NOW()
                    """
                ),
                {
                    "email": email_norm,
                    "carrito_json": json.dumps(payload, ensure_ascii=False)
                }
            )
    except Exception as exc:
        print(f"No se pudo guardar carrito para {email_norm}: {exc}")


def _sincronizar_carrito_usuario_desde_sesion():
    if session.get('rol') != 'normal':
        return
    _guardar_carrito_guardado_usuario(session.get('usuario', ''), session.get('carrito', []))


def _obtener_carrito_sesion_usuario():
    if session.get('rol') != 'normal':
        return []

    carrito_actual = session.get('carrito', None)
    if not isinstance(carrito_actual, list):
        carrito_actual = _obtener_carrito_guardado_usuario(session.get('usuario', ''))
        session['carrito'] = carrito_actual
        session.modified = True

    return carrito_actual


def _enriquecer_carrito_con_imagenes(carrito):
    if not isinstance(carrito, list) or not carrito:
        return []

    productos = cargar_productos_df()
    referencias = {}
    if not productos.empty:
        for _, row in productos.iterrows():
            id_producto = pd.to_numeric(row.get('id_producto'), errors='coerce')
            if pd.isna(id_producto):
                continue
            referencias[int(id_producto)] = {
                'imagen_url': str(row.get('imagen_url', '') or '').strip(),
                'descripcion': str(row.get('descripcion', '') or '').strip()
            }

    carrito_enriquecido = []
    for item in carrito:
        item_enriquecido = dict(item) if isinstance(item, dict) else {}
        id_producto = pd.to_numeric(item_enriquecido.get('id_producto'), errors='coerce')
        if pd.isna(id_producto):
            item_enriquecido['imagen_url'] = normalizar_imagen_url(item_enriquecido.get('imagen_url', ''))
            carrito_enriquecido.append(item_enriquecido)
            continue

        id_producto = int(id_producto)
        ref = referencias.get(id_producto, {})
        imagen_actual = str(item_enriquecido.get('imagen_url', '') or '').strip()
        imagen_principal = str(ref.get('imagen_url', '') or '').strip()
        galeria = obtener_galeria_producto(id_producto, imagen_principal)
        imagen_final = ''
        if galeria:
            imagen_final = normalizar_imagen_url(galeria[0])
        elif imagen_actual:
            imagen_final = normalizar_imagen_url(imagen_actual)

        item_enriquecido['imagen_url'] = imagen_final
        if not str(item_enriquecido.get('descripcion', '') or '').strip() and ref.get('descripcion'):
            item_enriquecido['descripcion'] = ref.get('descripcion')
        carrito_enriquecido.append(item_enriquecido)

    return carrito_enriquecido


def _hash_carrito_checkout(carrito):
    payload = carrito if isinstance(carrito, list) else []
    texto = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha256(texto.encode('utf-8')).hexdigest()


def _stripe_checkout_guardar_creado(session_id, usuario_email, codigo_promo, carrito, cart_hash, total_esperado):
    carrito_json = "[]"
    if isinstance(carrito, list):
        carrito_json = json.dumps(carrito, ensure_ascii=False)
    with engine.begin() as conn:
        conn.execute(
            sa.text(
                """
                INSERT INTO stripe_checkout (
                    session_id, usuario_email, codigo_promo, carrito_json, cart_hash, total_esperado,
                    estado, id_pedido, created_at, updated_at
                )
                VALUES (
                    :session_id, :usuario_email, :codigo_promo, :carrito_json, :cart_hash, :total_esperado,
                    'creado', NULL, NOW(), NOW()
                )
                ON CONFLICT (session_id)
                DO UPDATE SET
                    usuario_email = EXCLUDED.usuario_email,
                    codigo_promo = EXCLUDED.codigo_promo,
                    carrito_json = EXCLUDED.carrito_json,
                    cart_hash = EXCLUDED.cart_hash,
                    total_esperado = EXCLUDED.total_esperado,
                    estado = 'creado',
                    id_pedido = NULL,
                    updated_at = NOW()
                """
            ),
            {
                "session_id": str(session_id or "").strip(),
                "usuario_email": normalizar_email(usuario_email),
                "codigo_promo": str(codigo_promo or "").strip().upper(),
                "carrito_json": carrito_json,
                "cart_hash": str(cart_hash or "").strip(),
                "total_esperado": float(total_esperado or 0.0),
            },
        )


def _stripe_checkout_obtener(session_id):
    sid = str(session_id or "").strip()
    if not sid:
        return None
    with engine.connect() as conn:
        row = conn.execute(
            sa.text(
                """
                SELECT session_id, usuario_email, codigo_promo, carrito_json, cart_hash,
                       total_esperado, estado, id_pedido
                FROM stripe_checkout
                WHERE session_id = :session_id
                """
            ),
            {"session_id": sid},
        ).mappings().first()
    return dict(row) if row else None


def _stripe_checkout_marcar_estado(session_id, estado, id_pedido=None):
    sid = str(session_id or "").strip()
    if not sid:
        return
    with engine.begin() as conn:
        conn.execute(
            sa.text(
                """
                UPDATE stripe_checkout
                SET estado = :estado,
                    id_pedido = :id_pedido,
                    updated_at = NOW()
                WHERE session_id = :session_id
                """
            ),
            {
                "session_id": sid,
                "estado": str(estado or "").strip() or "creado",
                "id_pedido": id_pedido,
            },
        )


def _stripe_obj_get(data, key, default=None):
    if isinstance(data, Mapping):
        return data.get(key, default)
    try:
        return data.get(key, default)
    except Exception:
        return getattr(data, key, default)


def _stripe_checkout_cargar_carrito(registro_checkout):
    raw = str((registro_checkout or {}).get("carrito_json", "") or "").strip()
    if not raw:
        return []
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return []
    return data if isinstance(data, list) else []


def _obtener_usuario_por_email(email):
    email_norm = normalizar_email(email)
    if not email_norm:
        return None
    usuarios = cargar_usuarios_df()
    usuarios['email'] = usuarios['email'].astype(str).str.strip().str.lower()
    candidatos = usuarios[usuarios['email'] == email_norm]
    if candidatos.empty:
        return None
    return candidatos.iloc[0].to_dict()


@app.template_filter('cop')
def cop_filter(valor):
    return formatear_cop(valor)


@app.context_processor
def inyectar_nombre_admin():
    admin_nombre = ''
    if session.get('rol') == 'admin':
        admin_nombre = obtener_nombre_sesion()
    return {
        'admin_nombre': admin_nombre,
        'currency_code': CURRENCY_CODE,
        'currency_name': CURRENCY_NAME
    }

@app.route('/')
def home():
    productos = cargar_productos_activos_df()
    lista_productos = productos.to_dict(orient='records')

    # Promoción vigente por producto para mostrar en catálogo
    promos = cargar_promociones_df()
    hoy = datetime.now().date()
    mejor_promo_por_producto = obtener_mejor_promocion_por_producto(productos, promos, hoy)

    for producto in lista_productos:
        precio_base = float(pd.to_numeric(producto.get('precio', 0), errors='coerce') or 0)
        promo = mejor_promo_por_producto.get(int(producto.get('id_producto', 0)))
        if promo:
            descuento = calcular_descuento_promocion(precio_base, promo)
            producto['precio_original'] = precio_base
            producto['precio_con_descuento'] = max(0.0, precio_base - descuento)
            producto['promo_activa'] = True
            producto['promo_nombre'] = promo.get('nombre', '')
            if promo.get('tipo_descuento') == 'valor_fijo':
                producto['promo_etiqueta'] = f"-{formatear_cop(promo.get('valor_descuento', 0))}"
            else:
                valor_pct = float(pd.to_numeric(promo.get('valor_descuento', 0), errors='coerce') or 0)
                producto['promo_etiqueta'] = f"-{valor_pct:g}%"
        else:
            producto['precio_original'] = precio_base
            producto['precio_con_descuento'] = precio_base
            producto['promo_activa'] = False
            producto['promo_nombre'] = ''
            producto['promo_etiqueta'] = ''
        galeria = obtener_galeria_producto(
            int(producto.get('id_producto', 0)),
            producto.get('imagen_url', '')
        )
        if not galeria:
            galeria = [producto.get('imagen_url', '')]
        while len(galeria) < 5:
            galeria.append(galeria[0])
        producto['galeria_dashboard'] = galeria[:5]

    productos_destacados_ejercito = [p for p in lista_productos if bool(p.get('destacado_dashboard', False)) and p.get('fuerza') == 'Ejercito'][:5]
    productos_destacados_policia = [p for p in lista_productos if bool(p.get('destacado_dashboard', False)) and p.get('fuerza') == 'Policia'][:5]
    productos_destacados_armada = [p for p in lista_productos if bool(p.get('destacado_dashboard', False)) and p.get('fuerza') == 'Armada'][:5]

    return render_template(
        'Usuarios/Autenticacion/login.html',
        productos=lista_productos,
        productos_destacados_ejercito=productos_destacados_ejercito,
        productos_destacados_policia=productos_destacados_policia,
        productos_destacados_armada=productos_destacados_armada
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('Usuarios/Autenticacion/login_form.html')
    usuarios = cargar_usuarios_df()  # leer cada vez
    email = normalizar_email(request.form.get('email', ''))
    password = request.form.get('password', '')

    usuarios['email'] = usuarios['email'].astype(str).str.strip()
    candidatos = usuarios[usuarios['email'].str.lower() == email]
    if candidatos.empty:
        flash('Correo equivocado.', 'email_error')
        return render_template('Usuarios/Autenticacion/login_form.html'), 401

    idx_usuario = candidatos.index[0]
    usuario = candidatos.loc[idx_usuario]

    if not password_coincide(usuario.get('password_hash', ''), password):
        flash('Contraseña incorrecta.', 'password_error')
        return render_template('Usuarios/Autenticacion/login_form.html'), 401

    # Migracion transparente: si la contraseña antigua estaba en texto plano, se hashea al iniciar sesión.
    password_guardado = str(usuario.get('password_hash', '') or '')
    if not password_esta_hasheado(password_guardado):
        usuarios.at[idx_usuario, 'password_hash'] = crear_hash_password(password)
        guardar_usuarios_df(usuarios)

    estado = str(usuario.get('estado', 'activo')).strip().lower()
    if estado != 'activo':
        return "Tu usuario está inactivo. Contacta al administrador."

    rol = usuario['rol']
    id_usuario = usuario['id_usuario']
    nombre = str(usuario.get('nombre', '')).strip()
    email_sesion = str(usuario.get('email', email)).strip() or email

    session.permanent = True  # evita perder la sesión al navegar entre pestañas
    session['usuario'] = email_sesion
    session['id_usuario'] = int(id_usuario)
    session['rol'] = rol
    session['nombre'] = nombre

    if rol == 'normal':
        session['carrito'] = _obtener_carrito_guardado_usuario(email_sesion)
    else:
        session.pop('carrito', None)

    registrar_actividad("Inicio de sesión exitoso")

    if rol == 'admin':
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('user_dashboard'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'GET':
        return render_template(
            'Usuarios/Autenticacion/forgot_password.html',
            reset_minutes=PASSWORD_RESET_EXP_MINUTES
        )

    email = normalizar_email(request.form.get('email', ''))
    if not email_es_valido(email):
        flash('Debes ingresar un correo electrónico válido.', 'danger')
        return render_template(
            'Usuarios/Autenticacion/forgot_password.html',
            reset_minutes=PASSWORD_RESET_EXP_MINUTES
        ), 400

    usuarios = cargar_usuarios_df()
    usuarios['email'] = usuarios['email'].astype(str).str.strip().str.lower()
    candidatos = usuarios[usuarios['email'] == email]

    if not candidatos.empty:
        idx_usuario = candidatos.index[0]
        token = generar_token_recuperacion()
        expiry_at = datetime.now() + timedelta(minutes=PASSWORD_RESET_EXP_MINUTES)

        usuarios.at[idx_usuario, 'reset_token'] = token
        usuarios.at[idx_usuario, 'reset_token_expiry'] = expiry_at.strftime('%Y-%m-%d %H:%M:%S')
        guardar_usuarios_df(usuarios)

        enlace = url_for('reset_password', token=token, _external=True)
        envio_ok = enviar_recuperacion_password(
            email=email,
            enlace_recuperacion=enlace,
            minutos_expiracion=PASSWORD_RESET_EXP_MINUTES
        )

        if envio_ok:
            registrar_actividad(f"Enlace de recuperación enviado a {email}")
        else:
            limpiar_token_recuperacion(usuarios, idx_usuario)
            guardar_usuarios_df(usuarios)
            print(f"No fue posible enviar el correo de recuperación para: {email}")

    flash(
        'Si el correo existe en el sistema, te enviamos un enlace para restablecer tu contraseña.',
        'info'
    )
    return redirect(url_for('forgot_password'))


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    token = str(token or '').strip()
    usuarios = cargar_usuarios_df()

    encontrado = obtener_usuario_por_token_recuperacion(usuarios, token)
    if not encontrado:
        flash('El enlace de recuperación no es válido o ya fue utilizado.', 'danger')
        return redirect(url_for('forgot_password'))

    idx_usuario, usuario = encontrado
    if token_recuperacion_expirado(usuario):
        limpiar_token_recuperacion(usuarios, idx_usuario)
        guardar_usuarios_df(usuarios)
        flash('El enlace de recuperación expiró. Solicita uno nuevo.', 'warning')
        return redirect(url_for('forgot_password'))

    if request.method == 'GET':
        return render_template(
            'Usuarios/Autenticacion/reset_password.html',
            token=token
        )

    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')

    if not new_password or not confirm_password:
        flash('Debes completar todos los campos.', 'danger')
        return render_template('Usuarios/Autenticacion/reset_password.html', token=token), 400

    if new_password != confirm_password:
        flash('Las contraseñas no coinciden.', 'danger')
        return render_template('Usuarios/Autenticacion/reset_password.html', token=token), 400

    if not password_cumple_estandares(new_password):
        flash(
            'La contraseña debe tener mínimo 8 caracteres, mayúscula, minúscula, número y carácter especial.',
            'danger'
        )
        return render_template('Usuarios/Autenticacion/reset_password.html', token=token), 400

    usuarios.at[idx_usuario, 'password_hash'] = crear_hash_password(new_password)
    limpiar_token_recuperacion(usuarios, idx_usuario)
    guardar_usuarios_df(usuarios)

    email_usuario = str(usuario.get('email', '')).strip()
    registrar_actividad(f"Contraseña restablecida por recuperación para {email_usuario}")
    flash('Contraseña actualizada correctamente. Ya puedes iniciar sesión.', 'success')
    return redirect(url_for('login'))

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    def _render_registro(nombre='', email=''):
        return render_template(
            'Usuarios/Autenticacion/registro.html',
            nombre_registro=nombre,
            email_registro=email,
            registro_code_minutes=REGISTER_CODE_EXP_MINUTES
        )

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        email = normalizar_email(request.form.get('email', ''))
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not nombre or not email or not password or not confirm_password:
            flash('Debes completar todos los campos del formulario.', 'danger')
            return _render_registro(nombre, email), 400

        if not email_es_valido(email):
            flash('Debes ingresar un correo electrónico válido.', 'danger')
            return _render_registro(nombre, email), 400

        if password != confirm_password:
            flash('Las contraseñas no coinciden.', 'danger')
            return _render_registro(nombre, email), 400

        if not password_cumple_estandares(password):
            flash(
                'La contraseña debe tener mínimo 8 caracteres, mayúscula, minúscula, número y carácter especial.',
                'danger'
            )
            return _render_registro(nombre, email), 400

        usuarios = cargar_usuarios_df()
        usuarios['email'] = usuarios['email'].astype(str).str.strip().str.lower()

        if email in usuarios['email'].values:
            flash(
                'El correo electrónico ya está registrado con otra cuenta. '
                'Por favor, utiliza otro correo electrónico válido.',
                'warning'
            )
            return _render_registro(nombre, email), 409

        codigo = generar_codigo_verificacion()
        password_hash = crear_hash_password(password)
        guardar_registro_pendiente(email, codigo, nombre=nombre, password=password_hash)
        session['registro_pendiente_email'] = email

        envio_ok, mensaje_envio = enviar_codigo_registro(email, codigo)
        if not envio_ok:
            PENDING_REGISTRATIONS.pop(email, None)
            session.pop('registro_pendiente_email', None)
            flash(mensaje_envio, 'danger')
            return _render_registro(nombre, email), 500

        flash(mensaje_envio, 'success')
        flash('Ingresa el código de verificación para activar tu cuenta.', 'info')
        return redirect(url_for('registro_verificacion'))

    return _render_registro()


@app.route('/registro/check-email', methods=['POST'])
def registro_check_email():
    try:
        payload = request.get_json(silent=True) or request.form
        email = normalizar_email(payload.get('email', ''))

        if not email:
            return jsonify({'success': False, 'exists': False, 'message': 'Debes ingresar un correo electrónico.'}), 400

        if not email_es_valido(email):
            return jsonify({'success': False, 'exists': False, 'message': 'El correo electrónico no es válido.'}), 400

        usuarios = cargar_usuarios_df()
        usuarios['email'] = usuarios['email'].astype(str).str.strip().str.lower()
        email_existente = email in usuarios['email'].values

        if email_existente:
            return jsonify({
                'success': True,
                'exists': True,
                'message': (
                    'El correo electrónico ya está registrado con otra cuenta. '
                    'Por favor, utiliza otro correo electrónico válido.'
                )
            })

        return jsonify({
            'success': True,
            'exists': False,
            'message': 'Correo electrónico disponible.'
        })
    except Exception as e:
        print(f"Error validando correo de registro: {str(e)}")
        return jsonify({'success': False, 'exists': False, 'message': 'Error interno validando el correo.'}), 500


@app.route('/registro/send-code', methods=['POST'])
def registro_send_code():
    try:
        payload = request.get_json(silent=True) or request.form
        email = normalizar_email(payload.get('email', ''))

        if not email:
            return jsonify({'success': False, 'message': 'Debes ingresar un correo electrónico.'}), 400

        if not email_es_valido(email):
            return jsonify({'success': False, 'message': 'El correo electrónico no es válido.'}), 400

        usuarios = cargar_usuarios_df()
        usuarios['email'] = usuarios['email'].astype(str).str.strip().str.lower()
        if email in usuarios['email'].values:
            return jsonify({
                'success': False,
                'message': (
                    'El correo electrónico ya está registrado con otra cuenta. '
                    'Por favor, utiliza otro correo electrónico válido.'
                )
            }), 409

        pendiente_actual = obtener_registro_pendiente(email) or {}
        nombre = str(pendiente_actual.get('nombre', '')).strip()
        password = str(pendiente_actual.get('password', ''))

        codigo = generar_codigo_verificacion()
        guardar_registro_pendiente(email, codigo, nombre=nombre, password=password)
        envio_ok, mensaje_envio = enviar_codigo_registro(email, codigo)

        if not envio_ok:
            PENDING_REGISTRATIONS.pop(email, None)
            session.pop('registro_pendiente_email', None)
            return jsonify({'success': False, 'message': mensaje_envio}), 500

        session['registro_pendiente_email'] = email
        return jsonify({
            'success': True,
            'message': mensaje_envio
        })
    except Exception as e:
        print(f"Error al enviar código de registro: {str(e)}")
        return jsonify({'success': False, 'message': 'Error interno enviando el código.'}), 500


@app.route('/registro/verificacion', methods=['GET', 'POST'])
def registro_verificacion():
    def _render_verificacion(email=''):
        return render_template(
            'Usuarios/Autenticacion/registro_verificacion.html',
            email_registro=email,
            registro_code_minutes=REGISTER_CODE_EXP_MINUTES
        )

    email = normalizar_email(session.get('registro_pendiente_email', ''))
    if not email:
        flash('Primero debes completar el formulario de registro.', 'warning')
        return redirect(url_for('registro'))

    registro_pendiente = obtener_registro_pendiente(email)
    if not registro_pendiente:
        session.pop('registro_pendiente_email', None)
        flash('No hay un registro pendiente o el código ya expiró. Registra tus datos nuevamente.', 'warning')
        return redirect(url_for('registro'))

    if request.method == 'POST':
        accion = request.form.get('action', 'verify')
        if accion == 'resend':
            codigo_nuevo = generar_codigo_verificacion()
            nombre = str(registro_pendiente.get('nombre', '')).strip()
            password = str(registro_pendiente.get('password', ''))
            guardar_registro_pendiente(email, codigo_nuevo, nombre=nombre, password=password)
            envio_ok, mensaje_envio = enviar_codigo_registro(email, codigo_nuevo)
            flash(mensaje_envio, 'success' if envio_ok else 'danger')
            if not envio_ok:
                PENDING_REGISTRATIONS.pop(email, None)
                session.pop('registro_pendiente_email', None)
                return redirect(url_for('registro'))
            return _render_verificacion(email)

        codigo_ingresado = request.form.get('verification_code', '').strip()
        if not codigo_ingresado:
            flash('Debes ingresar el código de verificación.', 'danger')
            return _render_verificacion(email), 400

        registro_pendiente = obtener_registro_pendiente(email)
        if not registro_pendiente:
            session.pop('registro_pendiente_email', None)
            flash('El código expiró. Debes iniciar el registro nuevamente.', 'warning')
            return redirect(url_for('registro'))

        if str(codigo_ingresado) != str(registro_pendiente.get('code', '')).strip():
            flash('Código de verificación incorrecto.', 'danger')
            return _render_verificacion(email), 400

        usuarios = cargar_usuarios_df()
        usuarios['email'] = usuarios['email'].astype(str).str.strip().str.lower()
        if email in usuarios['email'].values:
            PENDING_REGISTRATIONS.pop(email, None)
            session.pop('registro_pendiente_email', None)
            flash(
                'El correo electrónico ya está registrado con otra cuenta. '
                'Por favor, utiliza otro correo electrónico válido.',
                'warning'
            )
            return redirect(url_for('registro'))

        nombre = str(registro_pendiente.get('nombre', '')).strip()
        password_guardado = str(registro_pendiente.get('password', ''))
        if not nombre or not password_guardado:
            PENDING_REGISTRATIONS.pop(email, None)
            session.pop('registro_pendiente_email', None)
            flash('No se encontraron los datos del registro pendiente. Intenta nuevamente.', 'warning')
            return redirect(url_for('registro'))

        if not password_esta_hasheado(password_guardado):
            password_guardado = crear_hash_password(password_guardado)

        ultimo_id = pd.to_numeric(usuarios['id_usuario'], errors='coerce').max()
        nuevo_id = int(ultimo_id + 1) if pd.notna(ultimo_id) else 1
        nuevo_usuario = {
            'id_usuario': nuevo_id,
            'nombre': nombre,
            'email': email,
            'password_hash': password_guardado,
            'rol': 'normal',
            'estado': 'activo',
            'fecha_registro': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'email_verified': True,
            'verification_code': '',
            'verification_code_expiry': '',
            'reset_token': '',
            'reset_token_expiry': ''
        }

        usuarios = pd.concat([usuarios, pd.DataFrame([nuevo_usuario])], ignore_index=True)
        guardar_usuarios_df(usuarios)
        PENDING_REGISTRATIONS.pop(email, None)
        session.pop('registro_pendiente_email', None)

        registrar_actividad(f"Nuevo usuario registrado y verificado: {nombre}")
        session['usuario'] = email
        session['id_usuario'] = int(nuevo_id)
        session['rol'] = 'normal'
        session['nombre'] = nombre
        flash('Cuenta creada correctamente. Correo verificado.', 'success')
        return redirect(url_for('user_dashboard'))

    return _render_verificacion(email)

@app.route('/logout')
def logout():
    if session.get('rol') == 'normal':
        _guardar_carrito_guardado_usuario(session.get('usuario', ''), _obtener_carrito_sesion_usuario())
    session.clear()
    return redirect(url_for('home'))

@app.route('/admin')
def admin_dashboard():
    if session.get('rol') == 'admin':
        productos = cargar_productos_activos_df()
        lista_productos = productos.to_dict(orient='records')
        productos_agotados = obtener_productos_agotados()
        admin_nombre = obtener_nombre_sesion()

        # Métricas de usuarios y ventas últimos 30 días
        hoy = datetime.now()
        hace_30 = hoy - timedelta(days=30)

        usuarios_df = cargar_usuarios_df()
        if not usuarios_df.empty and 'fecha_registro' in usuarios_df.columns:
            # Normaliza la zona horaria para evitar comparaciones entre tz-aware y tz-naive
            usuarios_df['fecha_registro_dt'] = (
                pd.to_datetime(usuarios_df['fecha_registro'], errors='coerce', utc=True)
                .dt.tz_convert(None)
            )
            usuarios_ult_mes = usuarios_df[usuarios_df['fecha_registro_dt'] >= hace_30]['id_usuario'].nunique()
        else:
            usuarios_ult_mes = 0

        pedidos_df = cargar_pedidos_df()
        usuarios_compraron_ult_mes = 0
        if not pedidos_df.empty:
            pedidos_df['fecha_pedido_dt'] = (
                pd.to_datetime(pedidos_df['fecha_pedido'], dayfirst=True, errors='coerce', utc=True)
                .dt.tz_convert(None)
            )
            pedidos_ult_mes = pedidos_df[pedidos_df['fecha_pedido_dt'] >= hace_30]
            usuarios_compraron_ult_mes = pedidos_ult_mes['id_usuario'].nunique()

        # Top productos vendidos (totales)
        detalle_df = cargar_detalle_pedido_df()
        top_productos = []
        if not detalle_df.empty:
            detalle_df['cantidad'] = pd.to_numeric(detalle_df.get('cantidad', 0), errors='coerce').fillna(0)
            top_df = detalle_df.groupby('id_producto', as_index=False)['cantidad'].sum().sort_values('cantidad', ascending=False).head(5)
            prod_ref = productos[['id_producto', 'nombre']] if not productos.empty else pd.DataFrame(columns=['id_producto', 'nombre'])
            top_df = pd.merge(top_df, prod_ref, on='id_producto', how='left')
            top_productos = top_df.to_dict(orient='records')

        productos_destacables = []
        destacados_count = 0
        if not productos.empty:
            productos_ordenados = productos.sort_values(['fuerza', 'intendencia', 'nombre'], ascending=[True, True, True])
            productos_destacables = productos_ordenados.to_dict(orient='records')
            destacados_count = int(pd.to_numeric(productos_ordenados.get('destacado_dashboard', False), errors='coerce').fillna(0).astype(bool).sum())

        # plantilla actual en el proyecto
        return render_template(
            'Administrador/admin_dashboard_principal.html',
            productos=lista_productos,
            admin_nombre=admin_nombre,
            productos_agotados=productos_agotados,
            usuarios_ult_mes=usuarios_ult_mes,
            usuarios_compraron_ult_mes=usuarios_compraron_ult_mes,
            top_productos=top_productos,
            productos_destacables=productos_destacables,
            destacados_count=destacados_count,
            destacados_max=5,
        )
    return "Acceso denegado"


@app.route('/admin/destacados', methods=['POST'])
def admin_actualizar_destacados():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos = cargar_productos_df()
    if productos.empty:
        flash('No hay productos para destacar.', 'warning')
        return redirect(url_for('admin_dashboard'))

    productos['id_producto'] = pd.to_numeric(productos['id_producto'], errors='coerce')
    ids_disponibles = set(
        productos.loc[productos['eliminado'] == False, 'id_producto']
        .dropna()
        .astype(int)
        .tolist()
    )

    # Obtener listas por categoría
    ids_ejercito_raw = request.form.getlist('destacados_ejercito')
    ids_policia_raw = request.form.getlist('destacados_policia')
    ids_armada_raw = request.form.getlist('destacados_armada')

    def procesar_ids(ids_raw):
        ids_seleccionados = []
        for valor in ids_raw:
            try:
                valor_int = int(valor)
            except (TypeError, ValueError):
                continue
            if valor_int in ids_disponibles and valor_int not in ids_seleccionados:
                ids_seleccionados.append(valor_int)
        return ids_seleccionados[:5]  # Máximo 5 por categoría

    ids_ejercito = procesar_ids(ids_ejercito_raw)
    ids_policia = procesar_ids(ids_policia_raw)
    ids_armada = procesar_ids(ids_armada_raw)

    # Resetear destacados
    productos['destacado_dashboard'] = False
    productos.loc[productos['eliminado'] == True, 'destacado_dashboard'] = False

    # Asignar destacados por categoría
    productos.loc[productos['id_producto'].isin(ids_ejercito), 'destacado_dashboard'] = True
    productos.loc[productos['id_producto'].isin(ids_policia), 'destacado_dashboard'] = True
    productos.loc[productos['id_producto'].isin(ids_armada), 'destacado_dashboard'] = True

    guardar_productos_df(productos)

    total_destacados = len(ids_ejercito) + len(ids_policia) + len(ids_armada)
    registrar_actividad(
        "Administrador actualizó prendas destacadas por categoría:\n"
        f"- Ejército: {len(ids_ejercito)} ({', '.join(str(x) for x in ids_ejercito) if ids_ejercito else 'ninguna'})\n"
        f"- Policía: {len(ids_policia)} ({', '.join(str(x) for x in ids_policia) if ids_policia else 'ninguna'})\n"
        f"- Armada: {len(ids_armada)} ({', '.join(str(x) for x in ids_armada) if ids_armada else 'ninguna'})\n"
        f"- Total destacadas: {total_destacados}"
    )

    flash('Prendas destacadas actualizadas correctamente.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/productos')
def admin_productos():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos_todos = cargar_productos_df()
    total_eliminados = int(productos_todos['eliminado'].fillna(False).astype(bool).sum())
    productos = cargar_productos_activos_df()
    total_productos = len(productos)
    busqueda = request.args.get('q', '').strip()
    fuerza_filtro = request.args.get('fuerza', '').strip()
    intendencia_filtro = request.args.get('intendencia', '').strip()

    if fuerza_filtro and fuerza_filtro not in FUERZAS_OPCIONES:
        fuerza_filtro = ''
    if intendencia_filtro and intendencia_filtro not in INTENDENCIAS_OPCIONES:
        intendencia_filtro = ''

    if busqueda:
        filtros = [
            productos['nombre'].str.contains(busqueda, case=False, na=False, regex=False),
            productos['descripcion'].str.contains(busqueda, case=False, na=False, regex=False),
            productos['fuerza'].str.contains(busqueda, case=False, na=False, regex=False),
            productos['intendencia'].str.contains(busqueda, case=False, na=False, regex=False),
            productos['id_producto'].fillna('').astype(str).str.contains(busqueda, case=False, na=False, regex=False),
        ]
        mascara = filtros[0]
        for filtro in filtros[1:]:
            mascara = mascara | filtro
        productos = productos[mascara].copy()

    if fuerza_filtro:
        productos = productos[productos['fuerza'].str.strip() == fuerza_filtro].copy()

    if intendencia_filtro:
        productos = productos[productos['intendencia'].str.strip() == intendencia_filtro].copy()

    hay_filtros = bool(busqueda or fuerza_filtro or intendencia_filtro)
    lista_productos = productos.to_dict(orient='records')
    productos_por_fuerza = {fuerza: [] for fuerza in FUERZAS_OPCIONES}

    for producto in lista_productos:
        imagen_principal = str(producto.get('imagen_url', '')).strip()
        galeria = obtener_galeria_producto(producto.get('id_producto'), imagen_principal)
        producto['imagenes'] = galeria
        if galeria:
            producto['imagen_url'] = galeria[0]

        fuerza_producto = str(producto.get('fuerza', '')).strip().lower()
        for fuerza in FUERZAS_OPCIONES:
            if fuerza_producto == fuerza.lower():
                productos_por_fuerza[fuerza].append(producto)
                break

    return render_template(
        'Administrador/Gestion productos/admin_prod_dashboard.html',
        productos=lista_productos,
        productos_por_fuerza=productos_por_fuerza,
        fuerzas=FUERZAS_OPCIONES,
        intendencias=INTENDENCIAS_OPCIONES,
        search_query=busqueda,
        selected_fuerza=fuerza_filtro,
        selected_intendencia=intendencia_filtro,
        has_filters=hay_filtros,
        total_productos=total_productos,
        total_eliminados=total_eliminados,
        max_images_per_product=MAX_IMAGES_PER_PRODUCT
    )


@app.route('/admin/agregar_producto', methods=['POST'])
def agregar_producto():
    if session.get('rol') == 'admin':
        productos = cargar_productos_df()
        nuevo_id = next_id('producto', 'id_producto')
        fuerza = request.form.get('fuerza', '').strip()
        intendencia = request.form.get('intendencia', '').strip()
        if fuerza not in FUERZAS_OPCIONES or intendencia not in INTENDENCIAS_OPCIONES:
            flash('Selecciona una fuerza e intendencia válidas.', 'danger')
            return redirect(url_for('admin_productos'))

        imagenes = [
            archivo for archivo in request.files.getlist('imagenes')
            if archivo and str(getattr(archivo, 'filename', '')).strip()
        ]
        if not imagenes:
            imagen_unica = request.files.get('imagen')
            if imagen_unica and str(getattr(imagen_unica, 'filename', '')).strip():
                imagenes = [imagen_unica]

        if len(imagenes) > MAX_IMAGES_PER_PRODUCT:
            flash(f'Solo puedes subir hasta {MAX_IMAGES_PER_PRODUCT} imágenes por producto.', 'danger')
            return redirect(url_for('admin_productos'))

        for imagen in imagenes:
            error_validacion = validar_archivo_imagen(imagen)
            if error_validacion:
                flash(error_validacion, 'danger')
                return redirect(url_for('admin_productos'))

        galeria_guardada = guardar_galeria_producto(nuevo_id, imagenes, reemplazar=True)
        imagen_url = galeria_guardada[0] if galeria_guardada else ''

        nuevo_producto = {
            'id_producto': nuevo_id,
            'nombre': request.form['nombre'],
            'descripcion': request.form['descripcion'],
            'precio': float(request.form['precio']),
            'stock': int(request.form['stock']),
            'id_categoria': 1,  # Se conserva por compatibilidad
            'fuerza': fuerza,
            'intendencia': intendencia,
            'imagen_url': imagen_url,
            'eliminado': False
        }

        productos = pd.concat([productos, pd.DataFrame([nuevo_producto])], ignore_index=True)
        guardar_productos_df(productos)

        registrar_actividad(
            f"Creo producto '{request.form['nombre']}' (ID {nuevo_id})\n"
            f"- precio: {formatear_cop(float(request.form['precio']))}\n"
            f"- stock: {int(request.form['stock'])}\n"
            f"- fuerza: {fuerza}\n"
            f"- intendencia: {intendencia}\n"
            f"- imágenes: {len(galeria_guardada)}"
        )

        return redirect(url_for('admin_productos'))
    return "Acceso denegado"

@app.route('/admin/sales/export')
def export_sales():
    if session.get('rol') != 'admin':
        return "Acceso denegado"
    
    pedidos = cargar_pedidos_df()
    pagos = cargar_pagos_df()
    detalle = cargar_detalle_pedido_df()
    
    # Combinar los datos
    informe = pd.merge(pedidos, pagos, on='id_pedido', how='left')
    
    # Agregar totales por pedido
    totales_pedido = detalle.groupby('id_pedido')['subtotal'].sum().reset_index()
    totales_pedido.columns = ['id_pedido', 'total_productos']
    informe = pd.merge(informe, totales_pedido, on='id_pedido', how='left')
    
    # Convertir a CSV
    csv_data = informe.to_csv(index=False)
    
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
    )


@app.route('/admin/imagen/<int:id_producto>', methods=['POST'])
def subir_imagen(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    imagenes = [
        archivo for archivo in request.files.getlist('imagenes')
        if archivo and str(getattr(archivo, 'filename', '')).strip()
    ]
    if not imagenes:
        imagen_unica = request.files.get('imagen')
        if imagen_unica and str(getattr(imagen_unica, 'filename', '')).strip():
            imagenes = [imagen_unica]

    if not imagenes:
        flash("No se seleccionó ninguna imagen.")
        return redirect(url_for('admin_productos'))

    if len(imagenes) > MAX_IMAGES_PER_PRODUCT:
        flash(f'Solo puedes subir hasta {MAX_IMAGES_PER_PRODUCT} imágenes por producto.', 'danger')
        return redirect(url_for('admin_productos'))

    for imagen in imagenes:
        error_validacion = validar_archivo_imagen(imagen)
        if error_validacion:
            flash(error_validacion, 'danger')
            return redirect(url_for('admin_productos'))

    galeria_guardada = guardar_galeria_producto(id_producto, imagenes, reemplazar=True)

    productos = cargar_productos_df()
    idx = productos[productos['id_producto'] == id_producto].index
    if not idx.empty:
        productos.at[idx[0], 'imagen_url'] = galeria_guardada[0] if galeria_guardada else ''
        guardar_productos_df(productos)
        flash(f"Galería reemplazada ({len(galeria_guardada)} imágenes).", 'success')
    else:
        flash("Producto no encontrado.", 'danger')

    return redirect(url_for('admin_productos'))


@app.route('/admin/imagen/agregar/<int:id_producto>', methods=['POST'])
def agregar_imagenes_producto(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    imagenes = [
        archivo for archivo in request.files.getlist('imagenes_agregar')
        if archivo and str(getattr(archivo, 'filename', '')).strip()
    ]
    if not imagenes:
        imagenes = [
            archivo for archivo in request.files.getlist('imagenes')
            if archivo and str(getattr(archivo, 'filename', '')).strip()
        ]

    if not imagenes:
        flash("No se seleccionaron imágenes para agregar.", 'warning')
        return redirect(url_for('admin_productos'))

    for imagen in imagenes:
        error_validacion = validar_archivo_imagen(imagen)
        if error_validacion:
            flash(error_validacion, 'danger')
            return redirect(url_for('admin_productos'))

    productos = cargar_productos_df()
    idx = productos[productos['id_producto'] == id_producto].index
    if idx.empty:
        flash("Producto no encontrado.", 'danger')
        return redirect(url_for('admin_productos'))

    imagen_principal = str(productos.at[idx[0], 'imagen_url']) if 'imagen_url' in productos.columns else ''
    galeria_actual = obtener_galeria_producto(id_producto, imagen_principal)
    if len(galeria_actual) + len(imagenes) > MAX_IMAGES_PER_PRODUCT:
        disponibles = max(0, MAX_IMAGES_PER_PRODUCT - len(galeria_actual))
        flash(
            f"Este producto ya tiene {len(galeria_actual)} imagen(es). "
            f"Solo puedes agregar {disponibles} más (máximo {MAX_IMAGES_PER_PRODUCT}).",
            'danger'
        )
        return redirect(url_for('admin_productos'))

    guardar_galeria_producto(id_producto, imagenes, reemplazar=False)
    galeria_actualizada = obtener_galeria_producto(id_producto, '')
    productos.at[idx[0], 'imagen_url'] = galeria_actualizada[0] if galeria_actualizada else ''
    guardar_productos_df(productos)

    flash(f"Se agregaron {len(imagenes)} imágenes. Total: {len(galeria_actualizada)}.", 'success')
    return redirect(url_for('admin_productos'))


@app.route('/admin/imagen/eliminar/<int:id_producto>', methods=['POST'])
def eliminar_imagen_producto(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    imagen_a_eliminar = normalizar_imagen_url(request.form.get('imagen_a_eliminar', '').strip())
    if not imagen_a_eliminar:
        flash("Selecciona una imagen para eliminar.", 'warning')
        return redirect(url_for('admin_productos'))

    productos = cargar_productos_df()
    idx = productos[productos['id_producto'] == id_producto].index
    if idx.empty:
        flash("Producto no encontrado.", 'danger')
        return redirect(url_for('admin_productos'))

    imagen_principal = str(productos.at[idx[0], 'imagen_url']) if 'imagen_url' in productos.columns else ''
    galeria_actual = obtener_galeria_producto(id_producto, imagen_principal)
    if imagen_a_eliminar not in galeria_actual:
        flash("La imagen seleccionada no pertenece a este producto.", 'danger')
        return redirect(url_for('admin_productos'))

    ruta_absoluta = ruta_imagen_producto_absoluta(imagen_a_eliminar)
    if ruta_absoluta and os.path.exists(ruta_absoluta):
        os.remove(ruta_absoluta)

    galeria_actualizada = obtener_galeria_producto(id_producto, '')
    productos.at[idx[0], 'imagen_url'] = galeria_actualizada[0] if galeria_actualizada else ''
    guardar_productos_df(productos)

    flash(f"Imagen eliminada. Total actual: {len(galeria_actualizada)}.", 'success')
    return redirect(url_for('admin_productos'))



@app.route('/admin/eliminar/<int:id_producto>', methods=['POST'])
def eliminar_producto(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos = cargar_productos_df()

    idx = productos[productos['id_producto'] == id_producto].index

    if not idx.empty:
        productos.at[idx[0], 'eliminado'] = True
        guardar_productos_df(productos)

        nombre = productos.at[idx[0], 'nombre']
        registrar_actividad(f"Elimino producto: {nombre} (ID {id_producto})")

    return redirect(url_for('admin_productos'))

@app.route('/admin/eliminar_definitivo/<int:id_producto>', methods=['POST'])
def eliminar_definitivo(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos = cargar_productos_df()
    idx = productos[productos['id_producto'] == id_producto].index

    if not idx.empty:
        nombre = productos.at[idx[0], 'nombre']
        productos = productos.drop(index=idx)
        guardar_productos_df(productos)

        registrar_actividad(f"Elimino definitivamente el producto: {nombre} (ID {id_producto})")

    return redirect(url_for('admin_papelera'))

@app.route('/admin/restaurar/<int:id_producto>', methods=['POST'])
def restaurar_producto(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos = cargar_productos_df()
    idx = productos[productos['id_producto'] == id_producto].index

    if not idx.empty:
        productos.at[idx[0], 'eliminado'] = False
        guardar_productos_df(productos)

        nombre = productos.at[idx[0], 'nombre']
        registrar_actividad(f"Restauro producto: {nombre} (ID {id_producto})")

    return redirect(url_for('admin_papelera'))


@app.route('/admin/papelera')
def admin_papelera():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos = cargar_productos_df()
    eliminados = productos[productos['eliminado'] == True].to_dict(orient='records')
    return render_template('Administrador/Papelera/admin_papelera.html', productos=eliminados)


@app.route('/admin/usuarios')
def admin_usuarios():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    usuarios = cargar_usuarios_df()
    usuarios['id_usuario'] = pd.to_numeric(usuarios['id_usuario'], errors='coerce')

    buscar = request.args.get('buscar', '').strip()
    rol = request.args.get('rol', '').strip()
    estado = request.args.get('estado', '').strip().lower()
    orden = request.args.get('orden', 'reciente').strip()
    edit_id = request.args.get('edit_id', type=int)

    filtrados = usuarios.copy()
    if buscar:
        mask = (
            filtrados['nombre'].astype(str).str.contains(buscar, case=False, na=False) |
            filtrados['email'].astype(str).str.contains(buscar, case=False, na=False)
        )
        filtrados = filtrados[mask]

    if rol:
        filtrados = filtrados[filtrados['rol'].astype(str).str.lower() == rol.lower()]
    if estado in ['activo', 'inactivo']:
        filtrados = filtrados[filtrados['estado'].astype(str).str.lower() == estado]

    if orden == 'antiguo':
        filtrados = filtrados.sort_values(by='id_usuario', ascending=True, na_position='last')
    elif orden == 'nombre':
        filtrados = filtrados.sort_values(by='nombre', ascending=True, na_position='last')
    else:
        filtrados = filtrados.sort_values(by='id_usuario', ascending=False, na_position='last')

    usuario_editar = None
    if edit_id is not None:
        candidato = usuarios[usuarios['id_usuario'] == edit_id]
        if not candidato.empty:
            usuario_editar = candidato.iloc[0].to_dict()
            if pd.notna(usuario_editar.get('id_usuario')):
                usuario_editar['id_usuario'] = int(usuario_editar['id_usuario'])
            usuario_editar['estado'] = str(usuario_editar.get('estado', 'activo')).strip().lower()

    lista_usuarios = filtrados.fillna('').to_dict(orient='records')
    for usuario in lista_usuarios:
        if usuario.get('id_usuario') != '':
            usuario['id_usuario'] = int(usuario['id_usuario'])
        usuario['estado'] = str(usuario.get('estado', 'activo')).strip().lower()

    filtros = {'buscar': buscar, 'rol': rol, 'estado': estado, 'orden': orden}
    return render_template(
        'Administrador/Gestion usuarios/admin_usuario.html',
        usuarios=lista_usuarios,
        usuario_editar=usuario_editar,
        filtros=filtros
    )


@app.route('/admin/usuarios/guardar', methods=['POST'])
def admin_guardar_usuario():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    id_usuario_raw = request.form.get('id_usuario', '').strip()
    nombre = request.form.get('nombre', '').strip()
    email = normalizar_email(request.form.get('email', ''))
    password = request.form.get('password', '').strip()
    rol = request.form.get('rol', 'normal').strip().lower()
    estado = request.form.get('estado', 'activo').strip().lower()

    if rol not in ['admin', 'normal']:
        rol = 'normal'
    if estado not in ['activo', 'inactivo']:
        estado = 'activo'

    if not nombre or not email:
        flash('Nombre y email son obligatorios.', 'danger')
        return redirect(url_for('admin_usuarios'))

    usuarios = cargar_usuarios_df()
    usuarios['id_usuario'] = pd.to_numeric(usuarios['id_usuario'], errors='coerce')
    usuarios['email'] = usuarios['email'].astype(str)

    edit_id = None
    if id_usuario_raw:
        try:
            edit_id = int(float(id_usuario_raw))
        except ValueError:
            flash('ID de usuario inválido.', 'danger')
            return redirect(url_for('admin_usuarios'))

    if edit_id is None and not password:
        flash('La contraseña es obligatoria para crear usuarios.', 'danger')
        return redirect(url_for('admin_usuarios'))

    existe_email = usuarios[usuarios['email'].str.lower() == email.lower()]
    if edit_id is not None:
        existe_email = existe_email[existe_email['id_usuario'] != edit_id]
    if not existe_email.empty:
        flash('Ese email ya está registrado por otro usuario.', 'danger')
        return redirect(url_for('admin_usuarios'))

    if edit_id is not None:
        idx = usuarios[usuarios['id_usuario'] == edit_id].index
        if idx.empty:
            flash('Usuario no encontrado para edición.', 'danger')
            return redirect(url_for('admin_usuarios'))

        usuarios.at[idx[0], 'nombre'] = nombre
        usuarios.at[idx[0], 'email'] = email
        usuarios.at[idx[0], 'rol'] = rol
        usuarios.at[idx[0], 'estado'] = estado
        if password:
            usuarios.at[idx[0], 'password_hash'] = crear_hash_password(password)

        guardar_usuarios_df(usuarios[USUARIO_COLUMNS])
        registrar_actividad(f"Actualizo usuario {email} (ID {edit_id})\n- rol: {rol}\n- estado: {estado}")
        flash('Usuario actualizado correctamente.', 'success')
    else:
        ultimo_id = pd.to_numeric(usuarios['id_usuario'], errors='coerce').max()
        nuevo_id = int(ultimo_id + 1) if pd.notna(ultimo_id) else 1

        nuevo_usuario = {
            'id_usuario': nuevo_id,
            'nombre': nombre,
            'email': email,
            'password_hash': crear_hash_password(password),
            'rol': rol,
            'estado': estado,
            'fecha_registro': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'reset_token': '',
            'reset_token_expiry': ''
        }
        usuarios = pd.concat([usuarios, pd.DataFrame([nuevo_usuario])], ignore_index=True)
        guardar_usuarios_df(usuarios[USUARIO_COLUMNS])
        registrar_actividad(f"Creo usuario {email} (ID {nuevo_id})\n- rol: {rol}\n- estado: {estado}")
        flash('Usuario creado correctamente.', 'success')

    return redirect(url_for('admin_usuarios'))


@app.route('/admin/registros')
def admin_registros():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    registros = cargar_registros_df()

    filtro_usuario = request.args.get('usuario', '').strip()
    filtro_fecha = request.args.get('fecha', '').strip()

    if filtro_usuario:
        registros = registros[
            registros['id_usuario'].astype(str).str.contains(filtro_usuario, case=False, na=False) |
            registros['accion'].astype(str).str.contains(filtro_usuario, case=False, na=False)
        ]

    if filtro_fecha:
        registros = registros[registros['fecha_accion'].astype(str).str.startswith(filtro_fecha)]

    registros['id_registro'] = pd.to_numeric(registros['id_registro'], errors='coerce')
    registros = registros.sort_values(by='id_registro', ascending=False, na_position='last')

    lista_registros = registros.fillna('').to_dict(orient='records')
    for registro in lista_registros:
        if registro.get('id_registro') != '':
            registro['id_registro'] = int(registro['id_registro'])

    filtros = {'usuario': filtro_usuario, 'fecha': filtro_fecha}
    return render_template(
        'Administrador/Gestion usuarios/admin_registros.html',
        registros=lista_registros,
        filtros=filtros
    )


@app.route('/admin/registros/export_excel')
def admin_registros_export_excel():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    registros = cargar_registros_df()
    filtro_usuario = request.args.get('usuario', '').strip()
    filtro_fecha = request.args.get('fecha', '').strip()

    if filtro_usuario:
        registros = registros[
            registros['id_usuario'].astype(str).str.contains(filtro_usuario, case=False, na=False) |
            registros['accion'].astype(str).str.contains(filtro_usuario, case=False, na=False)
        ]
    if filtro_fecha:
        registros = registros[registros['fecha_accion'].astype(str).str.startswith(filtro_fecha)]

    buffer = BytesIO()
    registros.to_excel(buffer, index=False)
    buffer.seek(0)

    nombre_archivo = f"registros_bd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/admin/ajustes')
def admin_ajustes():
    if session.get('rol') != 'admin':
        return "Acceso denegado"
    pedidos = cargar_pedidos_df()
    pagos = cargar_pagos_df()
    detalle = cargar_detalle_pedido_df()
    productos = cargar_productos_df()
    usuarios = cargar_usuarios_df()

    pedidos, pagos, detalle, productos = _normalizar_dataframes_admin_pedidos(
        pedidos, pagos, detalle, productos
    )
    pedidos_view = _construir_vista_pedidos(pedidos, pagos, detalle, usuarios, productos)
    pedidos_view = pedidos_view.sort_values(by='id_pedido', ascending=False, na_position='last')

    pagina_actual = _parse_positive_int(request.args.get('ajustes_page', 1), default=1)
    lista_pedidos = _enriquecer_pedidos_con_tracking(_serializar_pedidos_admin(pedidos_view))
    pedidos_activos = [pedido for pedido in lista_pedidos if pedido.get('estado_activo')]
    pedidos_activos_vista, paginacion_ajustes = _paginar_lista(pedidos_activos, pagina_actual, per_page=5)

    resumen_estados = {
        'activos': len(pedidos_activos),
        'pendientes': sum(1 for pedido in pedidos_activos if pedido.get('estado_ui') == 'pendiente'),
        'preparacion': sum(
            1 for pedido in pedidos_activos
            if pedido.get('estado_ui') in {'confirmado', 'en_preparacion', 'empaquetado'}
        ),
        'enviados': sum(1 for pedido in pedidos_activos if pedido.get('estado_ui') == 'enviado'),
    }

    return render_template(
        'Administrador/Ajustes/admin_ajustes_dashboard.html',
        pedidos_activos=pedidos_activos_vista,
        resumen_estados=resumen_estados,
        paginacion_ajustes=paginacion_ajustes,
        estados_pedido=PEDIDO_STATUS_FLOW[:-1] + [('entregado', 'Entregado'), ('cancelado', 'Cancelado')],
    )


@app.route('/admin/informes')
def admin_informes():
    if session.get('rol') != 'admin':
        return "Acceso denegado"
    return render_template('Administrador/Informes/admin_infor_dashboard.html')


@app.route('/admin/pos')
def admin_pos():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    recibo_pedido_id = None
    descargar_recibo = request.args.get('descargar_recibo', '').strip()
    if descargar_recibo:
        try:
            recibo_pedido_id = int(descargar_recibo)
            if recibo_pedido_id <= 0:
                recibo_pedido_id = None
        except ValueError:
            recibo_pedido_id = None

    productos = cargar_productos_activos_df()
    promos = cargar_promociones_df()
    hoy = datetime.now().date()
    mejor_promo_por_producto = obtener_mejor_promocion_por_producto(productos, promos, hoy)

    lista_productos = productos.to_dict(orient='records')
    for producto in lista_productos:
        id_producto = int(pd.to_numeric(producto.get('id_producto', 0), errors='coerce') or 0)
        galeria_producto = obtener_galeria_producto(id_producto, producto.get('imagen_url', ''))
        if galeria_producto:
            producto['imagen_url'] = normalizar_imagen_url(galeria_producto[0])
        else:
            producto['imagen_url'] = normalizar_imagen_url(producto.get('imagen_url', ''))
        producto['requiere_talla'] = bool(producto_requiere_talla(producto.get('intendencia', '')))
        precio_base = float(pd.to_numeric(producto.get('precio', 0), errors='coerce') or 0)
        promo = mejor_promo_por_producto.get(id_producto)

        if promo:
            descuento_unitario = calcular_descuento_promocion(precio_base, promo)
            precio_venta = max(0.0, precio_base - descuento_unitario)
            producto['promo_activa'] = True
            producto['promo_id'] = promo.get('id_promo', '')
            producto['promo_codigo'] = promo.get('codigo', '')
            producto['promo_nombre'] = promo.get('nombre', '')
            producto['promo_tipo_descuento'] = promo.get('tipo_descuento', '')
            producto['promo_valor_descuento'] = float(pd.to_numeric(promo.get('valor_descuento', 0), errors='coerce') or 0)
            producto['promo_descuento_unitario'] = float(descuento_unitario)
            producto['precio_original'] = float(precio_base)
            producto['precio_con_descuento'] = float(precio_venta)
            producto['precio_venta'] = float(precio_venta)
        else:
            producto['promo_activa'] = False
            producto['promo_id'] = ''
            producto['promo_codigo'] = ''
            producto['promo_nombre'] = ''
            producto['promo_tipo_descuento'] = ''
            producto['promo_valor_descuento'] = 0.0
            producto['promo_descuento_unitario'] = 0.0
            producto['precio_original'] = float(precio_base)
            producto['precio_con_descuento'] = float(precio_base)
            producto['precio_venta'] = float(precio_base)

    return render_template(
        'Administrador/Sistema POS/admin_pos_dashboard.html',
        productos=lista_productos,
        fuerzas=FUERZAS_OPCIONES,
        tallas=TALLAS_OPCIONES,
        recibo_pedido_id=recibo_pedido_id
    )


@app.route('/admin/pos/recibo/<int:id_pedido>.pdf')
def admin_pos_recibo_pdf(id_pedido):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    try:
        pdf_bytes = generar_pdf_recibo_pos(id_pedido)
    except ValueError as exc:
        return Response(str(exc), status=404, mimetype='text/plain; charset=utf-8')
    except Exception as exc:
        return Response(
            f'No se pudo generar el recibo PDF. {str(exc)}',
            status=500,
            mimetype='text/plain; charset=utf-8'
        )

    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"recibo_pos_{id_pedido}.pdf"
    )


@app.route('/admin/pos/recibo/<int:id_pedido>/html')
def admin_pos_recibo_html(id_pedido):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    try:
        return render_html_recibo_pos(id_pedido)
    except ValueError as exc:
        return Response(str(exc), status=404, mimetype='text/plain; charset=utf-8')
    except Exception as exc:
        return Response(
            f'No se pudo generar la vista del recibo. {str(exc)}',
            status=500,
            mimetype='text/plain; charset=utf-8'
        )


def _normalizar_dataframes_admin_pedidos(pedidos, pagos, detalle, productos):
    if 'id_pedido' not in pedidos.columns:
        pedidos['id_pedido'] = pd.Series(dtype='int')
    if 'id_usuario' not in pedidos.columns:
        pedidos['id_usuario'] = ''
    if 'fecha_pedido' not in pedidos.columns:
        pedidos['fecha_pedido'] = ''
    if 'estado' not in pedidos.columns:
        pedidos['estado'] = 'pendiente'
    if 'cliente_telefono' not in pedidos.columns:
        pedidos['cliente_telefono'] = ''
    if 'cliente_direccion' not in pedidos.columns:
        pedidos['cliente_direccion'] = ''

    if 'id_pedido' not in pagos.columns:
        pagos['id_pedido'] = pd.Series(dtype='int')
    if 'id_pago' not in pagos.columns:
        pagos['id_pago'] = pd.Series(dtype='int')
    if 'monto' not in pagos.columns:
        pagos['monto'] = 0
    if 'metodo_pago' not in pagos.columns:
        pagos['metodo_pago'] = ''
    if 'fecha_pago' not in pagos.columns:
        pagos['fecha_pago'] = ''
    if 'estado_pago' not in pagos.columns:
        pagos['estado_pago'] = ''

    if 'id_pedido' not in detalle.columns:
        detalle['id_pedido'] = pd.Series(dtype='int')
    if 'id_producto' not in detalle.columns:
        detalle['id_producto'] = pd.Series(dtype='int')
    if 'cantidad' not in detalle.columns:
        detalle['cantidad'] = 0
    if 'subtotal' not in detalle.columns:
        detalle['subtotal'] = 0
    if 'talla' not in detalle.columns:
        detalle['talla'] = ''

    if 'id_producto' not in productos.columns:
        productos['id_producto'] = pd.Series(dtype='int')
    if 'nombre' not in productos.columns:
        productos['nombre'] = ''

    pedidos['id_pedido'] = pd.to_numeric(pedidos['id_pedido'], errors='coerce')
    pagos['id_pedido'] = pd.to_numeric(pagos['id_pedido'], errors='coerce')
    pagos['id_pago'] = pd.to_numeric(pagos['id_pago'], errors='coerce')
    pagos['monto'] = pd.to_numeric(pagos['monto'], errors='coerce').fillna(0)
    detalle['id_pedido'] = pd.to_numeric(detalle['id_pedido'], errors='coerce')
    detalle['id_producto'] = pd.to_numeric(detalle['id_producto'], errors='coerce')
    detalle['cantidad'] = pd.to_numeric(detalle['cantidad'], errors='coerce').fillna(0)
    detalle['subtotal'] = pd.to_numeric(detalle['subtotal'], errors='coerce').fillna(0)
    detalle['talla'] = detalle['talla'].fillna('').astype(str).str.strip().str.upper()
    productos['id_producto'] = pd.to_numeric(productos['id_producto'], errors='coerce')

    return pedidos, pagos, detalle, productos[['id_producto', 'nombre']].copy()


def _construir_mapa_usuarios(usuarios):
    usuarios_map = {}
    for _, usuario in usuarios.iterrows():
        uid = pd.to_numeric(usuario.get('id_usuario'), errors='coerce')
        if pd.notna(uid):
            usuarios_map[int(uid)] = str(usuario.get('nombre', '')).strip()
    return usuarios_map


def _resolver_nombre_usuario(valor, usuarios_map):
    uid = pd.to_numeric(valor, errors='coerce')
    if pd.notna(uid):
        uid = int(uid)
        if uid in usuarios_map and usuarios_map[uid]:
            return usuarios_map[uid]
        return f"Usuario #{uid}"
    return str(valor) if str(valor).strip() else 'N/A'


def _construir_mapa_productos(productos):
    productos_map = {}
    for _, producto in productos.iterrows():
        pid = pd.to_numeric(producto.get('id_producto'), errors='coerce')
        if pd.notna(pid):
            productos_map[int(pid)] = str(producto.get('nombre', '')).strip()
    return productos_map


def _construir_productos_por_pedido(detalle, productos_map):
    productos_por_pedido = {}

    for id_pedido, grupo in detalle.groupby('id_pedido'):
        if pd.isna(id_pedido):
            continue

        items = []
        for _, item in grupo.iterrows():
            pid = pd.to_numeric(item.get('id_producto'), errors='coerce')
            cantidad = pd.to_numeric(item.get('cantidad'), errors='coerce')
            talla = str(item.get('talla', '') or '').strip().upper()
            nombre_producto = ''

            if pd.notna(pid):
                nombre_producto = productos_map.get(int(pid), f"Producto #{int(pid)}")
            if not nombre_producto:
                nombre_producto = 'Producto sin nombre'

            etiqueta_talla = f" ({talla})" if talla else ''
            nombre_con_talla = f"{nombre_producto}{etiqueta_talla}"

            if pd.notna(cantidad) and int(cantidad) > 1:
                items.append(f"{nombre_con_talla} x{int(cantidad)}")
            else:
                items.append(nombre_con_talla)

        vistos = []
        for item in items:
            if item not in vistos:
                vistos.append(item)
        productos_por_pedido[int(id_pedido)] = ', '.join(vistos) if vistos else 'Sin productos'

    return productos_por_pedido


def _construir_vista_pedidos(pedidos, pagos, detalle, usuarios, productos):
    totales_pedido = detalle.groupby('id_pedido', as_index=False)['subtotal'].sum()
    totales_pedido = totales_pedido.rename(columns={'subtotal': 'total_productos'})

    pagos_ultimos = pagos.sort_values(by='id_pago', ascending=False).drop_duplicates(
        subset=['id_pedido'],
        keep='first'
    )
    pagos_ultimos = pagos_ultimos[['id_pedido', 'monto', 'metodo_pago', 'fecha_pago', 'estado_pago']]

    pedidos_view = pedidos.merge(totales_pedido, on='id_pedido', how='left')
    pedidos_view = pedidos_view.merge(pagos_ultimos, on='id_pedido', how='left')
    pedidos_view['total_productos'] = pedidos_view['total_productos'].fillna(0)
    pedidos_view['monto'] = pedidos_view['monto'].fillna(0)

    usuarios_map = _construir_mapa_usuarios(usuarios)
    productos_map = _construir_mapa_productos(productos)
    productos_por_pedido = _construir_productos_por_pedido(detalle, productos_map)

    pedidos_view['usuario_nombre'] = pedidos_view['id_usuario'].apply(
        lambda valor: _resolver_nombre_usuario(valor, usuarios_map)
    )
    pedidos_view['productos_pedido'] = pedidos_view['id_pedido'].apply(
        lambda valor: productos_por_pedido.get(int(valor), 'Sin productos') if pd.notna(valor) else 'Sin productos'
    )

    return pedidos_view


def _leer_filtros_admin_pedidos(args):
    filtros = {
        'q': str(args.get('q', '')).strip(),
        'estado': str(args.get('estado', 'todos')).strip().lower(),
        'fecha_desde': str(args.get('fecha_desde', '')).strip(),
        'fecha_hasta': str(args.get('fecha_hasta', '')).strip(),
        'page': str(args.get('page', '1')).strip(),
    }
    pago_filtros = {
        'q': str(args.get('pago_q', '')).strip(),
        'metodo': str(args.get('pago_metodo', 'todos')).strip().lower(),
        'estado': str(args.get('pago_estado', 'todos')).strip().lower(),
        'fecha_desde': str(args.get('pago_fecha_desde', '')).strip(),
        'fecha_hasta': str(args.get('pago_fecha_hasta', '')).strip(),
        'page': str(args.get('pago_page', '1')).strip(),
    }
    return filtros, pago_filtros


def _parse_positive_int(value, default=1):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(1, parsed)


def _build_pagination_buttons(total_paginas, pagina_actual):
    if total_paginas <= 7:
        return list(range(1, total_paginas + 1))

    botones = [1]
    inicio = max(2, pagina_actual - 1)
    fin = min(total_paginas - 1, pagina_actual + 1)
    if inicio > 2:
        botones.append('...')
    botones.extend(range(inicio, fin + 1))
    if fin < total_paginas - 1:
        botones.append('...')
    botones.append(total_paginas)
    return botones


def _paginar_lista(items, pagina_actual, per_page=5):
    total = len(items)
    total_paginas = max(1, (total + per_page - 1) // per_page)
    pagina_actual = min(max(1, pagina_actual), total_paginas)

    inicio = (pagina_actual - 1) * per_page
    fin = inicio + per_page
    vista = items[inicio:fin]

    paginacion = {
        'page': pagina_actual,
        'per_page': per_page,
        'total': total,
        'total_paginas': total_paginas,
        'desde': inicio + 1 if total else 0,
        'hasta': min(fin, total),
        'botones': _build_pagination_buttons(total_paginas, pagina_actual),
    }
    return vista, paginacion


def _estado_pedido_ui(estado):
    estado_normalizado = str(estado or '').strip().lower() or 'pendiente'
    return PEDIDO_STATUS_ALIAS.get(estado_normalizado, estado_normalizado)


def _etiqueta_estado_pedido(estado):
    estado_normalizado = str(estado or '').strip().lower() or 'pendiente'
    return PEDIDO_STATUS_LABELS.get(
        estado_normalizado,
        estado_normalizado.replace('_', ' ').capitalize()
    )


def _construir_pasos_estado_pedido(estado):
    estado_ui = _estado_pedido_ui(estado)
    flujo = [clave for clave, _ in PEDIDO_STATUS_FLOW]

    if estado_ui == 'cancelado':
        return []

    try:
        indice_actual = flujo.index(estado_ui)
    except ValueError:
        indice_actual = 0

    pasos = []
    for indice, (clave, etiqueta) in enumerate(PEDIDO_STATUS_FLOW):
        if indice < indice_actual:
            estado_paso = 'done'
        elif indice == indice_actual:
            estado_paso = 'current'
        else:
            estado_paso = 'upcoming'
        pasos.append({
            'key': clave,
            'label': etiqueta,
            'state': estado_paso,
        })
    return pasos


def _enriquecer_pedidos_con_tracking(registros):
    pedidos_enriquecidos = []
    for pedido in registros:
        pedido_item = dict(pedido)
        estado_original = str(pedido_item.get('estado', '')).strip().lower() or 'pendiente'
        estado_ui = _estado_pedido_ui(estado_original)
        pedido_item['estado'] = estado_original
        pedido_item['estado_ui'] = estado_ui
        pedido_item['estado_label'] = _etiqueta_estado_pedido(estado_original)
        pedido_item['estado_activo'] = estado_ui not in {'entregado', 'cancelado'}
        pedido_item['tracking_steps'] = _construir_pasos_estado_pedido(estado_original)
        pedidos_enriquecidos.append(pedido_item)
    return pedidos_enriquecidos


def _filtrar_y_paginar_pedidos(pedidos_view, filtros, per_page=10):
    filtros = dict(filtros)
    estados_validos = {clave for clave, _ in PEDIDO_STATUS_FLOW} | {'cancelado'}
    if filtros.get('estado') not in estados_validos:
        filtros['estado'] = 'todos'

    vista = pedidos_view.copy()
    vista['estado'] = vista['estado'].fillna('pendiente').astype(str).str.strip().str.lower()
    vista['estado_ui'] = vista['estado'].apply(_estado_pedido_ui)
    vista['fecha_pedido'] = vista['fecha_pedido'].fillna('').astype(str)
    vista['fecha_pedido_dt'] = pd.to_datetime(vista['fecha_pedido'], errors='coerce')
    vista['id_pedido_txt'] = vista['id_pedido'].apply(
        lambda valor: str(int(valor)) if pd.notna(valor) else ''
    )

    if filtros['estado'] != 'todos':
        vista = vista[vista['estado_ui'] == filtros['estado']]

    fecha_desde_dt = pd.to_datetime(filtros.get('fecha_desde', ''), errors='coerce')
    if pd.notna(fecha_desde_dt):
        vista = vista[vista['fecha_pedido_dt'].dt.date >= fecha_desde_dt.date()]
    else:
        filtros['fecha_desde'] = ''

    fecha_hasta_dt = pd.to_datetime(filtros.get('fecha_hasta', ''), errors='coerce')
    if pd.notna(fecha_hasta_dt):
        vista = vista[vista['fecha_pedido_dt'].dt.date <= fecha_hasta_dt.date()]
    else:
        filtros['fecha_hasta'] = ''

    filtro_q = filtros.get('q', '')
    if filtro_q:
        texto = filtro_q.lower()
        vista = vista[
            vista['id_pedido_txt'].str.contains(texto, na=False)
            | vista['usuario_nombre'].fillna('').astype(str).str.lower().str.contains(texto, na=False)
            | vista['productos_pedido'].fillna('').astype(str).str.lower().str.contains(texto, na=False)
            | vista['fecha_pedido'].str.lower().str.contains(texto, na=False)
        ]

    vista = vista.sort_values(by='id_pedido', ascending=False, na_position='last')
    pagina_actual = _parse_positive_int(filtros.get('page', 1), default=1)
    total_filtrados = int(len(vista.index))
    total_paginas = max(1, (total_filtrados + per_page - 1) // per_page)
    if pagina_actual > total_paginas:
        pagina_actual = total_paginas

    inicio = (pagina_actual - 1) * per_page
    fin = inicio + per_page
    vista = vista.iloc[inicio:fin].copy()

    filtros['page'] = pagina_actual
    paginacion = {
        'page': pagina_actual,
        'per_page': per_page,
        'total': total_filtrados,
        'total_paginas': total_paginas,
        'desde': (inicio + 1) if total_filtrados > 0 else 0,
        'hasta': min(fin, total_filtrados),
        'opciones': list(range(1, total_paginas + 1)),
        'botones': _build_pagination_buttons(total_paginas, pagina_actual),
    }
    hay_filtros_activos = bool(
        filtros.get('q') or filtros.get('fecha_desde') or filtros.get('fecha_hasta') or filtros.get('estado') != 'todos'
    )
    return vista, filtros, paginacion, hay_filtros_activos


def _filtrar_y_paginar_pagos(pagos, pago_filtros, per_page=10):
    filtros = dict(pago_filtros)

    vista = pagos.copy()
    vista['metodo_pago'] = vista['metodo_pago'].fillna('').astype(str).str.strip().str.lower()
    vista['estado_pago'] = vista['estado_pago'].fillna('').astype(str).str.strip().str.lower()
    vista['fecha_pago'] = vista['fecha_pago'].fillna('').astype(str)
    vista['fecha_pago_dt'] = pd.to_datetime(vista['fecha_pago'], errors='coerce')
    vista['id_pago_txt'] = vista['id_pago'].apply(lambda valor: str(int(valor)) if pd.notna(valor) else '')
    vista['id_pedido_txt'] = vista['id_pedido'].apply(lambda valor: str(int(valor)) if pd.notna(valor) else '')
    vista['monto_txt'] = vista['monto'].apply(lambda valor: f"{float(valor):.2f}" if pd.notna(valor) else '')

    metodos_pago_opciones = sorted([m for m in vista['metodo_pago'].unique().tolist() if m])
    estados_pago_opciones = sorted([e for e in vista['estado_pago'].unique().tolist() if e])

    if filtros.get('metodo') != 'todos' and filtros.get('metodo') not in metodos_pago_opciones:
        filtros['metodo'] = 'todos'
    if filtros.get('estado') != 'todos' and filtros.get('estado') not in estados_pago_opciones:
        filtros['estado'] = 'todos'

    if filtros.get('metodo') != 'todos':
        vista = vista[vista['metodo_pago'] == filtros['metodo']]
    if filtros.get('estado') != 'todos':
        vista = vista[vista['estado_pago'] == filtros['estado']]

    pago_fecha_desde_dt = pd.to_datetime(filtros.get('fecha_desde', ''), errors='coerce')
    if pd.notna(pago_fecha_desde_dt):
        vista = vista[vista['fecha_pago_dt'].dt.date >= pago_fecha_desde_dt.date()]
    else:
        filtros['fecha_desde'] = ''

    pago_fecha_hasta_dt = pd.to_datetime(filtros.get('fecha_hasta', ''), errors='coerce')
    if pd.notna(pago_fecha_hasta_dt):
        vista = vista[vista['fecha_pago_dt'].dt.date <= pago_fecha_hasta_dt.date()]
    else:
        filtros['fecha_hasta'] = ''

    filtro_q = filtros.get('q', '')
    if filtro_q:
        pago_texto = filtro_q.lower()
        vista = vista[
            vista['id_pago_txt'].str.contains(pago_texto, na=False)
            | vista['id_pedido_txt'].str.contains(pago_texto, na=False)
            | vista['metodo_pago'].str.contains(pago_texto, na=False)
            | vista['estado_pago'].str.contains(pago_texto, na=False)
            | vista['fecha_pago'].str.lower().str.contains(pago_texto, na=False)
            | vista['monto_txt'].str.contains(pago_texto, na=False)
        ]

    vista = vista.sort_values(by='id_pago', ascending=False, na_position='last')
    pagina_actual = _parse_positive_int(filtros.get('page', 1), default=1)
    total_filtrados = int(len(vista.index))
    total_paginas = max(1, (total_filtrados + per_page - 1) // per_page)
    if pagina_actual > total_paginas:
        pagina_actual = total_paginas

    inicio = (pagina_actual - 1) * per_page
    fin = inicio + per_page
    vista = vista.iloc[inicio:fin].copy()

    filtros['page'] = pagina_actual
    paginacion = {
        'page': pagina_actual,
        'per_page': per_page,
        'total': total_filtrados,
        'total_paginas': total_paginas,
        'desde': (inicio + 1) if total_filtrados > 0 else 0,
        'hasta': min(fin, total_filtrados),
        'botones': _build_pagination_buttons(total_paginas, pagina_actual),
    }
    hay_filtros_activos = bool(
        filtros.get('q')
        or filtros.get('fecha_desde')
        or filtros.get('fecha_hasta')
        or filtros.get('metodo') != 'todos'
        or filtros.get('estado') != 'todos'
    )

    return vista, filtros, paginacion, metodos_pago_opciones, estados_pago_opciones, hay_filtros_activos


def _serializar_pedidos_admin(pedidos_view):
    lista_pedidos = pedidos_view.fillna('').to_dict(orient='records')
    for pedido in lista_pedidos:
        if pedido.get('id_pedido') != '':
            pedido['id_pedido'] = int(pedido['id_pedido'])
        pedido['estado'] = str(pedido.get('estado', 'pendiente')).strip().lower() or 'pendiente'
    return lista_pedidos


def _serializar_pagos_admin(pagos_view):
    lista_pagos = pagos_view.fillna('').to_dict(orient='records')
    for pago in lista_pagos:
        if pago.get('id_pago') != '':
            pago['id_pago'] = int(pago['id_pago'])
        if pago.get('id_pedido') != '':
            pago['id_pedido'] = int(pago['id_pedido'])
        pago['metodo_pago'] = str(pago.get('metodo_pago', '')).strip().lower()
        pago['estado_pago'] = str(pago.get('estado_pago', '')).strip().lower()
    return lista_pagos


def _construir_params_redireccion_admin_pedidos(filtros, pago_filtros):
    params = {}

    if filtros.get('q'):
        params['q'] = filtros['q']
    if filtros.get('estado') and filtros['estado'] != 'todos':
        params['estado'] = filtros['estado']
    if filtros.get('fecha_desde'):
        params['fecha_desde'] = filtros['fecha_desde']
    if filtros.get('fecha_hasta'):
        params['fecha_hasta'] = filtros['fecha_hasta']

    page = _parse_positive_int(filtros.get('page', 1), default=1)
    if page > 1:
        params['page'] = page

    if pago_filtros.get('q'):
        params['pago_q'] = pago_filtros['q']
    if pago_filtros.get('metodo') and pago_filtros['metodo'] != 'todos':
        params['pago_metodo'] = pago_filtros['metodo']
    if pago_filtros.get('estado') and pago_filtros['estado'] != 'todos':
        params['pago_estado'] = pago_filtros['estado']
    if pago_filtros.get('fecha_desde'):
        params['pago_fecha_desde'] = pago_filtros['fecha_desde']
    if pago_filtros.get('fecha_hasta'):
        params['pago_fecha_hasta'] = pago_filtros['fecha_hasta']

    pago_page = _parse_positive_int(pago_filtros.get('page', 1), default=1)
    if pago_page > 1:
        params['pago_page'] = pago_page

    return params


def _redirigir_admin_pedidos_con_filtros(filtros, pago_filtros):
    params = _construir_params_redireccion_admin_pedidos(filtros, pago_filtros)
    return redirect(url_for('admin_pedidos', **params))


@app.route('/admin/pedidos')
def admin_pedidos():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    pedidos = cargar_pedidos_df()
    pagos = cargar_pagos_df()
    detalle = cargar_detalle_pedido_df()
    productos = cargar_productos_df()
    usuarios = cargar_usuarios_df()

    pedidos, pagos, detalle, productos = _normalizar_dataframes_admin_pedidos(
        pedidos, pagos, detalle, productos
    )
    pedidos_view = _construir_vista_pedidos(pedidos, pagos, detalle, usuarios, productos)

    filtros, pago_filtros = _leer_filtros_admin_pedidos(request.args)
    pedidos_view, filtros, paginacion, hay_filtros_activos = _filtrar_y_paginar_pedidos(
        pedidos_view, filtros, per_page=10
    )
    pagos_view, pago_filtros, paginacion_pagos, metodos_pago_opciones, estados_pago_opciones, hay_filtros_pagos_activos = _filtrar_y_paginar_pagos(
        pagos, pago_filtros, per_page=10
    )

    lista_pedidos = _enriquecer_pedidos_con_tracking(_serializar_pedidos_admin(pedidos_view))
    lista_pagos = _serializar_pagos_admin(pagos_view)

    return render_template(
        'Administrador/Gestion pedidos/admin_orders.html',
        pedidos=lista_pedidos,
        pagos=lista_pagos,
        filtros=filtros,
        pago_filtros=pago_filtros,
        paginacion=paginacion,
        paginacion_pagos=paginacion_pagos,
        metodos_pago_opciones=metodos_pago_opciones,
        estados_pago_opciones=estados_pago_opciones,
        hay_filtros_activos=hay_filtros_activos,
        hay_filtros_pagos_activos=hay_filtros_pagos_activos,
    )


@app.route('/admin/pedidos/estado/<int:id_pedido>', methods=['POST'])
def admin_pedidos_estado(id_pedido):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    filtros = {
        'q': str(request.form.get('f_q', '')).strip(),
        'estado': str(request.form.get('f_estado', 'todos')).strip().lower(),
        'fecha_desde': str(request.form.get('f_fecha_desde', '')).strip(),
        'fecha_hasta': str(request.form.get('f_fecha_hasta', '')).strip(),
        'page': str(request.form.get('f_page', '1')).strip(),
    }
    pago_filtros = {
        'q': str(request.form.get('f_pago_q', '')).strip(),
        'metodo': str(request.form.get('f_pago_metodo', 'todos')).strip().lower(),
        'estado': str(request.form.get('f_pago_estado', 'todos')).strip().lower(),
        'fecha_desde': str(request.form.get('f_pago_fecha_desde', '')).strip(),
        'fecha_hasta': str(request.form.get('f_pago_fecha_hasta', '')).strip(),
        'page': str(request.form.get('f_pago_page', '1')).strip(),
    }
    ajustes_page = _parse_positive_int(request.form.get('ajustes_page', '1'), default=1)

    estado_nuevo = request.form.get('estado', '').strip().lower()
    origen = str(request.form.get('origen', '')).strip().lower()
    estados_validos = {clave for clave, _ in PEDIDO_STATUS_FLOW} | {'cancelado'}
    if estado_nuevo not in estados_validos:
        flash('Estado de pedido inválido.', 'danger')
        if origen == 'ajustes':
            return redirect(url_for('admin_ajustes', ajustes_page=ajustes_page))
        return _redirigir_admin_pedidos_con_filtros(filtros, pago_filtros)

    pedidos = cargar_pedidos_df()
    if 'id_pedido' not in pedidos.columns:
        flash('Estructura de pedidos inválida.', 'danger')
        if origen == 'ajustes':
            return redirect(url_for('admin_ajustes', ajustes_page=ajustes_page))
        return _redirigir_admin_pedidos_con_filtros(filtros, pago_filtros)
    if 'estado' not in pedidos.columns:
        pedidos['estado'] = 'pendiente'

    pedidos['id_pedido'] = pd.to_numeric(pedidos['id_pedido'], errors='coerce')
    idx = pedidos[pedidos['id_pedido'] == id_pedido].index
    if idx.empty:
        flash('Pedido no encontrado.', 'warning')
        if origen == 'ajustes':
            return redirect(url_for('admin_ajustes', ajustes_page=ajustes_page))
        return _redirigir_admin_pedidos_con_filtros(filtros, pago_filtros)

    estado_anterior = str(pedidos.at[idx[0], 'estado']).strip().lower()
    pedidos.at[idx[0], 'estado'] = estado_nuevo
    guardar_pedidos_df(pedidos)

    if estado_anterior != estado_nuevo:
        registrar_actividad(f"Actualizo estado de pedido #{id_pedido}: {estado_anterior} -> {estado_nuevo}")
        flash(f'Pedido #{id_pedido} actualizado a "{estado_nuevo}".', 'success')
    else:
        flash(f'El pedido #{id_pedido} ya estaba en "{estado_nuevo}".', 'info')

    if origen == 'ajustes':
        return redirect(url_for('admin_ajustes', ajustes_page=ajustes_page))
    return _redirigir_admin_pedidos_con_filtros(filtros, pago_filtros)

def _validar_cliente_pos(cliente_nombre, cliente_correo, cliente_documento, cliente_telefono):
    if not cliente_nombre or not cliente_correo or not cliente_documento or not cliente_telefono:
        return ('Debes completar todos los datos del cliente para registrar la venta.', 'warning')
    if not re.fullmatch(r'[A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+', cliente_nombre):
        return ('El nombre del cliente solo puede contener letras y espacios.', 'warning')
    if not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', cliente_correo):
        return ('Debes ingresar un correo electrónico válido.', 'warning')
    if not cliente_documento.isdigit():
        return ('La cédula solo puede contener números.', 'warning')
    if not cliente_telefono.isdigit() or len(cliente_telefono) > 10:
        return ('El teléfono solo puede contener números y máximo 10 dígitos.', 'warning')
    return None


def _normalizar_metodo_pago_pos(metodo_pago):
    metodo = str(metodo_pago or 'efectivo').strip().lower()
    metodos_validos = {'efectivo', 'tarjeta', 'transferencia', 'qr'}
    if metodo not in metodos_validos:
        return 'efectivo'
    return metodo


def _parsear_items_checkout_pos(items_raw):
    try:
        items = json.loads(items_raw)
    except json.JSONDecodeError:
        return None, ('No se pudo procesar el carrito POS.', 'danger')

    if not isinstance(items, list) or not items:
        return None, ('Agrega al menos un producto para cobrar.', 'warning')

    return items, None


def _validar_y_preparar_carrito_pos(items, productos, mejor_promo_por_producto):
    carrito_validado = []
    total_bruto = 0.0
    total_descuento = 0.0
    total = 0.0

    for item in items:
        try:
            id_producto = int(item.get('id_producto', 0))
            cantidad = int(item.get('cantidad', 0))
        except (TypeError, ValueError):
            return None, ('Hay productos inválidos en el carrito POS.', 'danger')

        if cantidad <= 0:
            return None, ('La cantidad debe ser mayor a cero.', 'danger')

        idx = productos[productos['id_producto'] == id_producto].index
        if idx.empty:
            return None, (f'El producto con ID {id_producto} ya no existe.', 'danger')

        row_idx = idx[0]
        if bool(productos.at[row_idx, 'eliminado']):
            return None, (f'El producto ID {id_producto} está eliminado.', 'danger')

        requiere_talla = producto_requiere_talla(productos.at[row_idx, 'intendencia'])
        talla = str(item.get('talla', '')).strip().upper()
        if requiere_talla:
            if talla not in TALLAS_OPCIONES:
                nombre = str(productos.at[row_idx, 'nombre'])
                return None, (f'Selecciona una talla valida para "{nombre}".', 'warning')
        else:
            talla = ''

        stock_actual = int(pd.to_numeric(productos.at[row_idx, 'stock'], errors='coerce') or 0)
        if cantidad > stock_actual:
            nombre = str(productos.at[row_idx, 'nombre'])
            return None, (f'Stock insuficiente para "{nombre}". Disponible: {stock_actual}.', 'warning')

        precio_base = float(pd.to_numeric(productos.at[row_idx, 'precio'], errors='coerce') or 0)
        promo = mejor_promo_por_producto.get(id_producto)
        descuento_unitario = calcular_descuento_promocion(precio_base, promo) if promo else 0.0

        subtotal_bruto = precio_base * cantidad
        subtotal_descuento = descuento_unitario * cantidad
        subtotal_final = max(0.0, subtotal_bruto - subtotal_descuento)

        total_bruto += subtotal_bruto
        total_descuento += subtotal_descuento
        total += subtotal_final

        productos.at[row_idx, 'stock'] = stock_actual - cantidad
        carrito_validado.append({
            'id_producto': id_producto,
            'cantidad': cantidad,
            'talla': talla,
            'subtotal': subtotal_final,
            'subtotal_bruto': subtotal_bruto,
            'monto_descuento': subtotal_descuento,
            'promo_id': promo.get('id_promo', '') if promo else '',
            'promo_codigo': promo.get('codigo', '') if promo else '',
            'promo_tipo_descuento': promo.get('tipo_descuento', '') if promo else '',
            'promo_valor_descuento': float(pd.to_numeric(promo.get('valor_descuento', 0), errors='coerce') or 0) if promo else 0.0
        })

    return {
        'carrito_validado': carrito_validado,
        'productos': productos,
        'total_bruto': total_bruto,
        'total_descuento': total_descuento,
        'total': total,
    }, None


def _resumen_promocion_pago_desde_carrito(carrito_validado):
    promo_ids = []
    promo_codigos = []
    promo_tipos = []
    promo_valores = []
    for item in carrito_validado:
        promo_id = str(item.get('promo_id', '')).strip()
        promo_codigo = str(item.get('promo_codigo', '')).strip()
        promo_tipo = str(item.get('promo_tipo_descuento', '')).strip()
        promo_valor = float(pd.to_numeric(item.get('promo_valor_descuento', 0), errors='coerce') or 0)

        if promo_id and promo_id not in promo_ids:
            promo_ids.append(promo_id)
        if promo_codigo and promo_codigo not in promo_codigos:
            promo_codigos.append(promo_codigo)
        if promo_tipo and promo_tipo not in promo_tipos:
            promo_tipos.append(promo_tipo)
        if promo_valor and promo_valor not in promo_valores:
            promo_valores.append(promo_valor)

    if not promo_ids:
        return {
            'id_promo': '',
            'codigo_promo': '',
            'tipo_descuento': '',
            'valor_descuento': 0.0
        }
    if len(promo_ids) == 1:
        return {
            'id_promo': promo_ids[0],
            'codigo_promo': promo_codigos[0] if promo_codigos else '',
            'tipo_descuento': promo_tipos[0] if promo_tipos else '',
            'valor_descuento': float(promo_valores[0]) if promo_valores else 0.0
        }
    return {
        'id_promo': 'multi',
        'codigo_promo': ','.join(promo_codigos),
        'tipo_descuento': 'multi',
        'valor_descuento': 0.0
    }


def _resumen_promocion_desde_promo_aplicada(promo_aplicada):
    if not promo_aplicada:
        return {
            'id_promo': '',
            'codigo_promo': '',
            'tipo_descuento': '',
            'valor_descuento': 0.0
        }

    id_promo_num = pd.to_numeric(promo_aplicada.get('id_promo'), errors='coerce')
    id_promo = int(id_promo_num) if pd.notna(id_promo_num) else ''
    return {
        'id_promo': id_promo,
        'codigo_promo': promo_aplicada.get('codigo', ''),
        'tipo_descuento': promo_aplicada.get('tipo_descuento', ''),
        'valor_descuento': float(pd.to_numeric(promo_aplicada.get('valor_descuento', 0), errors='coerce') or 0.0)
    }


def _construir_items_detalle_desde_carrito(carrito):
    items_detalle = []
    for item in carrito:
        items_detalle.append({
            'id_producto': item['id_producto'],
            'cantidad': item['cantidad'],
            'subtotal': item['subtotal'],
            'talla': str(item.get('talla', '')).strip()
        })
    return items_detalle


def _crear_pedido_y_detalle(
    pedidos,
    detalle_pedido,
    id_usuario,
    estado_pedido,
    items_detalle,
    cliente_telefono='',
    cliente_direccion='',
):
    nuevo_id_pedido = next_id('pedidos', 'id_pedido')
    nuevo_pedido = {
        'id_pedido': nuevo_id_pedido,
        'id_usuario': id_usuario,
        'fecha_pedido': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'estado': estado_pedido,
        'cliente_telefono': str(cliente_telefono or '').strip(),
        'cliente_direccion': str(cliente_direccion or '').strip(),
    }
    pedidos = pd.concat([pedidos, pd.DataFrame([nuevo_pedido])], ignore_index=True)
    guardar_pedidos_df(pedidos)

    next_detalle_id = next_id('detalle_pedido', 'id_detalle')
    nuevos_detalles = []
    for item in items_detalle:
        nuevos_detalles.append({
            'id_detalle': next_detalle_id,
            'id_pedido': nuevo_id_pedido,
            'id_producto': item['id_producto'],
            'cantidad': item['cantidad'],
            'subtotal': item['subtotal'],
            'talla': str(item.get('talla', '')).strip()
        })
        next_detalle_id += 1
    detalle_pedido = pd.concat([detalle_pedido, pd.DataFrame(nuevos_detalles)], ignore_index=True)
    guardar_detalle_pedido_df(detalle_pedido)
    return nuevo_id_pedido


def _crear_pago_para_pedido(pagos, id_pedido, monto, metodo_pago, resumen_promos, monto_descuento):
    resumen = resumen_promos or {
        'id_promo': '',
        'codigo_promo': '',
        'tipo_descuento': '',
        'valor_descuento': 0.0
    }

    nuevo_id_pago = next_id('pagos', 'id_pago')
    nuevo_pago = {
        'id_pago': nuevo_id_pago,
        'id_pedido': id_pedido,
        'monto': monto,
        'metodo_pago': metodo_pago,
        'fecha_pago': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'estado_pago': 'aprobado',
        'id_promo': resumen.get('id_promo', ''),
        'codigo_promo': resumen.get('codigo_promo', ''),
        'tipo_descuento': resumen.get('tipo_descuento', ''),
        'valor_descuento': resumen.get('valor_descuento', 0.0),
        'monto_descuento': monto_descuento
    }
    pagos = pd.concat([pagos, pd.DataFrame([nuevo_pago])], ignore_index=True)
    guardar_pagos_df(pagos)
    return nuevo_id_pago


def _registrar_venta_pos_admin(carrito_validado, metodo_pago, total, total_descuento):
    pedidos = cargar_pedidos_df()
    detalle_pedido = cargar_detalle_pedido_df()
    pagos = cargar_pagos_df()
    items_detalle = _construir_items_detalle_desde_carrito(carrito_validado)
    next_pedido_id = _crear_pedido_y_detalle(
        pedidos=pedidos,
        detalle_pedido=detalle_pedido,
        id_usuario=session.get('usuario', 'admin_pos'),
        estado_pedido='completado',
        items_detalle=items_detalle
    )

    resumen_promos = _resumen_promocion_pago_desde_carrito(carrito_validado)
    _crear_pago_para_pedido(
        pagos=pagos,
        id_pedido=next_pedido_id,
        monto=round(total, 2),
        metodo_pago=metodo_pago,
        resumen_promos={
            'id_promo': resumen_promos['id_promo'],
            'codigo_promo': resumen_promos['codigo_promo'],
            'tipo_descuento': resumen_promos['tipo_descuento'],
            'valor_descuento': round(float(pd.to_numeric(resumen_promos['valor_descuento'], errors='coerce') or 0.0), 2)
        },
        monto_descuento=round(total_descuento, 2)
    )

    return next_pedido_id


@app.route('/admin/pos/checkout', methods=['POST'])
def admin_pos_checkout():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    cliente_nombre = request.form.get('cliente_nombre', '').strip()
    cliente_correo = request.form.get('cliente_correo', '').strip().lower()
    cliente_documento = request.form.get('cliente_documento', '').strip()
    cliente_telefono = request.form.get('cliente_telefono', '').strip()
    error_cliente = _validar_cliente_pos(
        cliente_nombre,
        cliente_correo,
        cliente_documento,
        cliente_telefono
    )
    if error_cliente:
        flash(error_cliente[0], error_cliente[1])
        return redirect(url_for('admin_pos'))

    items_raw = request.form.get('items_json', '[]')
    metodo_pago = _normalizar_metodo_pago_pos(request.form.get('metodo_pago', 'efectivo'))

    items, error_items = _parsear_items_checkout_pos(items_raw)
    if error_items:
        flash(error_items[0], error_items[1])
        return redirect(url_for('admin_pos'))

    productos = cargar_productos_df()
    if productos.empty:
        flash('No existe el catálogo de productos.', 'danger')
        return redirect(url_for('admin_pos'))

    promos = cargar_promociones_df()
    hoy = datetime.now().date()
    mejor_promo_por_producto = obtener_mejor_promocion_por_producto(productos, promos, hoy)

    resultado_carrito, error_carrito = _validar_y_preparar_carrito_pos(
        items,
        productos,
        mejor_promo_por_producto
    )
    if error_carrito:
        flash(error_carrito[0], error_carrito[1])
        return redirect(url_for('admin_pos'))

    carrito_validado = resultado_carrito['carrito_validado']
    total_bruto = resultado_carrito['total_bruto']
    total_descuento = resultado_carrito['total_descuento']
    total = resultado_carrito['total']
    guardar_productos_df(resultado_carrito['productos'])

    next_pedido_id = _registrar_venta_pos_admin(
        carrito_validado,
        metodo_pago,
        total,
        total_descuento
    )

    total_bruto = round(total_bruto, 2)
    total_descuento = round(total_descuento, 2)
    total = round(total, 2)
    total_cop = formatear_cop(total)
    total_bruto_cop = formatear_cop(total_bruto)
    descuento_cop = formatear_cop(total_descuento)
    registrar_actividad(
        f"POS registro venta #{next_pedido_id} por {total_cop} ({len(carrito_validado)} producto(s))\n"
        f"- total bruto: {total_bruto_cop}\n"
        f"- descuento aplicado: {descuento_cop}\n"
        f"- cliente: {cliente_nombre}\n"
        f"- correo: {cliente_correo}\n"
        f"- documento: {cliente_documento}\n"
        f"- teléfono: {cliente_telefono}"
    )
    if total_descuento > 0:
        flash(
            f'Venta POS registrada. Pedido #{next_pedido_id} por {total_cop}. '
            f'Descuento aplicado: {descuento_cop}.',
            'success'
        )
    else:
        flash(f'Venta POS registrada. Pedido #{next_pedido_id} por {total_cop}.', 'success')

    session['ultimo_recibo_pos'] = {
        'id_pedido': int(next_pedido_id),
        'cliente_nombre': cliente_nombre,
        'cliente_correo': cliente_correo,
        'cliente_documento': cliente_documento,
        'cliente_telefono': cliente_telefono
    }
    session.modified = True
    return redirect(url_for('admin_pos', descargar_recibo=next_pedido_id))


@app.route('/admin/editar/<int:id_producto>', methods=['GET', 'POST'])
def editar_producto(id_producto):
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    productos = cargar_productos_df()
    producto = productos[productos['id_producto'] == id_producto]

    if producto.empty:
        return "Producto no encontrado."

    if request.method == 'POST':
        idx = producto.index[0]

        anterior = {
            'nombre': str(productos.at[idx, 'nombre']) if 'nombre' in productos.columns else '',
            'descripcion': str(productos.at[idx, 'descripcion']) if 'descripcion' in productos.columns else '',
            'precio': float(pd.to_numeric(productos.at[idx, 'precio'], errors='coerce')) if 'precio' in productos.columns else 0.0,
            'stock': int(pd.to_numeric(productos.at[idx, 'stock'], errors='coerce')) if 'stock' in productos.columns else 0,
            'id_categoria': str(productos.at[idx, 'id_categoria']) if 'id_categoria' in productos.columns else '',
            'fuerza': str(productos.at[idx, 'fuerza']) if 'fuerza' in productos.columns else '',
            'intendencia': str(productos.at[idx, 'intendencia']) if 'intendencia' in productos.columns else ''
        }

        nuevo_nombre = request.form['nombre'].strip()
        nueva_descripcion = request.form['descripcion'].strip()
        nuevo_precio = float(request.form['precio'])
        nuevo_stock = int(request.form['stock'])
        nueva_fuerza = request.form.get('fuerza', anterior['fuerza']).strip()
        nueva_intendencia = request.form.get('intendencia', anterior['intendencia']).strip()
        if nueva_fuerza not in FUERZAS_OPCIONES or nueva_intendencia not in INTENDENCIAS_OPCIONES:
            flash('Selecciona una fuerza e intendencia válidas.', 'danger')
            return redirect(url_for('admin_productos'))

        nueva_categoria = anterior['id_categoria']
        if 'id_categoria' in productos.columns:
            categoria_form = request.form.get('id_categoria', '').strip()
            if categoria_form:
                try:
                    nueva_categoria = int(float(categoria_form))
                except ValueError:
                    nueva_categoria = anterior['id_categoria']
            # Evitar errores de tipo en pandas (int64): asegurar valor numerico
            nueva_categoria_num = pd.to_numeric(nueva_categoria, errors='coerce')
            if pd.isna(nueva_categoria_num):
                nueva_categoria_num = pd.to_numeric(productos.at[idx, 'id_categoria'], errors='coerce')
            nueva_categoria = int(nueva_categoria_num) if pd.notna(nueva_categoria_num) else 0

        productos.at[idx, 'nombre'] = nuevo_nombre
        productos.at[idx, 'descripcion'] = nueva_descripcion
        productos.at[idx, 'precio'] = nuevo_precio
        productos.at[idx, 'stock'] = nuevo_stock
        productos.at[idx, 'fuerza'] = nueva_fuerza
        productos.at[idx, 'intendencia'] = nueva_intendencia
        if 'id_categoria' in productos.columns:
            productos.at[idx, 'id_categoria'] = nueva_categoria

        guardar_productos_df(productos)

        cambios = []
        if anterior['nombre'] != nuevo_nombre:
            cambios.append(f"- nombre: '{anterior['nombre']}' -> '{nuevo_nombre}'")
        if anterior['descripcion'] != nueva_descripcion:
            cambios.append(f"- descripción: '{anterior['descripcion']}' -> '{nueva_descripcion}'")
        if round(float(anterior['precio']), 2) != round(float(nuevo_precio), 2):
            cambios.append(f"- precio: {formatear_cop(anterior['precio'])} -> {formatear_cop(nuevo_precio)}")
        if int(anterior['stock']) != int(nuevo_stock):
            cambios.append(f"- stock: {anterior['stock']} -> {nuevo_stock}")
        if anterior['fuerza'] != nueva_fuerza:
            cambios.append(f"- fuerza: {anterior['fuerza']} -> {nueva_fuerza}")
        if anterior['intendencia'] != nueva_intendencia:
            cambios.append(f"- intendencia: {anterior['intendencia']} -> {nueva_intendencia}")
        if str(anterior['id_categoria']) != str(nueva_categoria):
            cambios.append(f"- categoria: {anterior['id_categoria']} -> {nueva_categoria}")

        encabezado = f"Actualizo producto '{nuevo_nombre}' (ID {id_producto})"
        if cambios:
            registrar_actividad(encabezado + "\n" + "\n".join(cambios))
        else:
            registrar_actividad(encabezado + "\n- sin cambios detectados")

        return redirect(url_for('admin_productos'))

    # Render formulario de edición (plantilla propia en Gestion productos)
    return render_template(
        'Administrador/Gestion productos/editar_producto.html',
        producto=producto.iloc[0],
        fuerzas=FUERZAS_OPCIONES,
        intendencias=INTENDENCIAS_OPCIONES
    )

    

@app.route('/user')
def user_dashboard():
    if session.get('rol') != 'normal':
        flash('Debes iniciar sesión como usuario para acceder al catálogo.', 'warning')
        return redirect(url_for('login'))

    productos = cargar_productos_activos_df()
    lista_productos = productos.to_dict(orient='records')

    # Promoción vigente por producto para mostrar en catálogo
    promos = cargar_promociones_df()
    hoy = datetime.now().date()
    mejor_promo_por_producto = obtener_mejor_promocion_por_producto(productos, promos, hoy)

    for producto in lista_productos:
        precio_base = float(pd.to_numeric(producto.get('precio', 0), errors='coerce') or 0)
        promo = mejor_promo_por_producto.get(int(producto.get('id_producto', 0)))
        if promo:
            descuento = calcular_descuento_promocion(precio_base, promo)
            producto['precio_original'] = precio_base
            producto['precio_con_descuento'] = max(0.0, precio_base - descuento)
            producto['promo_activa'] = True
            producto['promo_nombre'] = promo.get('nombre', '')
            if promo.get('tipo_descuento') == 'valor_fijo':
                producto['promo_etiqueta'] = f"-{formatear_cop(promo.get('valor_descuento', 0))}"
            else:
                valor_pct = float(pd.to_numeric(promo.get('valor_descuento', 0), errors='coerce') or 0)
                producto['promo_etiqueta'] = f"-{valor_pct:g}%"
        else:
            producto['precio_original'] = precio_base
            producto['precio_con_descuento'] = precio_base
            producto['promo_activa'] = False
            producto['promo_nombre'] = ''
            producto['promo_etiqueta'] = ''
        
        # Galería para destacados (hasta 5 imágenes: 1 principal + 4 miniaturas; rellena con la principal si faltan)
        galeria = obtener_galeria_producto(
            int(producto.get('id_producto', 0)),
            producto.get('imagen_url', '')
        )
        if not galeria:
            galeria = [producto.get('imagen_url', '')]
        while len(galeria) < 5:
            galeria.append(galeria[0])
        producto['galeria_dashboard'] = galeria[:5]
    
    # Mostrar en los carruseles solo las prendas marcadas como destacadas en panel admin, filtradas por categoría.
    productos_destacados_ejercito = [p for p in lista_productos if bool(p.get('destacado_dashboard', False)) and p.get('fuerza') == 'Ejercito'][:5]
    productos_destacados_policia = [p for p in lista_productos if bool(p.get('destacado_dashboard', False)) and p.get('fuerza') == 'Policia'][:5]
    productos_destacados_armada = [p for p in lista_productos if bool(p.get('destacado_dashboard', False)) and p.get('fuerza') == 'Armada'][:5]
    
    # Obtener el contador del carrito
    carrito = _obtener_carrito_sesion_usuario()
    cart_count = len(carrito)
    
    return render_template(
        'Usuarios/user_dashboard.html',
        productos=lista_productos,
        productos_destacados_ejercito=productos_destacados_ejercito,
        productos_destacados_policia=productos_destacados_policia,
        productos_destacados_armada=productos_destacados_armada,
        cart_count=cart_count
    )


@app.route('/product/<int:id_producto>')
def product_detail(id_producto):
    if session.get('rol') == 'normal':
        productos = cargar_productos_activos_df()
        producto = productos[productos['id_producto'] == id_producto]
        
        if producto.empty:
            return "Producto no encontrado"
        
        producto_dict = producto.iloc[0].to_dict()
        galeria = obtener_galeria_producto(id_producto, producto_dict.get('imagen_url', ''))[:MAX_IMAGES_PER_PRODUCT]
        producto_dict['imagenes'] = galeria
        if galeria:
            producto_dict['imagen_url'] = galeria[0]
        producto_dict['requiere_talla'] = bool(producto_requiere_talla(producto_dict.get('intendencia', '')))
        
        # Obtener el contador del carrito
        carrito = _obtener_carrito_sesion_usuario()
        cart_count = len(carrito)
        
        return render_template(
            'Usuarios/Detalle pedido/product_detail.html',
            producto=producto_dict,
            cart_count=cart_count,
            tallas_opciones=TALLAS_OPCIONES
        )
    return "Acceso denegado"


@app.route('/add_to_cart/<int:id_producto>', methods=['POST'])
def add_to_cart(id_producto):
    if session.get('rol') != 'normal':
        return "Acceso denegado"

    productos = cargar_productos_activos_df()

    try:
        cantidad = int(request.form.get('cantidad', 1))
    except (TypeError, ValueError):
        cantidad = 1

    if cantidad <= 0:
        flash('Cantidad inválida.', 'warning')
        return redirect(url_for('user_dashboard'))

    producto_df = productos[productos['id_producto'] == id_producto]
    if producto_df.empty:
        flash('El producto no existe o ya no está disponible.', 'warning')
        return redirect(url_for('user_dashboard'))

    producto = producto_df.iloc[0]
    stock_actual = int(producto.get('stock', 0))
    if stock_actual <= 0:
        flash(f'El producto "{producto["nombre"]}" está agotado.', 'warning')
        return redirect(url_for('user_dashboard'))
    if cantidad > stock_actual:
        flash(f'Solo hay {stock_actual} unidad(es) disponibles de "{producto["nombre"]}".', 'warning')
        return redirect(url_for('user_dashboard'))

    carrito = _obtener_carrito_sesion_usuario()
    talla = str(request.form.get('talla', '')).strip().upper()
    requiere_talla = producto_requiere_talla(producto.get('intendencia', ''))
    if requiere_talla:
        if talla not in TALLAS_OPCIONES:
            flash('Selecciona una talla valida antes de agregar al carrito.', 'warning')
            return redirect(url_for('product_detail', id_producto=id_producto))
    else:
        talla = ''

    galeria_producto = obtener_galeria_producto(id_producto, producto.get('imagen_url', ''))
    imagen_principal = normalizar_imagen_url(galeria_producto[0]) if galeria_producto else ''

    item_existente = next(
        (
            item
            for item in carrito
            if int(item.get('id_producto', 0)) == int(id_producto)
            and str(item.get('talla', '')).strip().upper() == talla
        ),
        None
    )
    if item_existente:
        nueva_cantidad = int(item_existente.get('cantidad', 0)) + cantidad
        if nueva_cantidad > stock_actual:
            flash(f'No puedes agregar más de {stock_actual} unidad(es) de "{producto["nombre"]}".', 'warning')
            return redirect(url_for('cart'))
        item_existente['cantidad'] = nueva_cantidad
        item_existente['subtotal'] = float(item_existente['precio']) * nueva_cantidad
        if not str(item_existente.get('imagen_url', '') or '').strip() and imagen_principal:
            item_existente['imagen_url'] = imagen_principal
    else:
        item_carrito = {
            'id_producto': int(id_producto),
            'nombre': producto['nombre'],
            'descripcion': str(producto.get('descripcion', '') or '').strip(),
            'cantidad': cantidad,
            'precio': float(producto['precio']),
            'subtotal': float(producto['precio']) * cantidad,
            'talla': talla,
            'imagen_url': imagen_principal
        }
        carrito.append(item_carrito)

    session['carrito'] = carrito
    session.modified = True
    _sincronizar_carrito_usuario_desde_sesion()

    flash(f'{producto["nombre"]} agregado al carrito correctamente.', 'success')

    referer = request.referrer
    if referer and 'product' in referer:
        return redirect(url_for('product_detail', id_producto=id_producto))

    return redirect(url_for('user_dashboard'))

@app.route('/cart')
def cart():
    if session.get('rol') == 'normal':
        metodos_validos = {'tarjeta', 'transferencia'}
        metodo_pago = str(request.args.get('metodo_pago', '') or '').strip().lower()
        if metodo_pago not in metodos_validos:
            metodo_pago = str(session.get('checkout_metodo_preferido', '') or '').strip().lower()
        if metodo_pago not in metodos_validos:
            metodo_pago = ''

        session['checkout_metodo_preferido'] = metodo_pago
        if 'codigo_promo' in request.args:
            codigo_promo = str(request.args.get('codigo_promo', '') or '').strip().upper()
        else:
            codigo_promo = str(session.get('checkout_codigo_promo', '') or '').strip().upper()
        session['checkout_codigo_promo'] = codigo_promo

        carrito = _obtener_carrito_sesion_usuario()
        carrito_enriquecido = _enriquecer_carrito_con_imagenes(carrito)
        if carrito_enriquecido != carrito:
            session['carrito'] = carrito_enriquecido
            session.modified = True
            _sincronizar_carrito_usuario_desde_sesion()
        total = sum(float(item.get('subtotal', 0)) for item in carrito_enriquecido)
        contacto_cliente = _obtener_contacto_checkout_predeterminado()
        return render_template(
            'Usuarios/Carrito/cart.html',
            carrito=carrito_enriquecido,
            total=total,
            selected_metodo_pago=metodo_pago,
            selected_codigo_promo=codigo_promo,
            cliente_telefono=contacto_cliente['telefono'],
            cliente_direccion=contacto_cliente['direccion'],
        )
    return "Acceso denegado"

@app.route('/get_cart_count')
def get_cart_count():
    carrito = _obtener_carrito_sesion_usuario()
    count = len(carrito)
    return {'count': count}

@app.route('/cart/remove/<int:index>', methods=['POST'])
def remove_from_cart(index):
    if session.get('rol') == 'normal':
        carrito = _obtener_carrito_sesion_usuario()
        if 0 <= index < len(carrito):
            carrito.pop(index)
            session['carrito'] = carrito
            session.modified = True
            _sincronizar_carrito_usuario_desde_sesion()
    return redirect(url_for('cart'))

@app.route('/checkout')
def checkout():
    if session.get('rol') == 'normal':
        metodos_validos = {'transferencia'}
        metodo_pago = str(request.args.get('metodo_pago', '') or '').strip().lower()
        if metodo_pago not in metodos_validos:
            metodo_pago = 'transferencia'
        if metodo_pago not in metodos_validos:
            metodo_pago = 'transferencia'

        session['checkout_metodo_preferido'] = metodo_pago
        if 'codigo_promo' in request.args:
            codigo_promo = str(request.args.get('codigo_promo', '') or '').strip().upper()
        else:
            codigo_promo = str(session.get('checkout_codigo_promo', '') or '').strip().upper()
        session['checkout_codigo_promo'] = codigo_promo

        carrito = _obtener_carrito_sesion_usuario()
        if not carrito:
            flash('No hay productos en el carrito.', 'warning')
            return redirect(url_for('cart'))

        total = sum(float(item.get('subtotal', 0)) for item in carrito)
        contacto_cliente = _obtener_contacto_checkout_predeterminado()
        return render_template(
            'Usuarios/Carrito/checkout.html',
            carrito=carrito,
            total=total,
            selected_metodo_pago=metodo_pago,
            selected_codigo_promo=codigo_promo,
            cliente_telefono=contacto_cliente['telefono'],
            cliente_direccion=contacto_cliente['direccion'],
        )
    return "Acceso denegado"


def _asegurar_columnas_descuento_pagos(pagos):
    if 'id_promo' not in pagos.columns:
        pagos['id_promo'] = pd.Series(dtype='float')
    if 'codigo_promo' not in pagos.columns:
        pagos['codigo_promo'] = ''
    if 'tipo_descuento' not in pagos.columns:
        pagos['tipo_descuento'] = ''
    if 'valor_descuento' not in pagos.columns:
        pagos['valor_descuento'] = 0.0
    if 'monto_descuento' not in pagos.columns:
        pagos['monto_descuento'] = 0.0
    return pagos


def _resolver_promocion_checkout(codigo_promo, promos, total):
    if not codigo_promo:
        return None, 0.0, None

    promo_aplicada = buscar_promocion_por_codigo(promos, codigo_promo, datetime.now().date())
    if promo_aplicada is None:
        return None, 0.0, ('El código promocional no es válido o no está vigente.', 'warning')

    descuento_promo = calcular_descuento_promocion(total, promo_aplicada)
    return promo_aplicada, descuento_promo, None


def _obtener_contacto_checkout_predeterminado():
    telefono = str(session.get('checkout_cliente_telefono', '') or '').strip()
    direccion = str(session.get('checkout_cliente_direccion', '') or '').strip()
    usuario_email = normalizar_email(session.get('usuario', ''))

    if not usuario_email:
        return {'telefono': telefono, 'direccion': direccion}

    usuarios = cargar_usuarios_df()
    if usuarios.empty:
        return {'telefono': telefono, 'direccion': direccion}

    usuario = usuarios[usuarios['email'] == usuario_email]
    if usuario.empty:
        return {'telefono': telefono, 'direccion': direccion}

    usuario_dict = usuario.iloc[0].to_dict()
    if not telefono:
        telefono = str(usuario_dict.get('telefono', '') or '').strip()
    if not direccion:
        direccion = str(usuario_dict.get('direccion', '') or '').strip()

    return {'telefono': telefono, 'direccion': direccion}


def _validar_datos_cliente_checkout(telefono, direccion):
    telefono_limpio = re.sub(r"\D", "", str(telefono or ""))
    direccion_limpia = re.sub(r"\s+", " ", str(direccion or "")).strip()

    if not telefono_limpio or not direccion_limpia:
        return None, None, (
            'Antes de finalizar tu compra debes registrar el telefono y la direccion de entrega.',
            'warning'
        )

    if len(telefono_limpio) < 7 or len(telefono_limpio) > 10:
        return None, None, (
            'El telefono debe contener solo numeros y tener entre 7 y 10 digitos.',
            'warning'
        )

    return telefono_limpio, direccion_limpia, None


def _guardar_contacto_checkout_usuario(telefono, direccion):
    usuario_email = normalizar_email(session.get('usuario', ''))
    if not usuario_email:
        return

    usuarios = cargar_usuarios_df()
    if 'telefono' not in usuarios.columns:
        usuarios['telefono'] = ''
    if 'direccion' not in usuarios.columns:
        usuarios['direccion'] = ''

    idx = usuarios[usuarios['email'] == usuario_email].index
    if idx.empty:
        return

    usuarios.loc[idx, 'telefono'] = str(telefono or '').strip()
    usuarios.loc[idx, 'direccion'] = str(direccion or '').strip()
    guardar_usuarios_df(usuarios)


def _validar_stock_checkout(productos, carrito):
    for item in carrito:
        id_producto = int(item.get('id_producto', 0))
        cantidad = int(item.get('cantidad', 0))
        fila = productos[(productos['id_producto'] == id_producto) & (productos['eliminado'] == False)]
        if fila.empty:
            return (
                f'El producto "{item.get("nombre", id_producto)}" ya no esta disponible o fue retirado del catalogo.',
                'warning'
            )
        stock_actual = int(fila.iloc[0].get('stock', 0))
        if stock_actual <= 0:
            nombre = str(fila.iloc[0].get('nombre', item.get('nombre', id_producto)))
            return (f'El producto "{nombre}" esta agotado y no se puede procesar el pedido.', 'warning')
        if cantidad > stock_actual:
            nombre = str(fila.iloc[0].get('nombre', item.get('nombre', id_producto)))
            return (
                f'Stock insuficiente para "{nombre}". Disponible: {stock_actual}. '
                'Ajusta la cantidad para continuar.',
                'warning'
            )
    return None


def _descontar_stock_checkout(productos, carrito):
    agotados_en_compra = []
    for item in carrito:
        id_producto = int(item.get('id_producto', 0))
        cantidad = int(item.get('cantidad', 0))
        idx = productos[productos['id_producto'] == id_producto].index[0]
        stock_anterior = int(productos.at[idx, 'stock'])
        nuevo_stock = stock_anterior - cantidad
        productos.at[idx, 'stock'] = nuevo_stock
        if stock_anterior > 0 and nuevo_stock == 0:
            agotados_en_compra.append(str(productos.at[idx, 'nombre']))
    return agotados_en_compra


def _registrar_compra_checkout_usuario(
    carrito,
    pedidos,
    detalle_pedido,
    pagos,
    metodo_pago,
    total_final,
    promo_aplicada,
    descuento_promo,
    estado_pedido='pendiente',
    cliente_telefono='',
    cliente_direccion='',
):
    items_detalle = _construir_items_detalle_desde_carrito(carrito)
    nuevo_id_pedido = _crear_pedido_y_detalle(
        pedidos=pedidos,
        detalle_pedido=detalle_pedido,
        id_usuario=session.get('id_usuario', session['usuario']),
        estado_pedido=estado_pedido,
        items_detalle=items_detalle,
        cliente_telefono=cliente_telefono,
        cliente_direccion=cliente_direccion,
    )

    resumen_promos = _resumen_promocion_desde_promo_aplicada(promo_aplicada)
    _crear_pago_para_pedido(
        pagos=pagos,
        id_pedido=nuevo_id_pedido,
        monto=total_final,
        metodo_pago=metodo_pago,
        resumen_promos=resumen_promos,
        monto_descuento=float(descuento_promo)
    )
    return nuevo_id_pedido


def _iniciar_pago_stripe_desde_carrito(carrito, codigo_promo, total_final):
    ok_stripe, msg_stripe = stripe_estado_configuracion()
    if not ok_stripe:
        flash(
            f'El pago con tarjeta no esta disponible en este momento. Detalle: {msg_stripe}',
            'warning'
        )
        return redirect(url_for('cart', metodo_pago='tarjeta', codigo_promo=codigo_promo))

    cart_hash = _hash_carrito_checkout(carrito)
    try:
        stripe_session = crear_checkout_sesion_tarjeta(
            amount_total=total_final,
            success_url=url_for('pay_stripe_success', _external=True),
            cancel_url=url_for('pay_stripe_cancel', _external=True),
            customer_email=session.get('usuario', ''),
            descripcion=f"Compra NACHOHER ({len(carrito)} item(s))",
            metadata={
                "usuario": session.get('usuario', ''),
                "codigo_promo": codigo_promo,
                "cart_hash": cart_hash,
                "total_esperado": f"{total_final:.2f}",
            },
        )
    except Exception as exc:
        app.logger.exception("Error creando sesion Stripe Checkout: %s", exc)
        flash('No fue posible iniciar el pago con tarjeta. Intenta de nuevo.', 'danger')
        return redirect(url_for('cart', metodo_pago='tarjeta', codigo_promo=codigo_promo))

    stripe_session_id = str(_stripe_obj_get(stripe_session, "id", "") or "").strip()
    stripe_session_url = str(_stripe_obj_get(stripe_session, "url", "") or "").strip()
    if not stripe_session_id or not stripe_session_url:
        flash('Stripe no devolvio una sesion valida para continuar el pago.', 'danger')
        return redirect(url_for('cart', metodo_pago='tarjeta', codigo_promo=codigo_promo))

    _stripe_checkout_guardar_creado(
        session_id=stripe_session_id,
        usuario_email=session.get('usuario', ''),
        codigo_promo=codigo_promo,
        carrito=carrito,
        cart_hash=cart_hash,
        total_esperado=total_final,
    )
    return redirect(stripe_session_url, code=303)


@app.route('/checkout/select', methods=['POST'])
def checkout_select():
    if session.get('rol') != 'normal':
        return "Acceso denegado"

    carrito = _obtener_carrito_sesion_usuario()
    if not carrito:
        flash('No hay productos en el carrito.', 'warning')
        return redirect(url_for('cart'))

    metodo_pago = str(request.form.get('metodo_pago', '') or '').strip().lower()
    codigo_promo = str(request.form.get('codigo_promo', '') or '').strip().upper()
    cliente_telefono, cliente_direccion, error_cliente = _validar_datos_cliente_checkout(
        request.form.get('cliente_telefono', ''),
        request.form.get('cliente_direccion', '')
    )
    session['checkout_cliente_telefono'] = str(request.form.get('cliente_telefono', '') or '').strip()
    session['checkout_cliente_direccion'] = str(request.form.get('cliente_direccion', '') or '').strip()
    metodos_validos = {'tarjeta', 'transferencia'}
    if metodo_pago not in metodos_validos:
        flash('Selecciona un metodo de pago valido.', 'warning')
        return redirect(url_for('cart', codigo_promo=codigo_promo))
    if error_cliente:
        flash(error_cliente[0], error_cliente[1])
        return redirect(url_for('cart', metodo_pago=metodo_pago, codigo_promo=codigo_promo))

    session['checkout_metodo_preferido'] = metodo_pago
    session['checkout_codigo_promo'] = codigo_promo
    session['checkout_cliente_telefono'] = cliente_telefono
    session['checkout_cliente_direccion'] = cliente_direccion
    _guardar_contacto_checkout_usuario(cliente_telefono, cliente_direccion)

    total = sum(float(item.get('subtotal', 0)) for item in carrito)
    promos = cargar_promociones_df()
    promo_aplicada, descuento_promo, error_promo = _resolver_promocion_checkout(codigo_promo, promos, total)
    if error_promo:
        flash(error_promo[0], error_promo[1])
        return redirect(url_for('cart', metodo_pago=metodo_pago, codigo_promo=codigo_promo))

    productos = cargar_productos_df()
    if productos.empty:
        flash('No existe la base de productos.', 'danger')
        return redirect(url_for('cart', metodo_pago=metodo_pago, codigo_promo=codigo_promo))

    error_stock = _validar_stock_checkout(productos, carrito)
    if error_stock:
        flash(error_stock[0], error_stock[1])
        return redirect(url_for('cart', metodo_pago=metodo_pago, codigo_promo=codigo_promo))

    total_final = max(0.0, float(total) - float(descuento_promo))
    if metodo_pago == 'tarjeta':
        return _iniciar_pago_stripe_desde_carrito(carrito, codigo_promo, total_final)

    return redirect(url_for('checkout', metodo_pago='transferencia', codigo_promo=codigo_promo))


@app.route('/pay', methods=['POST'])
def pay():
    if session.get('rol') != 'normal':
        return "Acceso denegado"

    carrito = _obtener_carrito_sesion_usuario()
    if not carrito:
        flash('No hay productos en el carrito.', 'warning')
        return redirect(url_for('cart'))

    total = sum(float(item.get('subtotal', 0)) for item in carrito)
    codigo_promo = request.form.get('codigo_promo', '').strip().upper()
    metodo_pago = str(request.form.get('metodo_pago', '') or '').strip().lower()
    cliente_telefono_input = request.form.get('cliente_telefono', '')
    cliente_direccion_input = request.form.get('cliente_direccion', '')
    if not str(cliente_telefono_input or '').strip() or not str(cliente_direccion_input or '').strip():
        contacto_cliente = _obtener_contacto_checkout_predeterminado()
        cliente_telefono_input = contacto_cliente.get('telefono', '')
        cliente_direccion_input = contacto_cliente.get('direccion', '')
    cliente_telefono, cliente_direccion, error_cliente = _validar_datos_cliente_checkout(
        cliente_telefono_input,
        cliente_direccion_input
    )
    metodos_validos = {'tarjeta', 'transferencia', 'efectivo'}
    session['checkout_metodo_preferido'] = metodo_pago if metodo_pago in metodos_validos else ''
    session['checkout_codigo_promo'] = codigo_promo
    session['checkout_cliente_telefono'] = str(cliente_telefono_input or '').strip()
    session['checkout_cliente_direccion'] = str(cliente_direccion_input or '').strip()
    if error_cliente:
        flash(error_cliente[0], error_cliente[1])
        return redirect(url_for('checkout', metodo_pago='transferencia', codigo_promo=codigo_promo))
    session['checkout_cliente_telefono'] = cliente_telefono
    session['checkout_cliente_direccion'] = cliente_direccion
    _guardar_contacto_checkout_usuario(cliente_telefono, cliente_direccion)

    pagos = cargar_pagos_df()
    pagos = _asegurar_columnas_descuento_pagos(pagos)
    pedidos = cargar_pedidos_df()
    detalle_pedido = cargar_detalle_pedido_df()

    productos = cargar_productos_df()
    if productos.empty:
        flash('No existe la base de productos.', 'danger')
        return redirect(url_for('cart'))

    promos = cargar_promociones_df()
    promo_aplicada, descuento_promo, error_promo = _resolver_promocion_checkout(codigo_promo, promos, total)
    if error_promo:
        flash(error_promo[0], error_promo[1])
        return redirect(url_for('cart', metodo_pago=metodo_pago, codigo_promo=codigo_promo))

    total_final = max(0.0, float(total) - float(descuento_promo))
    error_stock = _validar_stock_checkout(productos, carrito)
    if error_stock:
        flash(error_stock[0], error_stock[1])
        return redirect(url_for('cart', metodo_pago=metodo_pago, codigo_promo=codigo_promo))

    if metodo_pago == 'tarjeta':
        return _iniciar_pago_stripe_desde_carrito(carrito, codigo_promo, total_final)
    agotados_en_compra = _descontar_stock_checkout(productos, carrito)
    guardar_productos_df(productos)

    nuevo_id_pedido = _registrar_compra_checkout_usuario(
        carrito=carrito,
        pedidos=pedidos,
        detalle_pedido=detalle_pedido,
        pagos=pagos,
        metodo_pago=metodo_pago,
        total_final=total_final,
        promo_aplicada=promo_aplicada,
        descuento_promo=descuento_promo,
        cliente_telefono=cliente_telefono,
        cliente_direccion=cliente_direccion,
    )

    registrar_actividad(
        f"Creo pedido #{nuevo_id_pedido} con {len(carrito)} producto(s) por {formatear_cop(total_final)}"
    )
    if agotados_en_compra:
        registrar_actividad("Stock agotado por pedido: " + ", ".join(agotados_en_compra))

    session['carrito'] = []
    session.modified = True
    _sincronizar_carrito_usuario_desde_sesion()

    metodo_pago_nombres = {
        'tarjeta': 'Tarjeta de Credito/Debito',
        'transferencia': 'Transferencia Bancaria',
        'efectivo': 'Efectivo',
        'paypal': 'PayPal'
    }

    return render_template(
        'Usuarios/Carrito/order_confirmation.html',
        pedido_id=nuevo_id_pedido,
        metodo_pago=metodo_pago_nombres.get(metodo_pago, metodo_pago),
        fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
        total=total_final,
        promo_codigo=promo_aplicada.get('codigo', '') if promo_aplicada else None,
        descuento=descuento_promo if promo_aplicada else 0
    )


@app.route('/pay/stripe/success')
def pay_stripe_success():
    if session.get('rol') != 'normal':
        return "Acceso denegado"

    session_id = str(request.args.get('session_id', '') or '').strip()
    if not session_id:
        flash('Stripe no envió el identificador de la sesión de pago.', 'warning')
        return redirect(url_for('cart', metodo_pago='tarjeta'))

    registro = _stripe_checkout_obtener(session_id)
    if not registro:
        flash('No se encontro el registro local del pago con tarjeta.', 'warning')
        return redirect(url_for('cart', metodo_pago='tarjeta'))

    usuario_registro = normalizar_email(registro.get('usuario_email', ''))
    usuario_sesion = normalizar_email(session.get('usuario', ''))
    if usuario_registro and usuario_sesion and usuario_registro != usuario_sesion:
        flash('La sesión de pago no pertenece al usuario autenticado.', 'danger')
        return redirect(url_for('cart', metodo_pago='tarjeta'))

    estado_actual = str(registro.get('estado', '') or '').strip().lower()
    id_pedido_existente = pd.to_numeric(registro.get('id_pedido'), errors='coerce')
    if estado_actual == 'pagado' and pd.notna(id_pedido_existente):
        flash(f'El pago ya fue procesado en el pedido #{int(id_pedido_existente)}.', 'info')
        return redirect(url_for('user_orders'))

    try:
        stripe_session = obtener_checkout_sesion(session_id)
    except Exception as exc:
        app.logger.exception("Error validando sesión Stripe %s: %s", session_id, exc)
        flash('No se pudo validar el pago con Stripe. Intenta nuevamente en unos segundos.', 'danger')
        return redirect(url_for('cart', metodo_pago='tarjeta'))

    payment_status = str(_stripe_obj_get(stripe_session, 'payment_status', '') or '').strip().lower()
    if payment_status != 'paid':
        _stripe_checkout_marcar_estado(session_id, 'pendiente')
        flash('El pago aun no aparece como confirmado por Stripe.', 'warning')
        return redirect(url_for('cart', metodo_pago='tarjeta'))

    carrito_checkout = _stripe_checkout_cargar_carrito(registro)
    if not carrito_checkout:
        carrito_checkout = _obtener_carrito_sesion_usuario()
    if not carrito_checkout:
        _stripe_checkout_marcar_estado(session_id, 'pagado_sin_carrito')
        flash(
            'El pago se confirmo, pero no se encontro el carrito asociado. '
            'Contacta al administrador con el id de sesión.',
            'danger'
        )
        return redirect(url_for('user_orders'))

    hash_registrado = str(registro.get('cart_hash', '') or '').strip()
    hash_actual = _hash_carrito_checkout(carrito_checkout)
    if hash_registrado and hash_registrado != hash_actual:
        app.logger.warning(
            "Diferencia de hash de carrito en Stripe session %s (registro=%s, actual=%s)",
            session_id,
            hash_registrado,
            hash_actual,
        )

    subtotal_checkout = sum(float(item.get('subtotal', 0)) for item in carrito_checkout)
    total_esperado = float(pd.to_numeric(registro.get('total_esperado', 0), errors='coerce') or 0.0)
    total_final = round(total_esperado if total_esperado > 0 else subtotal_checkout, 2)
    descuento_promo = round(max(0.0, float(subtotal_checkout) - float(total_final)), 2)
    contacto_cliente = _obtener_contacto_checkout_predeterminado()
    cliente_telefono, cliente_direccion, error_cliente = _validar_datos_cliente_checkout(
        contacto_cliente.get('telefono', ''),
        contacto_cliente.get('direccion', '')
    )
    if error_cliente:
        app.logger.warning(
            "Pedido Stripe %s confirmado sin datos completos del cliente. telefono=%r direccion=%r",
            session_id,
            contacto_cliente.get('telefono', ''),
            contacto_cliente.get('direccion', ''),
        )
        cliente_telefono = str(contacto_cliente.get('telefono', '') or '').strip()
        cliente_direccion = str(contacto_cliente.get('direccion', '') or '').strip()

    codigo_promo = str(registro.get('codigo_promo', '') or '').strip().upper()
    promo_aplicada = None
    if codigo_promo:
        promos = cargar_promociones_df()
        promo_aplicada = buscar_promocion_por_codigo(promos, codigo_promo, datetime.now().date())
        if promo_aplicada is None:
            promo_aplicada = {
                'id_promo': '',
                'codigo': codigo_promo,
                'tipo_descuento': '',
                'valor_descuento': 0.0,
            }

    productos = cargar_productos_df()
    pedidos = cargar_pedidos_df()
    detalle_pedido = cargar_detalle_pedido_df()
    pagos = _asegurar_columnas_descuento_pagos(cargar_pagos_df())

    estado_pedido = 'pendiente'
    agotados_en_compra = []
    error_stock = _validar_stock_checkout(productos, carrito_checkout)
    if error_stock:
        estado_pedido = 'pendiente_revision'
        registrar_actividad(
            f"Pago Stripe {session_id} confirmado con stock en conflicto. Pedido quedo en revision."
        )
    else:
        agotados_en_compra = _descontar_stock_checkout(productos, carrito_checkout)
        guardar_productos_df(productos)

    nuevo_id_pedido = _registrar_compra_checkout_usuario(
        carrito=carrito_checkout,
        pedidos=pedidos,
        detalle_pedido=detalle_pedido,
        pagos=pagos,
        metodo_pago='tarjeta',
        total_final=total_final,
        promo_aplicada=promo_aplicada,
        descuento_promo=descuento_promo,
        estado_pedido=estado_pedido,
        cliente_telefono=cliente_telefono,
        cliente_direccion=cliente_direccion,
    )
    _stripe_checkout_marcar_estado(session_id, 'pagado', nuevo_id_pedido)

    registrar_actividad(
        f"Stripe confirmado. Pedido #{nuevo_id_pedido} creado por {formatear_cop(total_final)}"
    )
    if agotados_en_compra:
        registrar_actividad("Stock agotado por pedido: " + ", ".join(agotados_en_compra))

    session['carrito'] = []
    session.modified = True
    _sincronizar_carrito_usuario_desde_sesion()

    if error_stock:
        flash(
            'El pago se registro correctamente, pero hay cambios de stock. '
            f'Tu pedido #{nuevo_id_pedido} quedo en revision.',
            'warning'
        )

    return render_template(
        'Usuarios/Carrito/order_confirmation.html',
        pedido_id=nuevo_id_pedido,
        metodo_pago='Tarjeta de Credito/Debito',
        fecha=datetime.now().strftime("%d/%m/%Y %H:%M"),
        total=total_final,
        promo_codigo=codigo_promo or None,
        descuento=descuento_promo if descuento_promo > 0 else 0
    )


@app.route('/pay/stripe/cancel')
def pay_stripe_cancel():
    if session.get('rol') != 'normal':
        return "Acceso denegado"
    flash('Pago con tarjeta cancelado. Tu carrito sigue guardado.', 'info')
    return redirect(url_for('cart', metodo_pago='tarjeta'))


@app.route('/user/orders')
def user_orders():
    if session.get('rol') != 'normal':
        return "Acceso denegado"
    
    pedidos = cargar_pedidos_df()
    id_usuario = pd.to_numeric(session.get('id_usuario'), errors='coerce')
    pedidos_usuario = pd.DataFrame(columns=PEDIDO_COLUMNS)
    if not pedidos.empty and pd.notna(id_usuario):
        pedidos = pedidos.copy()
        pedidos['id_usuario_num'] = pd.to_numeric(pedidos['id_usuario'], errors='coerce')
        pedidos_usuario = pedidos[pedidos['id_usuario_num'] == id_usuario].copy()
    
    # Obtener montos de pagos
    if not pedidos_usuario.empty:
        pagos = cargar_pagos_df()
        pedidos_usuario = pd.merge(pedidos_usuario, pagos[['id_pedido', 'monto', 'metodo_pago']], 
                                   on='id_pedido', how='left')
        pedidos_usuario = pedidos_usuario.sort_values(by='id_pedido', ascending=False, na_position='last')
    
    lista_pedidos = _enriquecer_pedidos_con_tracking(pedidos_usuario.fillna('').to_dict(orient='records'))
    return render_template('Usuarios/Carrito/user_orders.html', pedidos=lista_pedidos)

@app.route('/user/orders/details/<int:id_pedido>')
def user_order_details(id_pedido):
    if session.get('rol') != 'normal':
        return "Acceso denegado"
    
    # Verificar que el pedido pertenece al usuario
    pedidos = cargar_pedidos_df()
    pedidos = pedidos.copy()
    pedidos['id_pedido_num'] = pd.to_numeric(pedidos['id_pedido'], errors='coerce')
    pedidos['id_usuario_num'] = pd.to_numeric(pedidos['id_usuario'], errors='coerce')
    id_usuario = pd.to_numeric(session.get('id_usuario'), errors='coerce')
    pedido = pedidos[
        (pedidos['id_pedido_num'] == id_pedido) &
        (pedidos['id_usuario_num'] == id_usuario)
    ]
    
    if pedido.empty:
        return "Pedido no encontrado o no tienes permiso para verlo"
    
    detalle_pedido = cargar_detalle_pedido_df()
    detalles = detalle_pedido[detalle_pedido['id_pedido'] == id_pedido]
    
    productos = cargar_productos_df()
    detalles = pd.merge(detalles, productos[['id_producto', 'nombre', 'precio']], on='id_producto')
    if 'talla' not in detalles.columns:
        detalles['talla'] = ''

    pedido_info = _enriquecer_pedidos_con_tracking([pedido.iloc[0].to_dict()])[0]
    
    return render_template('Usuarios/Informacion compras pedido/user_order_details.html',
                          pedido=pedido_info,
                          detalles=detalles.to_dict(orient='records'))

@app.route('/user/profile')
def user_profile():
    if session.get('rol') != 'normal':
        return "Acceso denegado"
    
    # Cargar informacion del usuario
    usuarios = cargar_usuarios_df()
    usuario_email = session.get('usuario')
    usuario = usuarios[usuarios['email'] == usuario_email]
    
    if usuario.empty:
        return "Usuario no encontrado"
    
    usuario_dict = usuario.iloc[0].to_dict()
    
    # Asegurar que existan las columnas telefono y direccion
    if 'telefono' not in usuario_dict:
        usuario_dict['telefono'] = ''
    if 'direccion' not in usuario_dict:
        usuario_dict['direccion'] = ''
    
    # Asegurar que existan las columnas de verificación
    if 'email_verified' not in usuario_dict:
        usuario_dict['email_verified'] = False
    
    # Asegurar que existan las columnas de configuracion con valores por defecto
    config_defaults = {
        'notif_email': True,
        'notif_pedidos': True,
        'notif_promociones': True,
        'perfil_publico': False,
        'mostrar_historial': False,
        'idioma': 'es',
        'moneda': 'COP'
    }
    for key, default_value in config_defaults.items():
        if key not in usuario_dict or pd.isna(usuario_dict[key]):
            usuario_dict[key] = default_value
    
    # Obtener estadísticas del usuario
    pedidos_total = 0
    gasto_total = 0.0
    
    pedidos = cargar_pedidos_df()
    if not pedidos.empty:
        id_usuario = session.get('id_usuario', usuario_email)
        pedidos_usuario = pedidos[pedidos['id_usuario'] == id_usuario]
        pedidos_total = len(pedidos_usuario)
        
        if not pedidos_usuario.empty:
            pagos = cargar_pagos_df()
            pagos_usuario = pagos[pagos['id_pedido'].isin(pedidos_usuario['id_pedido'])]
            gasto_total = pagos_usuario['monto'].sum() if not pagos_usuario.empty else 0.0
    
    return render_template('Usuarios/Ajustes/user_profile.html', 
                          usuario=usuario_dict,
                          pedidos_total=pedidos_total,
                          gasto_total=gasto_total,
                          PASSWORD_CHANGE_CODE_EXP_MINUTES=PASSWORD_CHANGE_CODE_EXP_MINUTES)

@app.route('/user/profile/update', methods=['POST'])
def update_profile():
    if session.get('rol') != 'normal':
        flash('Acceso denegado', 'danger')
        return redirect(url_for('home'))
    
    # Obtener datos del formulario
    nombre = request.form.get('nombre')
    telefono = request.form.get('telefono', '')
    direccion = request.form.get('direccion', '')
    
    # Cargar usuarios
    usuarios = cargar_usuarios_df()
    usuario_email = session.get('usuario')
    
    # Asegurar que existan las columnas
    if 'telefono' not in usuarios.columns:
        usuarios['telefono'] = ''
    if 'direccion' not in usuarios.columns:
        usuarios['direccion'] = ''
    
    # Actualizar informacion
    idx = usuarios[usuarios['email'] == usuario_email].index
    if not idx.empty:
        usuarios.loc[idx, 'nombre'] = nombre
        usuarios.loc[idx, 'telefono'] = telefono
        usuarios.loc[idx, 'direccion'] = direccion
        
        guardar_usuarios_df(usuarios)
        registrar_actividad(f"Usuario {usuario_email} actualizo su perfil")
        flash('Perfil actualizado correctamente', 'success')
    else:
        flash('Error al actualizar el perfil', 'danger')
    
    return redirect(url_for('user_profile'))

@app.route('/user/profile/send-password-change-code', methods=['POST'])
def send_password_change_code():
    if session.get('rol') != 'normal':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403

    try:
        usuarios = cargar_usuarios_df()
        usuarios = asegurar_columnas_cambio_password(usuarios)

        usuario_email = normalizar_email(session.get('usuario', ''))
        usuario_idx = usuarios[usuarios['email'] == usuario_email].index
        if usuario_idx.empty:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404

        codigo = generar_codigo_verificacion()
        expiry = datetime.now() + timedelta(minutes=PASSWORD_CHANGE_CODE_EXP_MINUTES)

        usuarios.loc[usuario_idx, 'password_change_code'] = str(codigo)
        usuarios.loc[usuario_idx, 'password_change_code_expiry'] = expiry.strftime('%Y-%m-%d %H:%M:%S')
        guardar_usuarios_df(usuarios)

        envio_ok = enviar_codigo_verificacion(
            usuario_email,
            codigo,
            tipo='autenticacion',
            minutos_expiracion=PASSWORD_CHANGE_CODE_EXP_MINUTES
        )
        if not envio_ok:
            limpiar_codigo_cambio_password(usuarios, usuario_idx[0])
            guardar_usuarios_df(usuarios)
            return jsonify({
                'success': False,
                'message': 'No fue posible enviar el código al correo. Intenta nuevamente.'
            }), 500

        registrar_actividad(f"Código de cambio de contraseña enviado a {usuario_email}")
        return jsonify({
            'success': True,
            'message': (
                f'Código enviado a {usuario_email}. '
                f'Es válido por {PASSWORD_CHANGE_CODE_EXP_MINUTES} minutos.'
            )
        })
    except Exception as e:
        print(f"Error enviando código de cambio de contraseña: {str(e)}")
        return jsonify({'success': False, 'message': 'Error interno enviando el código.'}), 500


@app.route('/user/profile/verify-password-change-code', methods=['POST'])
def verify_password_change_code():
    if session.get('rol') != 'normal':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403

    try:
        payload = request.get_json(silent=True) or {}
        codigo_ingresado = str(payload.get('code', '') or '').strip()

        if not re.fullmatch(r'\d{6}', codigo_ingresado):
            return jsonify({
                'success': False,
                'message': 'El código debe tener 6 dígitos numéricos.'
            }), 400

        usuarios = cargar_usuarios_df()
        usuarios = asegurar_columnas_cambio_password(usuarios)
        usuario_email = normalizar_email(session.get('usuario', ''))
        usuario_idx = usuarios[usuarios['email'] == usuario_email].index

        if usuario_idx.empty:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404

        idx = usuario_idx[0]
        usuario = usuarios.loc[idx]
        codigo_guardado = str(usuario.get('password_change_code', '') or '').replace('.0', '').strip()
        if not codigo_guardado:
            return jsonify({
                'success': False,
                'message': 'No hay un código activo. Solicita uno nuevo.'
            }), 400

        if codigo_cambio_password_expirado(usuario):
            limpiar_codigo_cambio_password(usuarios, idx)
            guardar_usuarios_df(usuarios)
            return jsonify({
                'success': False,
                'message': 'El código expiró. Solicita uno nuevo.'
            }), 400

        if codigo_guardado != codigo_ingresado:
            return jsonify({
                'success': False,
                'message': 'El código es incorrecto. Verifica e intenta nuevamente.'
            }), 400

        return jsonify({
            'success': True,
            'message': 'Código validado. Ya puedes continuar con el cambio de contraseña.'
        })
    except Exception as e:
        print(f"Error verificando código de cambio de contraseña: {str(e)}")
        return jsonify({'success': False, 'message': 'Error interno verificando el código.'}), 500


@app.route('/user/profile/change-password', methods=['POST'])
def change_password():
    if session.get('rol') != 'normal':
        flash('Acceso denegado', 'danger')
        return redirect(url_for('home'))

    current_password = request.form.get('current_password', '')
    new_password = request.form.get('new_password', '')
    password_change_code = request.form.get('password_change_code', '').strip()

    if not password_change_code:
        flash('Debes ingresar el código de seguridad enviado a tu correo.', 'danger')
        return redirect(url_for('user_profile'))

    if not re.fullmatch(r'\d{6}', password_change_code):
        flash('El código de seguridad debe tener 6 dígitos numéricos.', 'danger')
        return redirect(url_for('user_profile'))

    if not current_password or not new_password:
        flash('Debes completar todos los campos para cambiar la contraseña.', 'danger')
        return redirect(url_for('user_profile'))

    if not password_cumple_estandares(new_password):
        flash(
            'La contraseña debe tener mínimo 8 caracteres, mayúscula, minúscula, número y carácter especial.',
            'danger'
        )
        return redirect(url_for('user_profile'))

    if current_password == new_password:
        flash('La nueva contraseña debe ser diferente a la actual.', 'warning')
        return redirect(url_for('user_profile'))

    usuarios = cargar_usuarios_df()
    usuarios = asegurar_columnas_cambio_password(usuarios)
    usuario_email = normalizar_email(session.get('usuario', ''))
    usuario_idx = usuarios[usuarios['email'] == usuario_email].index

    if usuario_idx.empty:
        flash('Usuario no encontrado', 'danger')
        return redirect(url_for('user_profile'))

    idx = usuario_idx[0]
    usuario = usuarios.loc[idx]

    codigo_guardado = str(usuario.get('password_change_code', '') or '').replace('.0', '').strip()
    if not codigo_guardado:
        flash('No hay un código activo. Solicita uno nuevo antes de cambiar la contraseña.', 'warning')
        return redirect(url_for('user_profile'))

    if codigo_cambio_password_expirado(usuario):
        limpiar_codigo_cambio_password(usuarios, idx)
        guardar_usuarios_df(usuarios)
        flash('El código de seguridad expiró. Solicita uno nuevo.', 'warning')
        return redirect(url_for('user_profile'))

    if codigo_guardado != password_change_code:
        flash('El código de seguridad es incorrecto.', 'danger')
        return redirect(url_for('user_profile'))

    password_guardado = usuario.get('password_hash', '')
    if not password_coincide(password_guardado, current_password):
        flash('La contraseña actual es incorrecta.', 'danger')
        return redirect(url_for('user_profile'))

    usuarios.at[idx, 'password_hash'] = crear_hash_password(new_password)
    usuarios.at[idx, 'reset_token'] = ''
    usuarios.at[idx, 'reset_token_expiry'] = ''
    limpiar_codigo_cambio_password(usuarios, idx)

    guardar_usuarios_df(usuarios)
    registrar_actividad(f"Usuario {usuario_email} cambió su contraseña con código de seguridad")
    flash('Contraseña cambiada correctamente.', 'success')
    return redirect(url_for('user_profile'))

@app.route('/user/profile/settings', methods=['POST'])
def update_settings():
    if session.get('rol') != 'normal':
        flash('Acceso denegado', 'danger')
        return redirect(url_for('home'))
    
    # Obtener datos del formulario
    notif_email = request.form.get('notif_email') == 'on'
    notif_pedidos = request.form.get('notif_pedidos') == 'on'
    notif_promociones = request.form.get('notif_promociones') == 'on'
    perfil_publico = request.form.get('perfil_publico') == 'on'
    mostrar_historial = request.form.get('mostrar_historial') == 'on'
    idioma = request.form.get('idioma', 'es')
    moneda = request.form.get('moneda', 'COP')
    
    # Cargar usuarios
    usuarios = cargar_usuarios_df()
    usuario_email = session.get('usuario')
    
    # Asegurar que existan las columnas de configuracion
    columnas_config = ['notif_email', 'notif_pedidos', 'notif_promociones', 
                       'perfil_publico', 'mostrar_historial', 'idioma', 'moneda']
    for col in columnas_config:
        if col not in usuarios.columns:
            usuarios[col] = '' if col in ['idioma', 'moneda'] else False
    
    # Actualizar configuracion
    idx = usuarios[usuarios['email'] == usuario_email].index
    if not idx.empty:
        usuarios.loc[idx, 'notif_email'] = notif_email
        usuarios.loc[idx, 'notif_pedidos'] = notif_pedidos
        usuarios.loc[idx, 'notif_promociones'] = notif_promociones
        usuarios.loc[idx, 'perfil_publico'] = perfil_publico
        usuarios.loc[idx, 'mostrar_historial'] = mostrar_historial
        usuarios.loc[idx, 'idioma'] = idioma
        usuarios.loc[idx, 'moneda'] = moneda
        
        guardar_usuarios_df(usuarios)
        registrar_actividad(f"Usuario {usuario_email} actualizo sus ajustes")
        flash('Ajustes guardados correctamente', 'success')
    else:
        flash('Error al guardar los ajustes', 'danger')
    
    return redirect(url_for('user_profile'))

@app.route('/user/send-verification-code', methods=['POST'])
def send_verification_code():
    """Envía un código de verificación al correo del usuario (funciona siempre)."""
    if session.get('rol') != 'normal':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403
    
    try:
        # Cargar usuarios
        usuarios = cargar_usuarios_df()
        usuario_email = session.get('usuario')
        
        # Asegurar que existan las columnas de verificación
        if 'email_verified' not in usuarios.columns:
            usuarios['email_verified'] = False
        if 'verification_code' not in usuarios.columns:
            usuarios['verification_code'] = ''
        if 'verification_code_expiry' not in usuarios.columns:
            usuarios['verification_code_expiry'] = ''
        
        # Verificar si el usuario existe
        usuario_idx = usuarios[usuarios['email'] == usuario_email].index
        if usuario_idx.empty:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404
        
        # Generar código de verificación (sin importar si ya está verificado)
        codigo = generar_codigo_verificacion()
        expiry = datetime.now() + timedelta(minutes=REGISTER_CODE_EXP_MINUTES)
        
        # Guardar código en la base de datos (forzar como string)
        usuarios.loc[usuario_idx, 'verification_code'] = str(codigo)
        usuarios.loc[usuario_idx, 'verification_code_expiry'] = expiry.strftime('%Y-%m-%d %H:%M:%S')
        
        # Asegurar que la columna sea tipo string
        usuarios['verification_code'] = usuarios['verification_code'].astype(str).str.replace('.0', '', regex=False).replace('nan', '')
        
        guardar_usuarios_df(usuarios)

        if enviar_codigo_verificacion(
            usuario_email,
            codigo,
            tipo='autenticacion',
            minutos_expiracion=REGISTER_CODE_EXP_MINUTES
        ):
            registrar_actividad(f"Código de autenticación enviado a {usuario_email}")
            return jsonify({
                'success': True, 
                'message': 'Código enviado correctamente. Revisa tu correo.'
            })

        usuarios.loc[usuario_idx, 'verification_code'] = ''
        usuarios.loc[usuario_idx, 'verification_code_expiry'] = ''
        guardar_usuarios_df(usuarios)
        return jsonify({
            'success': False,
            'message': (
                'No fue posible enviar el código de verificación al correo indicado. '
                'Verifica la configuración SMTP del sistema e intenta nuevamente.'
            )
        }), 500
            
    except Exception as e:
        print(f"Error al enviar código: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Error: {str(e)}'
        }), 500

@app.route('/user/verify-email', methods=['POST'])
def verify_email():
    """Verifica el código ingresado por el usuario (funciona siempre)."""
    if session.get('rol') != 'normal':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403
    
    try:
        codigo_ingresado = request.json.get('code', '').strip()
        
        if not codigo_ingresado:
            return jsonify({'success': False, 'message': 'Debe ingresar un código'}), 400
        
        # Cargar usuarios
        usuarios = cargar_usuarios_df()
        usuario_email = session.get('usuario')
        
        # Asegurar que existan las columnas
        if 'email_verified' not in usuarios.columns:
            usuarios['email_verified'] = False
        if 'verification_code' not in usuarios.columns:
            usuarios['verification_code'] = ''
        if 'verification_code_expiry' not in usuarios.columns:
            usuarios['verification_code_expiry'] = ''
        
        usuario_idx = usuarios[usuarios['email'] == usuario_email].index
        if usuario_idx.empty:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404
        
        usuario = usuarios.loc[usuario_idx[0]]
        
        # Verificar si existe un código
        if not usuario['verification_code']:
            return jsonify({'success': False, 'message': 'No hay código de verificación. Solicita uno nuevo.'}), 400
        
        # Verificar si el código expiró
        if timestamp_expirado(usuario.get('verification_code_expiry', '')):
            return jsonify({'success': False, 'message': 'El código ha expirado. Solicita uno nuevo.'}), 400
        
        # Verificar el código (limpiar .0 por si acaso)
        codigo_guardado = str(usuario['verification_code']).replace('.0', '').strip()
        if codigo_guardado == str(codigo_ingresado):
            # Código correcto
            # Marcar como verificado si no lo estaba
            if not usuario['email_verified']:
                usuarios.loc[usuario_idx, 'email_verified'] = True
                mensaje_exito = '¡Correo verificado exitosamente!'
                actividad = f"Usuario {usuario_email} verificó su correo electrónico"
            else:
                mensaje_exito = '¡Autenticación exitosa!'
                actividad = f"Usuario {usuario_email} se autenticó correctamente"
            
            # Limpiar código usado
            usuarios.loc[usuario_idx, 'verification_code'] = ''
            usuarios.loc[usuario_idx, 'verification_code_expiry'] = ''
            guardar_usuarios_df(usuarios)
            
            registrar_actividad(actividad)
            return jsonify({
                'success': True, 
                'message': mensaje_exito
            })
        else:
            return jsonify({
                'success': False, 
                'message': 'Código incorrecto. Verifica e intenta nuevamente.'
            }), 400
            
    except Exception as e:
        print(f"Error al verificar código: {str(e)}")
        return jsonify({
            'success': False, 
            'message': 'Error interno del servidor'
        }), 500

# Ruta para el redireccionamiento para ver el catálogo
@app.route('/armada')
def armada():
    productos = cargar_productos_por_fuerza('Armada')
    return render_template('Usuarios/catalogo/armada.html', productos=productos)


@app.route('/policia')
def policia():
    productos = cargar_productos_por_fuerza('Policia')
    return render_template('Usuarios/catalogo/policia.html', productos=productos)


@app.route('/gaula')
def gaula():
    productos = cargar_productos_por_fuerza('Gaula')
    return render_template('Usuarios/catalogo/gaula.html', productos=productos)


@app.route('/ejercito')
def ejercito():
    productos = cargar_productos_por_fuerza('Ejercito')
    return render_template('Usuarios/catalogo/ejercito.html', productos=productos)

@app.route('/producto/<int:id_producto>')
def producto_detalle(id_producto):
    productos = cargar_productos_activos_df()

    producto = productos[productos['id_producto'] == id_producto]
    if producto.empty:
        return "Producto no encontrado"

    producto_dict = producto.iloc[0].to_dict()
    galeria = obtener_galeria_producto(id_producto, producto_dict.get('imagen_url', ''))[:MAX_IMAGES_PER_PRODUCT]
    producto_dict['imagenes'] = galeria
    if galeria:
        producto_dict['imagen_url'] = galeria[0]
    producto_dict['requiere_talla'] = bool(producto_requiere_talla(producto_dict.get('intendencia', '')))

    carrito = _obtener_carrito_sesion_usuario() if session.get('rol') == 'normal' else []
    cart_count = len(carrito)
    return render_template(
        'Usuarios/Detalle pedido/product_detail.html',
        producto=producto_dict,
        cart_count=cart_count,
        tallas_opciones=TALLAS_OPCIONES
    )

@app.route('/admin/promo', methods=['GET','POST'])
def admin_promo():
    # Página de gestión de promociones para administradores
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    if request.method == 'POST':
        id_producto = pd.to_numeric(request.form.get('id_producto', ''), errors='coerce')
        nombre = request.form.get('nombre', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        tipo_descuento = request.form.get('tipo_descuento', 'porcentaje').strip().lower()
        if tipo_descuento not in ['porcentaje', 'valor_fijo']:
            tipo_descuento = 'porcentaje'
        try:
            valor_descuento = float(request.form.get('valor_descuento', 0))
        except ValueError:
            valor_descuento = 0
        if valor_descuento < 0:
            valor_descuento = 0
        if tipo_descuento == 'porcentaje':
            valor_descuento = min(valor_descuento, 100)

        codigo = request.form.get('codigo', '').strip().upper()
        fecha_inicio = request.form.get('fecha_inicio', '')
        fecha_fin = request.form.get('fecha_fin', '')
        activo = request.form.get('activo') == 'on'
        inicio_date = parsear_fecha_promocion(fecha_inicio)
        fin_date = parsear_fecha_promocion(fecha_fin)
        if inicio_date and fin_date and inicio_date > fin_date:
            flash('La fecha de inicio no puede ser mayor que la fecha de fin.', 'warning')
            return redirect(url_for('admin_promo'))
        if pd.isna(id_producto):
            flash('Debes seleccionar un producto para la promoción.', 'warning')
            return redirect(url_for('admin_promo'))
        productos_ref = cargar_productos_activos_df()
        id_producto_num = int(id_producto)
        existe_producto = not productos_ref[
            pd.to_numeric(productos_ref.get('id_producto', 0), errors='coerce') == id_producto_num
        ].empty
        if not existe_producto:
            flash('El producto seleccionado no existe o está inactivo.', 'warning')
            return redirect(url_for('admin_promo'))

        promos = cargar_promociones_df()
        if codigo:
            existe_codigo = promos[promos['codigo'].astype(str).str.upper() == codigo]
            if not existe_codigo.empty:
                flash('El código promocional ya existe. Usa otro.', 'warning')
                return redirect(url_for('admin_promo'))
        next_id = int(pd.to_numeric(promos['id_promo'], errors='coerce').max() + 1) if not promos.empty else 1
        nuevo = {
            'id_promo': next_id,
            'nombre': nombre,
            'descripcion': descripcion,
            'tipo_descuento': tipo_descuento,
            'valor_descuento': valor_descuento,
            'id_producto': int(id_producto),
            'codigo': codigo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'activo': activo
        }
        promos = pd.concat([promos, pd.DataFrame([nuevo])], ignore_index=True)
        guardar_promociones_df(promos)
        detalle_desc = formatear_cop(valor_descuento) if tipo_descuento == 'valor_fijo' else f"{valor_descuento:.2f}%"
        registrar_actividad(
            f"Promoción creada: {nombre}\n"
            f"- producto ID: {int(id_producto)}\n"
            f"- descuento: {detalle_desc}\n"
            f"- código: {codigo or 'N/A'}"
        )
        return redirect(url_for('admin_promo'))

    promos = cargar_promociones_df()
    hoy = datetime.now().date()
    productos = cargar_productos_activos_df()
    productos['id_producto'] = pd.to_numeric(productos.get('id_producto', 0), errors='coerce').fillna(0).astype(int)
    lista_productos = productos.to_dict(orient='records')
    productos_por_fuerza = {fuerza: [] for fuerza in FUERZAS_OPCIONES}

    for producto in lista_productos:
        imagen_principal = str(producto.get('imagen_url', '')).strip()
        galeria = obtener_galeria_producto(producto.get('id_producto'), imagen_principal)
        producto['imagenes'] = galeria
        if galeria:
            producto['imagen_url'] = galeria[0]

        fuerza_producto = str(producto.get('fuerza', '')).strip().lower()
        for fuerza in FUERZAS_OPCIONES:
            if fuerza_producto == fuerza.lower():
                productos_por_fuerza[fuerza].append(producto)
                break

    pagos = cargar_pagos_df()
    if not pagos.empty:
        if 'id_promo' not in pagos.columns:
            pagos['id_promo'] = pd.Series(dtype='float')
        if 'monto_descuento' not in pagos.columns:
            pagos['monto_descuento'] = 0.0
        pagos['id_promo'] = pd.to_numeric(pagos['id_promo'], errors='coerce')
        pagos['monto_descuento'] = pd.to_numeric(pagos['monto_descuento'], errors='coerce').fillna(0.0)
        impacto = pagos.dropna(subset=['id_promo']).groupby('id_promo', as_index=False).agg(
            usos=('id_pago', 'count'),
            descuento_total=('monto_descuento', 'sum')
        )
    else:
        impacto = pd.DataFrame(columns=['id_promo', 'usos', 'descuento_total'])

    lista_promos = promos.to_dict(orient='records')
    impacto_map = {}
    for _, row in impacto.iterrows():
        impacto_map[int(row['id_promo'])] = {
            'usos': int(row['usos']),
            'descuento_total': float(row['descuento_total'])
        }
    for p in lista_promos:
        p['estado_vigencia'] = estado_vigencia_promocion(p, hoy)
        p['es_aplicable'] = promocion_esta_aplicable(p, hoy)
        p['id_producto'] = int(pd.to_numeric(p.get('id_producto'), errors='coerce') or 0)
        key = int(pd.to_numeric(p.get('id_promo'), errors='coerce') or 0)
        metrica = impacto_map.get(key, {'usos': 0, 'descuento_total': 0.0})
        p['usos'] = metrica['usos']
        p['descuento_total'] = metrica['descuento_total']
    nombre_producto = {int(row['id_producto']): str(row.get('nombre', 'Producto')) for _, row in productos.iterrows()}
    for p in lista_promos:
        p['producto_nombre'] = nombre_producto.get(p['id_producto'], 'Producto no encontrado')
    return render_template(
        'Administrador/Promociones/adim_promo.html',
        promos=lista_promos,
        productos=lista_productos,
        productos_por_fuerza=productos_por_fuerza,
        fuerzas=FUERZAS_OPCIONES
    )


@app.route('/admin/promo/toggle/<int:id_promo>', methods=['POST'])
def admin_promo_toggle(id_promo):
    if session.get('rol') != 'admin':
        return "Acceso denegado"
    promos = cargar_promociones_df()
    idx = promos[promos['id_promo'] == id_promo].index
    if not idx.empty:
        promos.loc[idx, 'activo'] = ~promos.loc[idx, 'activo']
        guardar_promociones_df(promos)
        registrar_actividad(f"Promocion {'activada' if promos.loc[idx, 'activo'].iloc[0] else 'desactivada'}: {promos.loc[idx, 'nombre'].iloc[0]}")
    return redirect(url_for('admin_promo'))


@app.route('/accesorios')
def accesorios():
    productos = cargar_productos_por_intendencia('Accesorios')
    return render_template('Usuarios/catalogo/accesorios.html', productos=productos)


@app.route('/admin/charts')
def admin_charts():
    if session.get('rol') == 'admin':
        periodo = request.args.get('periodo', 'all').strip().lower()
        fecha_desde_raw = request.args.get('fecha_desde', '').strip()
        fecha_hasta_raw = request.args.get('fecha_hasta', '').strip()
        datos = obtener_datos_charts(periodo, fecha_desde_raw, fecha_hasta_raw)

        return render_template(
            'Administrador/Informes/admin_charts_dashboard.html',
            ventas_producto=datos['ventas_producto'],
            ventas_mes=datos['ventas_mes'],
            metodos_pago=datos['metodos_pago'],
            top_productos=datos['top_productos'],
            kpis=datos['kpis'],
            kpi_variaciones=datos['kpi_variaciones'],
            filtros=datos['filtros'],
            rango_texto=datos['rango_texto'],
            comparacion_texto=datos['comparacion_texto']
        )
    return "Acceso denegado"


def obtener_datos_charts(periodo, fecha_desde_raw, fecha_hasta_raw):
    detalle = cargar_detalle_pedido_df()
    productos = cargar_productos_df()
    pedidos = cargar_pedidos_df()
    pagos = cargar_pagos_df()

    if 'id_producto' not in detalle.columns:
        detalle['id_producto'] = pd.Series(dtype='int')
    if 'cantidad' not in detalle.columns:
        detalle['cantidad'] = 0
    if 'subtotal' not in detalle.columns:
        detalle['subtotal'] = 0
    if 'id_pedido' not in detalle.columns:
        detalle['id_pedido'] = pd.Series(dtype='int')

    detalle['cantidad'] = pd.to_numeric(detalle['cantidad'], errors='coerce').fillna(0)
    detalle['subtotal'] = pd.to_numeric(detalle['subtotal'], errors='coerce').fillna(0)

    if 'id_producto' not in productos.columns:
        productos['id_producto'] = pd.Series(dtype='int')
    if 'nombre' not in productos.columns:
        productos['nombre'] = ''
    if 'precio' not in productos.columns:
        productos['precio'] = 0
    productos['precio'] = pd.to_numeric(productos['precio'], errors='coerce').fillna(0)

    if 'id_pedido' not in pedidos.columns:
        pedidos['id_pedido'] = pd.Series(dtype='int')
    if 'fecha_pedido' not in pedidos.columns:
        pedidos['fecha_pedido'] = ''
    pedidos['fecha_pedido'] = pd.to_datetime(pedidos['fecha_pedido'], errors='coerce')

    if 'monto' not in pagos.columns:
        pagos['monto'] = 0
    if 'metodo_pago' not in pagos.columns:
        pagos['metodo_pago'] = ''
    pagos['monto'] = pd.to_numeric(pagos['monto'], errors='coerce').fillna(0)
    pagos['metodo_pago'] = pagos['metodo_pago'].fillna('').astype(str)

    def filtrar_pedidos_por_rango(df_pedidos, desde=None, hasta=None):
        filtrado = df_pedidos.copy().dropna(subset=['fecha_pedido'])
        if desde is not None:
            filtrado = filtrado[filtrado['fecha_pedido'].dt.date >= desde]
        if hasta is not None:
            filtrado = filtrado[filtrado['fecha_pedido'].dt.date <= hasta]
        return filtrado

    def calcular_kpis_desde_pedidos(pedidos_base):
        ids = pedidos_base['id_pedido'].tolist()
        detalle_base = detalle[detalle['id_pedido'].isin(ids)].copy()
        pagos_base = pagos[pagos['id_pedido'].isin(ids)].copy()
        total_ventas_base = float(pagos_base['monto'].sum()) if not pagos_base.empty else float(detalle_base['subtotal'].sum())
        total_pedidos_base = int(len(pedidos_base))
        total_items_base = int(detalle_base['cantidad'].sum()) if not detalle_base.empty else 0
        ticket_promedio_base = (total_ventas_base / total_pedidos_base) if total_pedidos_base > 0 else 0
        return {
            'total_ventas': total_ventas_base,
            'total_pedidos': total_pedidos_base,
            'ticket_promedio': ticket_promedio_base,
            'total_items': total_items_base
        }

    def variacion_kpi(valor_actual, valor_anterior):
        if valor_anterior == 0:
            if valor_actual == 0:
                return {'trend': 'flat', 'texto': '0.0% vs periodo anterior'}
            return {'trend': 'up', 'texto': 'Nuevo vs periodo anterior'}

        delta_pct = ((valor_actual - valor_anterior) / abs(valor_anterior)) * 100
        trend = 'up' if delta_pct > 0 else 'down' if delta_pct < 0 else 'flat'
        return {'trend': trend, 'texto': f"{delta_pct:+.1f}% vs periodo anterior"}

    hoy = datetime.now().date()
    fecha_desde = None
    fecha_hasta = None

    if periodo == 'today':
        fecha_desde = hoy
        fecha_hasta = hoy
    elif periodo == 'week':
        fecha_desde = hoy - timedelta(days=6)
        fecha_hasta = hoy
    elif periodo == 'month':
        fecha_desde = hoy.replace(day=1)
        fecha_hasta = hoy
    elif periodo == 'range':
        if fecha_desde_raw:
            try:
                fecha_desde = datetime.strptime(fecha_desde_raw, '%Y-%m-%d').date()
            except ValueError:
                fecha_desde = None
        if fecha_hasta_raw:
            try:
                fecha_hasta = datetime.strptime(fecha_hasta_raw, '%Y-%m-%d').date()
            except ValueError:
                fecha_hasta = None
        if fecha_desde is not None and fecha_hasta is not None and fecha_desde > fecha_hasta:
            fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    pedidos_filtrados = filtrar_pedidos_por_rango(pedidos, fecha_desde, fecha_hasta)
    ids_pedidos = pedidos_filtrados['id_pedido'].tolist()
    detalle_filtrado = detalle[detalle['id_pedido'].isin(ids_pedidos)].copy()
    pagos_filtrados = pagos[pagos['id_pedido'].isin(ids_pedidos)].copy()

    ventas_por_producto_df = detalle_filtrado.groupby('id_producto', as_index=False)['cantidad'].sum()
    ventas_por_producto_df = pd.merge(
        ventas_por_producto_df,
        productos[['id_producto', 'nombre', 'precio']],
        on='id_producto',
        how='left'
    )
    ventas_por_producto_df['nombre'] = ventas_por_producto_df['nombre'].fillna('Producto sin nombre')
    ventas_por_producto_df['precio'] = pd.to_numeric(ventas_por_producto_df['precio'], errors='coerce').fillna(0)

    pedidos_fechas = pedidos_filtrados.copy()
    pedidos_fechas['mes'] = pedidos_fechas['fecha_pedido'].dt.to_period('M').astype(str)

    ventas_por_mes_df = detalle_filtrado.groupby('id_pedido', as_index=False)['subtotal'].sum()
    ventas_por_mes_df = pd.merge(ventas_por_mes_df, pedidos_fechas[['id_pedido', 'mes']], on='id_pedido', how='left')
    ventas_por_mes_df = ventas_por_mes_df.dropna(subset=['mes'])
    ventas_por_mes_df = ventas_por_mes_df.groupby('mes', as_index=False)['subtotal'].sum().sort_values('mes')

    metodos_pago_df = pagos_filtrados.groupby('metodo_pago', as_index=False)['monto'].sum()
    metodos_pago_df = metodos_pago_df[metodos_pago_df['metodo_pago'].str.strip() != '']

    kpis = calcular_kpis_desde_pedidos(pedidos_filtrados)

    top_productos_df = ventas_por_producto_df.sort_values('cantidad', ascending=False).head(5).copy()
    if 'precio' not in top_productos_df.columns:
        top_productos_df = pd.merge(
            top_productos_df,
            productos[['id_producto', 'precio']],
            on='id_producto',
            how='left'
        )
    top_productos_df['precio'] = pd.to_numeric(top_productos_df.get('precio', 0), errors='coerce').fillna(0)

    kpis_anterior = {'total_ventas': 0, 'total_pedidos': 0, 'ticket_promedio': 0, 'total_items': 0}
    comparacion_texto = 'Sin comparacion de periodo'
    if fecha_desde is not None and fecha_hasta is not None:
        dias_periodo = max(1, (fecha_hasta - fecha_desde).days + 1)
        anterior_hasta = fecha_desde - timedelta(days=1)
        anterior_desde = anterior_hasta - timedelta(days=dias_periodo - 1)
        pedidos_anterior = filtrar_pedidos_por_rango(pedidos, anterior_desde, anterior_hasta)
        kpis_anterior = calcular_kpis_desde_pedidos(pedidos_anterior)
        comparacion_texto = (
            f"Comparando contra {anterior_desde.strftime('%Y-%m-%d')} a "
            f"{anterior_hasta.strftime('%Y-%m-%d')}"
        )
    elif periodo == 'all':
        comparacion_texto = 'Selecciona un periodo para activar comparacion'

    kpi_variaciones = {
        'total_ventas': variacion_kpi(kpis['total_ventas'], kpis_anterior['total_ventas']),
        'total_pedidos': variacion_kpi(kpis['total_pedidos'], kpis_anterior['total_pedidos']),
        'ticket_promedio': variacion_kpi(kpis['ticket_promedio'], kpis_anterior['ticket_promedio']),
        'total_items': variacion_kpi(kpis['total_items'], kpis_anterior['total_items'])
    }

    rango_texto = 'Todo el historial'
    if fecha_desde is not None and fecha_hasta is not None:
        rango_texto = f"{fecha_desde.strftime('%Y-%m-%d')} a {fecha_hasta.strftime('%Y-%m-%d')}"
    elif fecha_desde is not None:
        rango_texto = f"Desde {fecha_desde.strftime('%Y-%m-%d')}"
    elif fecha_hasta is not None:
        rango_texto = f"Hasta {fecha_hasta.strftime('%Y-%m-%d')}"

    filtros = {
        'periodo': periodo if periodo in {'all', 'today', 'week', 'month', 'range'} else 'all',
        'fecha_desde': fecha_desde_raw,
        'fecha_hasta': fecha_hasta_raw
    }

    return {
        'ventas_producto': ventas_por_producto_df.to_dict(orient='records'),
        'ventas_mes': ventas_por_mes_df.to_dict(orient='records'),
        'metodos_pago': metodos_pago_df.to_dict(orient='records'),
        'top_productos': top_productos_df.to_dict(orient='records'),
        'kpis': kpis,
        'kpi_variaciones': kpi_variaciones,
        'filtros': filtros,
        'rango_texto': rango_texto,
        'comparacion_texto': comparacion_texto,
        'ventas_producto_df': ventas_por_producto_df,
        'ventas_mes_df': ventas_por_mes_df,
        'metodos_pago_df': metodos_pago_df,
        'top_productos_df': top_productos_df
    }


@app.route('/admin/charts/export_excel')
def admin_charts_export_excel():
    if session.get('rol') != 'admin':
        return "Acceso denegado"

    periodo = request.args.get('periodo', 'all').strip().lower()
    fecha_desde_raw = request.args.get('fecha_desde', '').strip()
    fecha_hasta_raw = request.args.get('fecha_hasta', '').strip()
    datos = obtener_datos_charts(periodo, fecha_desde_raw, fecha_hasta_raw)

    resumen_df = pd.DataFrame([
        {'metrica': 'Ventas totales', 'valor': formatear_cop(datos['kpis']['total_ventas'])},
        {'metrica': 'Pedidos', 'valor': datos['kpis']['total_pedidos']},
        {'metrica': 'Ticket promedio', 'valor': formatear_cop(datos['kpis']['ticket_promedio'])},
        {'metrica': 'Items vendidos', 'valor': datos['kpis']['total_items']},
        {'metrica': 'Rango aplicado', 'valor': datos['rango_texto']},
        {'metrica': 'Comparacion', 'valor': datos['comparacion_texto']}
    ])

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        resumen_df.to_excel(writer, sheet_name='Resumen', index=False)
        datos['ventas_producto_df'].to_excel(writer, sheet_name='Ventas por producto', index=False)
        datos['ventas_mes_df'].to_excel(writer, sheet_name='Ventas por mes', index=False)
        datos['metodos_pago_df'].to_excel(writer, sheet_name='Métodos de pago', index=False)
        datos['top_productos_df'].to_excel(writer, sheet_name='Top productos', index=False)

        wb = writer.book
        ws_producto = wb['Ventas por producto']
        ws_mes = wb['Ventas por mes']
        ws_metodos = wb['Métodos de pago']
        ws_top = wb['Top productos']
        formato_cop = '[>=1000]#,##0.00 "COP";0.00 "COP"'

        # Formatos numericos COP en columnas monetarias
        for row in range(2, ws_mes.max_row + 1):
            ws_mes.cell(row=row, column=2).number_format = formato_cop  # subtotal
        for row in range(2, ws_metodos.max_row + 1):
            ws_metodos.cell(row=row, column=2).number_format = formato_cop  # monto
        for row in range(2, ws_top.max_row + 1):
            ws_top.cell(row=row, column=4).number_format = formato_cop  # precio

        if ws_producto.max_row > 1:
            chart_prod = BarChart()
            chart_prod.title = 'Cantidad vendida por producto'
            chart_prod.y_axis.title = 'Cantidad'
            chart_prod.x_axis.title = 'Producto'
            data_prod = Reference(ws_producto, min_col=3, min_row=1, max_row=ws_producto.max_row)
            cats_prod = Reference(ws_producto, min_col=2, min_row=2, max_row=ws_producto.max_row)
            chart_prod.add_data(data_prod, titles_from_data=True)
            chart_prod.set_categories(cats_prod)
            chart_prod.height = 8
            chart_prod.width = 14
            if chart_prod.series:
                chart_prod.series[0].graphicalProperties.solidFill = "1882A9"
            ws_producto.add_chart(chart_prod, 'E2')

        if ws_mes.max_row > 1:
            chart_mes = LineChart()
            chart_mes.title = 'Ventas por mes'
            chart_mes.y_axis.title = 'Subtotal'
            chart_mes.x_axis.title = 'Mes'
            data_mes = Reference(ws_mes, min_col=2, min_row=1, max_row=ws_mes.max_row)
            cats_mes = Reference(ws_mes, min_col=1, min_row=2, max_row=ws_mes.max_row)
            chart_mes.add_data(data_mes, titles_from_data=True)
            chart_mes.set_categories(cats_mes)
            chart_mes.height = 8
            chart_mes.width = 14
            if chart_mes.series:
                chart_mes.series[0].graphicalProperties.line.solidFill = "0A2962"
            ws_mes.add_chart(chart_mes, 'D2')

        if ws_metodos.max_row > 1:
            chart_met = PieChart()
            chart_met.title = 'Distribucion por metodo de pago'
            data_met = Reference(ws_metodos, min_col=2, min_row=1, max_row=ws_metodos.max_row)
            cats_met = Reference(ws_metodos, min_col=1, min_row=2, max_row=ws_metodos.max_row)
            chart_met.add_data(data_met, titles_from_data=True)
            chart_met.set_categories(cats_met)
            chart_met.height = 8
            chart_met.width = 12
            if chart_met.series:
                colores = ["0A2962", "1882A9", "1ABC9C", "F1C40F", "E74C3C", "6C5CE7"]
                points = []
                num_points = max(0, ws_metodos.max_row - 1)
                for i in range(num_points):
                    pt = DataPoint(idx=i)
                    pt.graphicalProperties.solidFill = colores[i % len(colores)]
                    points.append(pt)
                chart_met.series[0].dPt = points
            ws_metodos.add_chart(chart_met, 'D2')

        if ws_top.max_row > 1:
            chart_top = BarChart()
            chart_top.title = 'Top productos vendidos'
            chart_top.y_axis.title = 'Cantidad'
            chart_top.x_axis.title = 'Producto'
            data_top = Reference(ws_top, min_col=3, min_row=1, max_row=ws_top.max_row)
            cats_top = Reference(ws_top, min_col=2, min_row=2, max_row=ws_top.max_row)
            chart_top.add_data(data_top, titles_from_data=True)
            chart_top.set_categories(cats_top)
            chart_top.height = 8
            chart_top.width = 14
            if chart_top.series:
                chart_top.series[0].graphicalProperties.solidFill = "0A2962"
            ws_top.add_chart(chart_top, 'F2')
    buffer.seek(0)

    nombre_archivo = f"graficas_informe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
#orden personalizada
@app.route('/orden-personalizada')
def orden_personalizada():
    return render_template('Usuarios/orden_personalizada/orden.html')

#sobre nosotros
@app.route('/sobre-nosotros')
def sobre_nosotros():
    return render_template('Usuarios/Informacion empresa/sobre_nosotros.html')

if __name__ == '__main__':
    app.run(debug=True, use_reloader=True)
    
